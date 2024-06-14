from django.db.models import F, Count, Q, Subquery, OuterRef
from django.db.models.functions import ExtractMonth, ExtractYear
from django.shortcuts import get_object_or_404
from collections import defaultdict

from django.utils import timezone
import openpyxl.styles
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from survei.models import DataSurvei, TipeSurvei
from survey.models import survey as FormulirElektronik
from users.models import Profile, Satker

import re
import os
import shutil
import codecs
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import json

from sidamas import pagination

from kegiatan import models
from kegiatan.api import serializers
from kegiatan.api import filters
from users.models import Satker

def get_kegiatan_satker_status(request):
    satker = request.user.profile.satker

    """
        Mapping :
        Satker Level       Status
        0 : BNNP        -> 2 : Pusat
        1 : BNNK        -> 1 : BNNP
        2 : Pusat       -> 2 : Pusat
    """

    status_map = {
        0: 1,
        1: 0,
        2: 2
    }

    status = status_map.get(satker.level, None)

    return status

def get_tanggal_kegiatan(self, tanggal_awal, tanggal_akhir=None):
    start = datetime.date.fromisoformat(f'{tanggal_awal}')
    start_date = start.strftime('%-d')
    start_month = start.strftime('%B')
    start_year = start.strftime('%Y')

    nama_bulan = {
        "January": "Januari",
        "February": "Februari",
        "March": "Maret",
        "April": "April",
        "May": "Mei",
        "June": "Juni",
        "July": "Juli",
        "August": "Agustus",
        "September": "September",
        "October": "Oktober",
        "November": "November",
        "December": "Desember"
    }

    if tanggal_akhir:
        end = datetime.date.fromisoformat(f'{tanggal_akhir}')
        end_date = end.strftime('%-d')
        end_month = end.strftime('%B')
        end_year = end.strftime('%Y')

        if start_month == end_month and start_year == end_year:
            return f"{start_date} - {end_date} {nama_bulan[start_month]} {start_year}"
        elif start_year != end_year:
            return f"{start.strftime('%-d')} {nama_bulan[start_month]} {start_year} - {end.strftime('%-d')} {nama_bulan[end_month]} {end_year}"
        else:
            return f"{start.strftime('%-d')} {nama_bulan[start_month]} - {end.strftime('%-d')} {nama_bulan[end_month]} {end_year}"
    else:
        return f"{start.strftime('%-d')} {nama_bulan[start_month]} {start_year}"

# ======= COUNT KEGIGATAN =======
class PSM_KegiatanCountView(APIView):

    def get(self, request):
        current_year = timezone.now().year
        satker_id = request.query_params.get('satker', None)

        def get_counts(queryset, satker_id):
            if satker_id:
                queryset = queryset.filter(satker_id=satker_id)
            counts = queryset.annotate(
                year=ExtractYear('created_at'),
                month=ExtractMonth('created_at')
            ).values('year', 'month').annotate(
                total=Count('id')
            ).order_by('year', 'month')

            result = {}
            for item in counts:
                year = item['year']
                month = item['month']
                if year not in result:
                    result[year] = {m: 0 for m in range(1, 13)}
                    result[year]['total'] = 0
                result[year][month] = item['total']
                result[year]['total'] += item['total']

            # Ensure all months are present with a value of 0 if no data is found
            for year in result:
                for month in range(1, 13):
                    if month not in result[year]:
                        result[year][month] = 0

            # Ensure the current year is present in the result
            if current_year not in result:
                result[current_year] = {m: 0 for m in range(1, 13)}
                result[current_year]['total'] = 0

            return result

        data = {
            'rakernis': get_counts(models.PSM_RAKERNIS.objects.all(), satker_id),
            'binaan_teknis': get_counts(models.PSM_BINAAN_TEKNIS.objects.all(), satker_id),
            'asistensi': get_counts(models.PSM_ASISTENSI.objects.all(), satker_id),
            'rakor_pemetaan': get_counts(models.PSM_RAKOR_PEMETAAN.objects.all(), satker_id),
            'audiensi': get_counts(models.PSM_AUDIENSI.objects.all(), satker_id),
            'konsolidasi_kebijakan': get_counts(models.PSM_KONSOLIDASI_KEBIJAKAN.objects.all(), satker_id),
            'workshop_penggiat': get_counts(models.PSM_WORKSHOP_PENGGIAT.objects.all(), satker_id),
            'bimtek_p4gn': get_counts(models.PSM_BIMTEK_P4GN.objects.all(), satker_id),
            'sinkronasi_kebijakan': get_counts(models.PSM_SINKRONISASI_KEBIJAKAN.objects.all(), satker_id),
            'workshop_tematik': get_counts(models.PSM_WORKSHOP_TEMATIK.objects.all(), satker_id),
            'asistensi': get_counts(models.PSM_ASISTENSI.objects.all(), satker_id),
            'tes_urine_deteksi_dini': get_counts(models.PSM_TES_URINE_DETEKSI_DINI.objects.all(), satker_id),
            'monev_supervisi_kegiatan_kotan': get_counts(models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all(), satker_id),
            'pengumpulan_data_ikotan': get_counts(models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all(), satker_id),
            'dukungan_data_stakeholder': get_counts(models.PSM_DUKUNGAN_STAKEHOLDER.objects.all(), satker_id),
            'kegiatan_lainnya': get_counts(models.PSM_KEGIATAN_LAINNYA.objects.all(), satker_id),
        }

        return Response({
            'status': True,
            'data': data
        }, status=status.HTTP_200_OK)

# ======= PSM RAKERNIS API =======
class PSM_SurveiCountView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        satker_id = request.query_params.get('satker', None)
        
        def get_counts_formulir(queryset):
            counts = queryset.annotate(
                year=ExtractYear('created'),
                month=ExtractMonth('created')
            ).values('year', 'month').annotate(
                total=Count('id')
            ).order_by('year', 'month')

            result = {}
            for item in counts:
                year = item['year']
                month = item['month']
                if year not in result:
                    result[year] = {m: 0 for m in range(1, 13)}
                    result[year]['total'] = 0
                result[year][month] = item['total']
                result[year]['total'] += item['total']

            # Ensure all months are present with a value of 0 if no data is found
            for year in result:
                for month in range(1, 13):
                    if month not in result[year]:
                        result[year][month] = 0

            # Ensure the current year is present in the result
            if current_year not in result:
                result[current_year] = {m: 0 for m in range(1, 13)}
                result[current_year]['total'] = 0

            return result

        def get_counts(queryset, satker_id):
            counts = queryset.annotate(
                year=ExtractYear('created_at'),
                month=ExtractMonth('created_at')
            ).values('year', 'month').annotate(
                total=Count('id')
            ).order_by('year', 'month')

            result = {}
            for item in counts:
                year = item['year']
                month = item['month']
                if year not in result:
                    result[year] = {m: 0 for m in range(1, 13)}
                    result[year]['total'] = 0
                result[year][month] = item['total']
                result[year]['total'] += item['total']

            # Ensure all months are present with a value of 0 if no data is found
            for year in result:
                for month in range(1, 13):
                    if month not in result[year]:
                        result[year][month] = 0

            # Ensure the current year is present in the result
            if current_year not in result:
                result[current_year] = {m: 0 for m in range(1, 13)}
                result[current_year]['total'] = 0

            return result
        
        tipe_survei_skm_tes_urine = TipeSurvei.objects.get(nama="SKM Tes Urine")
        data_survei_skm_tes_urine = DataSurvei.objects.filter(tipe=tipe_survei_skm_tes_urine.id)
        data_formulir_elektronik = FormulirElektronik.objects.filter(pemilik="1")
        
        data = {
            'formulir_elektronik': get_counts_formulir(data_formulir_elektronik),
            'skm_tes_urine': get_counts(data_survei_skm_tes_urine, satker_id)
        }

        return Response({
            'status': True,
            'data': data
        }, status=status.HTTP_200_OK)

# ======= PSM RAKERNIS API =======

class PSM_RAKERNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_RAKERNIS.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_RAKERNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_RAKERNIS_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_RAKERNIS.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_RAKERNIS_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_RAKERNIS.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_RAKERNIS.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI RAPAT KERJA TEKNIS {tahun}'
        base_path = 'media/kegiatan/psm/rakernis/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN RAKERNIS PEMBERDAYAAN MASYARAKAT")
            sheet.merge_cells('A1:K1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:K2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'TANGGAL', 'SATUAN KERJA YANG DIUNDANG', 'DESKRIPSI', 'KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('D4:D5')
            sheet.merge_cells('E4:E5')
            sheet.merge_cells('F4:F5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('K4:K5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['I4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'ANGGARAN'
            sheet.merge_cells('I4:J4')

            # Menambahkan subheader untuk 'ANGGARAN'
            subheaders = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']
            for col_num, subheader in enumerate(subheaders, start=9):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Mengatur tinggi baris untuk header dan subheader
            sheet.row_dimensions[1].height = 30
            sheet.row_dimensions[2].height = 30

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_RAKERNIS.objects.values(
                'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by(
                '-satker_id'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'satker_target': item['satker_target'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker_target__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            current_row = 6
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']
                data_row = [
                    group_count,
                    row['nama_satker'],
                    formated_date,
                    row['nama_satker_target'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                # Tambahkan data_row ke sheet
                for item, value in enumerate(data_row[:11], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = Alignment(vertical='center')
                current_row += 1
                numering += 1

            # Pastikan untuk menggabungkan sel terakhir jika diperlukan
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_RAKERNIS_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_RAKERNIS.objects.get(id=request.data['id'])
        except models.PSM_RAKERNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_RAKERNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_RAKERNIS.objects.get(id=id)
        except models.PSM_RAKERNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_RAKERNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM RAKERNIS"

# ======= PSM BINAAN TEKNIS API =======
class PSM_BINAAN_TEKNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_BINAAN_TEKNIS.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_BINAAN_TEKNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_BINAAN_TEKNIS_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_BINAAN_TEKNIS.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_BINAAN_TEKNIS_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_BINAAN_TEKNIS.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_BINAAN_TEKNIS.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_BINAAN_TEKNIS_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_BINAAN_TEKNIS.objects.get(id=request.data['id'])
        except models.PSM_BINAAN_TEKNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_BINAAN_TEKNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_BINAAN_TEKNIS.objects.get(id=id)
        except models.PSM_BINAAN_TEKNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_BINAAN_TEKNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI PEMBINAAN TEKNIS {tahun}'
        base_path = 'media/kegiatan/psm/bintek/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="REKAPITULASI PEMBINAAN TEKNIS BNN, BNNP KE BNNK ")
            sheet.merge_cells('A1:M1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            tahun = "2024"  # Ganti dengan tahun yang sesuai
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:M2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'NO', 'TANGGAL', 'SATKER YANG DIBINTEK', 'DESKRIPSI HASIL ', 'HAMBATAN/ KENDALA  ', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('D4:D5')
            sheet.merge_cells('E4:E5')
            sheet.merge_cells('F4:F5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('M4:M5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['K4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'ANGGARAN'
            sheet.merge_cells('K4:L4')

            # Menambahkan subheader untuk 'ANGGARAN'
            subheaders = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']
            for col_num, subheader in enumerate(subheaders, start=11):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Mengatur tinggi baris untuk header dan subheader
            sheet.row_dimensions[1].height = 30
            sheet.row_dimensions[2].height = 30

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_BINAAN_TEKNIS.objects.values(
                'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by(
                '-satker__id'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'satker_target': item['satker_target'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker_target__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    row['nama_satker_target'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:13], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def get_view_name(self):
        return "PSM BINTEK"

# ======= PSM ASISTENSI API =======
class PSM_ASISTENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_ASISTENSI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__order').distinct('satker__id')
    serializer_class = serializers.PSM_ASISTENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_ASISTENSI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()

        if satker_level == 1:
            # bnnk
            queryset = queryset.filter(satker__id=satker)
        elif satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        elif satker_level == 2:
            # pusat
            queryset = queryset.filter(status=2)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_ASISTENSI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_ASISTENSI_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_ASISTENSI.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
            'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_ASISTENSI.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN ASISTENSI KOTAN {tahun}'
        base_path = 'media/kegiatan/psm/asistensi/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN ASISTENSI KOTAN")
            sheet.merge_cells('A1:N1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:N2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','', 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')
            sheet.merge_cells('N4:N5')
            sheet.merge_cells('O4:O5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['L4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('L4:M4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'STAKEHOLDER/ PENDAMPING YANG DIUNDANG']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=12):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_ASISTENSI.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
                'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    row['jumlah_peserta'],
                    row['stakeholder'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:14], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_ASISTENSI_CREATE_UPDATE_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_ASISTENSI.objects.get(id=request.data['id'])
        except models.PSM_ASISTENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_ASISTENSI_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_ASISTENSI.objects.get(id=id)
        except models.PSM_ASISTENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_ASISTENSI_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM ASISTENSI"

# ======= PSM SINKRONISASI KEBIJAKAN API =======
class PSM_SINKRONISASI_KEBIJAKAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_SINKRONISASI_KEBIJAKAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__order').distinct('satker__id')
    serializer_class = serializers.PSM_SINKRONISASI_KEBIJAKAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_SINKRONISASI_KEBIJAKAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()

        if satker_level == 1:
            # bnnk
            queryset = queryset.filter(satker__id=satker)
        elif satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        elif satker_level == 2:
            # pusat
            queryset = queryset.filter(status=2)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_SINKRONISASI_KEBIJAKAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_SINKRONISASI_KEBIJAKAN_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_SINKRONISASI_KEBIJAKAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
            'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_SINKRONISASI_KEBIJAKAN.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN RAPAT SINKRONISASI KEBIJAKAN STAKEHOLDERS DALAM RANGKA KOTAN {tahun}'
        base_path = 'media/kegiatan/psm/sinkronisasi_kebijakan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'
        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN RAPAT SINKRONISASI KEBIJAKAN STAKEHOLDERS DALAM RANGKA KOTAN")
            sheet.merge_cells('A1:N1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:N2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','', 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')
            sheet.merge_cells('N4:N5')
            sheet.merge_cells('O4:O5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['L4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('L4:M4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'STAKEHOLDER/ PENDAMPING YANG DIUNDANG']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=12):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_SINKRONISASI_KEBIJAKAN.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
                'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url', 'dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    row['jumlah_peserta'],
                    row['stakeholder'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:14], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_SINKRONISASI_KEBIJAKAN_CREATE_UPDATE_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_SINKRONISASI_KEBIJAKAN.objects.get(id=request.data['id'])
        except models.PSM_SINKRONISASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_SINKRONISASI_KEBIJAKAN_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_SINKRONISASI_KEBIJAKAN.objects.get(id=id)
        except models.PSM_SINKRONISASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_SINKRONISASI_KEBIJAKAN_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM SINKRONISASI KEBIJAKAN"

# ======= PSM WORKSHOP TEMATIK API =======
class PSM_WORKSHOP_TEMATIK_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_WORKSHOP_TEMATIK.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__order').distinct('satker__id')
    serializer_class = serializers.PSM_WORKSHOP_TEMATIK_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_WORKSHOP_TEMATIK_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()

        if satker_level == 1:
            # bnnk
            queryset = queryset.filter(satker__id=satker)
        elif satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        elif satker_level == 2:
            # pusat
            queryset = queryset.filter(status=2)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_WORKSHOP_TEMATIK.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_WORKSHOP_TEMATIK_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_WORKSHOP_TEMATIK.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
            'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_WORKSHOP_TEMATIK.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI KEGIATAN WORKSHOP TEMATIK {tahun}'
        base_path = 'media/kegiatan/psm/sinkronisasi_kebijakan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="REKAPITULASI KEGIATAN WORKSHOP TEMATIK")
            sheet.merge_cells('A1:N1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:N2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','', 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')
            sheet.merge_cells('N4:N5')
            sheet.merge_cells('O4:O5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['L4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('L4:M4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'STAKEHOLDER/ PENDAMPING YANG DIUNDANG']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=12):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_WORKSHOP_TEMATIK.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
                'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    row['jumlah_peserta'],
                    row['stakeholder'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:14], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_WORKSHOP_TEMATIK_CREATE_UPDATE_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_WORKSHOP_TEMATIK.objects.get(id=request.data['id'])
        except models.PSM_WORKSHOP_TEMATIK.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_WORKSHOP_TEMATIK_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_WORKSHOP_TEMATIK.objects.get(id=id)
        except models.PSM_WORKSHOP_TEMATIK.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_WORKSHOP_TEMATIK_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM WORKSHOP TEMATIK"

# ======= PSM TES URIN DETEKSI DINI API =======
class PSM_TES_URINE_DETEKSI_DINI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_TES_URINE_DETEKSI_DINI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_TES_URINE_DETEKSI_DINI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_TES_URINE_DETEKSI_DINI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran' , 'drive_url', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran' , 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'])
    def delete_kegiatan_peserta(self, request, pk=None):
        # hapus data kegiatan sama seluruh peserta yang terkoneksi
        try:
            deleted_peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(pk=pk).delete()
            return Response({
                'message': 'Berhasil menghapus data',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}',
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)

        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}

            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN TES URINE DALAM RANGKA DETEKSI DINI {tahun}'
        base_path = 'media/kegiatan/psm/tes_urine/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN TES URINE DALAM RANGKA DETEKSI DINI")
            sheet.merge_cells('A1:O1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:O2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '','LINGKUNGAN', '','','JUMLAH PESERTA', 'HASIL TES URIN','','TINDAK LANJUT', 'ANGGARAN','', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A6')
            sheet.merge_cells('B4:B6')
            sheet.merge_cells('C4:C6')

            sheet.merge_cells('I4:I6')
            sheet.merge_cells('L4:L6')
            sheet.merge_cells('O4:O6')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['F4'].value = 'LINGKUNGAN'
            sheet['J4'].value = 'HASIL TES URIN'
            sheet['M4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:E4')
            sheet.merge_cells('F4:H4')
            sheet.merge_cells('J4:K4')
            sheet.merge_cells('M4:N4')

            subheaders1 = ['NO', 'TANGGAL']
            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet.merge_cells('D5:D6')
            sheet.merge_cells('E5:E6')

            subheaders2 = ['NAMA LINGKUNGAN', 'PESERTA']
            for col_num, subheader in enumerate(subheaders2, start=6):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet.merge_cells('G5:H5')
            sheet.merge_cells('F5:F6')

            subheaders2_1 = ['PRIA', 'WANITA']
            for col_num, subheader in enumerate(subheaders2_1, start=7):
                cell = sheet.cell(row=6, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders3 = ['POSITIF', 'NEGATIF']
            for col_num, subheader in enumerate(subheaders3, start=10):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet.merge_cells('J5:J6')
            sheet.merge_cells('K5:K6')

            subheaders4 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']
            for col_num, subheader in enumerate(subheaders4, start=13):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet.merge_cells('M5:M6')
            sheet.merge_cells('N5:N6')

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_TES_URINE_DETEKSI_DINI.objects.values(
                'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran' , 'drive_url', 'dokumentasi', 'status', 'satker_id__level'
            ).order_by(
                '-satker_id'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'satker_target': item['satker_target'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker_target__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'nama_lingkungan': item['nama_lingkungan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'status': item['status'],
                    'dokumentasi': item['dokumentasi'],
                    'satker_level': item['satker_id__level'],
                    'peserta': self.get_data_peserta(item['id']),
                }
                serialized_data.append(serialized_item)

            jumlah_kegiatan_per_satker = {}

            for row in serialized_data:
                nama_satker = row['nama_satker']
                if nama_satker not in jumlah_kegiatan_per_satker:
                    jumlah_kegiatan_per_satker[nama_satker] = 0
                jumlah_kegiatan_per_satker[nama_satker] += 1

            current_row = 7
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                # Menghitung jumlah data dengan jenis kelamin "L"
                jumlah_laki_laki = sum(1 for item in row['peserta'] if item["jenis_kelamin"] == "Laki-Laki") or 0
                # Menghitung jumlah data dengan jenis kelamin "P"
                jumlah_perempuan = sum(1 for item in row['peserta'] if item["jenis_kelamin"] == "Perempuan") or 0
                # Menghitung jumlah peserta
                jumlah_peserta = (jumlah_laki_laki + jumlah_perempuan)

                # Jumlah positif
                positif = sum(1 for item in row['peserta'] if item["hasil_test"] == "Positif") or 0
                # Jumlah negatif
                negatif = sum(1 for item in row['peserta'] if item["hasil_test"] == "Negatif") or 0

                data_row = [
                    group_count,
                    row['nama_satker'],
                    jumlah_kegiatan_per_satker[row['nama_satker']],
                    numering,
                    formated_date,
                    row['nama_lingkungan'],
                    jumlah_laki_laki,
                    jumlah_perempuan,
                    jumlah_peserta,
                    positif,
                    negatif,
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:18], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Detail Peserta
            sheet_peserta = workbook.create_sheet('Detail Peserta')

            cell = sheet_peserta.cell(row=1, column=1, value="DETAIL PESERTA TES URIN DETEKSI DINI")
            sheet_peserta.merge_cells('A1:J1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            cell = sheet_peserta.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet_peserta.merge_cells('A2:J2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PESERTA', '', '','', '', '', '']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet_peserta.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet_peserta.merge_cells('A4:A5')
            sheet_peserta.merge_cells('B4:B5')
            sheet_peserta.merge_cells('C4:C5')

            sheet_peserta['D4'].value = 'PESERTA'
            sheet_peserta.merge_cells('D4:J4')

            subheaders = ['NO', 'NAMA PESERTA', 'JENIS KELAMIN', 'HASIL TES', 'ALAMAT', 'PARAMETER', 'KETERANGAN ISI PARAMETER']
            for col_num, subheader in enumerate(subheaders, start=4):
                cell = sheet_peserta.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )


            # Inisialisasi variabel
            current_group_peserta = None
            current_row_peserta = 6
            current_parent = None
            group_count_peserta = 0
            numbering_peserta = None
            
            jumlah_kegiatan_per_satker = {}

            for row in serialized_data:
                nama_satker = row['nama_satker']
                if nama_satker not in jumlah_kegiatan_per_satker:
                    jumlah_kegiatan_per_satker[nama_satker] = 0
                jumlah_kegiatan_per_satker[nama_satker] += 1
                

            for row in serialized_data:
                if row['nama_satker'] != current_group_peserta:
                    if current_group_peserta is not None:
                        # Merge sebelumnya berdasarkan jumlah peserta sebelumnya
                        end_merge_row = current_row_peserta - 1
                        sheet_peserta.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet_peserta.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet_peserta.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group_peserta = row['nama_satker']
                    start_merge_row = current_row_peserta
                    group_count_peserta += 1

                for peserta in row['peserta']:
                    if peserta['parent_id'] != current_parent:
                        current_parent = peserta['parent_id']
                        numbering_peserta = 1

                    # Isi data peserta
                    data_row = [
                        group_count_peserta,
                        row['nama_satker'],
                        jumlah_kegiatan_per_satker[row['nama_satker']],
                        numbering_peserta,
                        peserta['nama_peserta'],
                        peserta['jenis_kelamin'],
                        peserta['hasil_test'],
                        peserta['alamat'],
                        peserta['isi_parameter'],
                        peserta['keterangan_isi_parameter']
                    ]

                    for item, value in enumerate(data_row[:10], start=1):
                        cell = sheet_peserta.cell(row=current_row_peserta, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    current_row_peserta += 1
                    numbering_peserta += 1

            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group_peserta is not None:
                end_merge_row = current_row_peserta - 1
                sheet_peserta.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet_peserta.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet_peserta.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            for column_cells in sheet_peserta.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet_peserta.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:
                gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                status_mapping = {'P': 'Positif', 'L': 'Negatif'}

                jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                hasil_test = status_mapping.get(entry['status'], 'Negatif')

                parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=inserted_id)

                peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                    jenis_kelamin=jenis_kelamin,
                    hasil_test=hasil_test,
                    alamat=entry['alamat'],
                    isi_parameter=entry['isi_parameter'],
                    keterangan_isi_parameter=entry['keterangan_isi_parameter']
                )

    def get_view_name(self):
        return "PSM TEST URINE"

    def get_data_peserta(self, parent):
        data = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'parent_id', 'jenis_kelamin', 'hasil_test', 'alamat', 'isi_parameter', 'keterangan_isi_parameter'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'parent_id': item['parent_id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'hasil_test': item['hasil_test'],
                'alamat': item['alamat'],
                'isi_parameter': item['isi_parameter'],
                'keterangan_isi_parameter': item['keterangan_isi_parameter']
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def process_peserta_array(array_json, current_record_id):
        if array_json is not None:
            data = json.loads(array_json)
            gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
            status_mapping = {'P': 'Positif', 'L': 'Negatif'}
            parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=current_record_id)

            for entry in data:
                try:
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat'],
                        isi_parameter=entry['isi_parameter']

                    )
                except KeyError:
                    # Handle missing keys in entry
                    pass  # Or log the error

    def get_flat_values(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran' , 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        return serialized_data

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_TES_URINE_DETEKSI_DINI_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # helper
    def count_participants(self, data, gender):
        count = 0
        for row in data:
            if row['jenis_kelamin'] == gender:
                count += 1
        return count

    def format_addresses(self, objects):
        print(objects)
        result = ""
        for obj in objects:
            result += f"{obj['nama_peserta']} - {obj['alamat']}, "
        # Remove the trailing comma and space
        result = result.rstrip(", ")
        return result

    def count_hasil_test(self, data):
        count_positif = 0
        count_negatif = 0

        for row in data:
            if row['hasil_test'] == "Positif":
                count_positif += 1
            elif row['hasil_test'] == "Negatif":
                count_negatif += 1

        return "Positif - {} Negatif - {}".format(count_positif, count_negatif)

    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return serializers.PSM_TES_URINE_DETEKSI_DINI_CREATE_UPDATE_Serializer
        return super().get_serializer_class()

class PSM_TES_URNIE_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_TES_URINE_DETEKSI_DINI.objects.all()
    serializer_class = serializers.PSM_TES_URINE_DETEKSI_DINI_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_update(self, serializer):
        current_record_id = serializer.instance.pk

        serializer.save()

        models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent_id=current_record_id).delete()

        peserta_array_lama_json = self.request.data.get('peserta_lamaArray', None)
        if peserta_array_lama_json:
            data = json.loads(peserta_array_lama_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                    status_mapping = {'P': 'Positif', 'L': 'Negatif'}

                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=current_record_id)

                    peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat'],
                        isi_parameter=entry['isi_parameter'],
                        keterangan_isi_parameter=entry['keterangan_isi_parameter']
                    )

        peserta_array_json = self.request.data.get('pesertaArray', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                    status_mapping = {'P': 'Positif', 'L': 'Negatif'}

                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=current_record_id)

                    peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat'],
                        isi_parameter=entry['isi_parameter'],
                        keterangan_isi_parameter=entry['keterangan_isi_parameter']
                    )

        serializer.save(updated_by=self.request.user)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran' , 'drive_url',  'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    def get_data_peserta(self, parent):
        data = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'hasil_test', 'alamat', 'isi_parameter', 'keterangan_isi_parameter'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'hasil_test': item['hasil_test'],
                'alamat': item['alamat'],
                'isi_parameter': item['isi_parameter'],
                'keterangan_isi_parameter': item['keterangan_isi_parameter']
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False)
    def get_detail_data_detail_all(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut',  'anggaran', 'penyerapan_anggaran' , 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

# ======= PSM MONITORING DAN EVALUASI SUPERVISI KEGIATAN KOTAN API =======
class PSM_MONITORING_DAN_EVALUASI_SUPERVISI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'])
    def delete_kegiatan_peserta(self, request, pk=None):
        # hapus data kegiatan sama seluruh peserta yang terkoneksi
        try:
            deleted_peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).delete()
            return Response({
                'message': 'Berhasil menghapus data',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}',
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)

        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}

            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'MONEV SUPERVISI KEGIATAN KOTAN {tahun}'
        base_path = 'media/kegiatan/psm/monev_supervisi_kegiatan_kotan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        # load template
        template_base_path = 'media/kegiatan/psm/monev_supervisi_kegiatan_kotan/template'
        template_file_path = f'{template_base_path}/template.xlsx'

        workbook = openpyxl.load_workbook(template_file_path)
        sheet = workbook.get_sheet_by_name('Data Kegiatan')
        sheet_peserta = workbook.get_sheet_by_name('Detail Peserta')
        workbook.active = sheet

        try:
            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:O2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'nama_lingkungan', 'status_indeks', 'nilai_ikp', 'status_ikp', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by(
                '-satker_id'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'nama_lingkungan': item['nama_lingkungan'],
                    'status_indeks': item['status_indeks'],
                    'status_ikp': item['status_ikp'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'nilai_ikp': item['nilai_ikp'],
                    'deskripsi_hasil': item['deskripsi_hasil'],
                    'simpulan': item['simpulan'],
                    'status': item['status'],
                    'dokumentasi': item['dokumentasi'],
                    'satker_level': satker_level,
                    'peserta': self.get_data_peserta(item['id']),
                }
                serialized_data.append(serialized_item)

            current_row = 7
            numering = 0
            group_count = 0
            current_group = None

            jumlah_kegiatan_per_satker = {}

            for row in serialized_data:
                nama_satker = row['nama_satker']
                if nama_satker not in jumlah_kegiatan_per_satker:
                    jumlah_kegiatan_per_satker[nama_satker] = 0
                jumlah_kegiatan_per_satker[nama_satker] += 1

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                # jumlah_kegiatan = count_kegiatan[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                # Menghitung jumlah data dengan jenis kelamin "L & P"
                jumlah_laki_laki = sum(1 for item in row['peserta'] if item["jenis_kelamin"] == "Laki-Laki") or 0
                jumlah_perempuan = sum(1 for item in row['peserta'] if item["jenis_kelamin"] == "Perempuan") or 0
                # Menghitung jumlah peserta
                jumlah_peserta = (jumlah_laki_laki + jumlah_perempuan)

                status_indeksMapping = {
                    1: "Sangat Tidak Mandiri",
                    2: "Tidak Mandiri",
                    3: "Mandiri",
                    4: "Sangat Mandiri"
                }
                status_indeks = status_indeksMapping.get(row['status_indeks'], "Nilai tidak valid")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    jumlah_kegiatan_per_satker[row['nama_satker']],
                    numering,
                    formated_date,
                    row['nama_lingkungan'],
                    jumlah_laki_laki,
                    jumlah_perempuan,
                    jumlah_peserta,
                    status_indeks,
                    row['nilai_ikp'],
                    row['status_ikp'],
                    row['deskripsi_hasil'],
                    row['simpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:18], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Detail Peserta
            cell = sheet_peserta.cell(row=2, column=1, value=f"TAHUN {tahun}")
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            current_group_peserta = None
            current_row_peserta = 6
            current_parent = None
            group_count_peserta = 0
            numbering_peserta = 1

            for row in serialized_data:
                if row['nama_satker'] != current_group_peserta:
                    if current_group_peserta is not None:
                        # Merge sebelumnya berdasarkan jumlah peserta sebelumnya
                        end_merge_row = current_row_peserta - 1
                        sheet_peserta.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet_peserta.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet_peserta.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group_peserta = row['nama_satker']
                    start_merge_row = current_row_peserta
                    group_count_peserta += 1

                for peserta in row['peserta']:
                    if peserta['parent_id'] != current_parent:
                        current_parent = peserta['parent_id']
                        numbering_peserta = 1

                    # Isi data peserta
                    data_row = [
                        group_count_peserta,
                        row['nama_satker'],
                        jumlah_kegiatan_per_satker[row['nama_satker']],
                        numbering_peserta,
                        # peserta['parent_id'],
                        peserta['nama_peserta'],
                        peserta['jabatan'],
                        peserta['jenis_kelamin'],
                    ]

                    for item, value in enumerate(data_row[:7], start=1):
                        cell = sheet_peserta.cell(row=current_row_peserta, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    current_row_peserta += 1
                    numbering_peserta += 1

            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group_peserta is not None:
                end_merge_row = current_row_peserta - 1
                sheet_peserta.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet_peserta.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet_peserta.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:
                gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                status_mapping = {'P': 'Positif', 'L': 'Negatif'}

                jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                hasil_test = status_mapping.get(entry['status'], 'Negatif')

                parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=inserted_id)

                peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                    jenis_kelamin=jenis_kelamin,
                    hasil_test=hasil_test,
                    alamat=entry['alamat']
                )

    def get_view_name(self):
        return "PSM MONITORING DAN EVALUASI SUPERVISI KEGIATAN KOTAN"

    def get_data_peserta(self, parent):
        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'jabatan', 'parent_id'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'parent_id': item['parent_id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'jabatan': item['jabatan'],
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def process_peserta_array(array_json, current_record_id):
        if array_json is not None:
            data = json.loads(array_json)
            gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
            status_mapping = {'P': 'Positif', 'L': 'Negatif'}
            parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=current_record_id)

            for entry in data:
                try:
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat']
                    )
                except KeyError:
                    # Handle missing keys in entry
                    pass  # Or log the error

    def get_flat_values(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        return serialized_data

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # helper
    def count_participants(self, data, gender):
        count = 0
        for row in data:
            if row['jenis_kelamin'] == gender:
                count += 1
        return count

    def format_addresses(self, objects):
        print(objects)
        result = ""
        for obj in objects:
            result += f"{obj['nama_peserta']} - {obj['alamat']}, "
        # Remove the trailing comma and space
        result = result.rstrip(", ")
        return result

    def count_hasil_test(self, data):
        count_positif = 0
        count_negatif = 0

        for row in data:
            if row['hasil_test'] == "Positif":
                count_positif += 1
            elif row['hasil_test'] == "Negatif":
                count_negatif += 1

        return "Positif - {} Negatif - {}".format(count_positif, count_negatif)

    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CREATE_UPDATE_Serializer
        return super().get_serializer_class()

class PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all()
    serializer_class = serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:
                gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}

                jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')

                parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=inserted_id)

                peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                    jenis_kelamin=jenis_kelamin,
                    jabatan=entry['jabatan']
                )

    def perform_update(self, serializer):
        current_record_id = serializer.instance.pk

        serializer.save()

        models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent_id=current_record_id).delete()

        peserta_array_lama_json = self.request.data.get('peserta_lamaArray', None)
        if peserta_array_lama_json:
            data = json.loads(peserta_array_lama_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}

                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')

                    parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=current_record_id)

                    peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        jabatan=entry['jabatan']
                    )

        peserta_array_json = self.request.data.get('pesertaArray', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}

                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')

                    parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=current_record_id)

                    peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        jabatan=entry['jabatan']
                    )

        serializer.save(updated_by=self.request.user)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'status_indeks', 'nilai_ikp', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'status_indeks': item['status_indeks'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'nilai_ikp': item['nilai_ikp'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    def get_data_peserta(self, parent):
        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'jabatan'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'jabatan': item['jabatan']
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def delete(self, request, *args, **kwargs):
        try:
            pk = self.kwargs.get('pk')
            deleted_peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).delete()
            return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}'}, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM PENGUMPULAN DATA IKOTAN API =======
class PSM_PENGUMPULAN_DATA_IKOTAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_PENGUMPULAN_DATA_IKOTAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_PENGUMPULAN_DATA_IKOTAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)

        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}

            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def get_view_name(self):
        return "PSM PENGUMPULAN DATA IKOTAN"

    def get_data_peserta(self, parent):
        data = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta']
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def get_flat_values(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        return serialized_data

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_PENGUMPULAN_DATA_IKOTAN_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # helper
    def count_participants(self, data, gender):
        count = 0
        for row in data:
            if row['jenis_kelamin'] == gender:
                count += 1
        return count

    def format_addresses(self, objects):
        print(objects)
        result = ""
        for obj in objects:
            result += f"{obj['nama_peserta']} - {obj['alamat']}, "
        # Remove the trailing comma and space
        result = result.rstrip(", ")
        return result

    def count_hasil_test(self, data):
        count_positif = 0
        count_negatif = 0

        for row in data:
            if row['hasil_test'] == "Positif":
                count_positif += 1
            elif row['hasil_test'] == "Negatif":
                count_negatif += 1

        return "Positif - {} Negatif - {}".format(count_positif, count_negatif)

    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CREATE_UPDATE_Serializer
        return super().get_serializer_class()

    def get_data_peserta(self, parent):
        data = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN PENGUMPULAN DATA IKOTAN {tahun}'
        base_path = 'media/kegiatan/psm/pengumpulan_data_ikotan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        # load template
        template_base_path = 'media/kegiatan/psm/pengumpulan_data_ikotan/template'
        template_file_path = f'{template_base_path}/template.xlsx'

        workbook = openpyxl.load_workbook(template_file_path)
        sheet = workbook.get_sheet_by_name('Data Kegiatan')
        # sheet_peserta = workbook.get_sheet_by_name('Detail Peserta')
        workbook.active = sheet

        try:
            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:L2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'satker_id__level','observasi__id','observasi__nama_unit', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by(
                'satker__nama_satker'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'observasi_id': item['observasi__id'],
                    'nama_observasi': item['observasi__nama_unit'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi_hasil': item['deskripsi_hasil'],
                    'simpulan': item['simpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'status': item['status'],
                    'dokumentasi': item['dokumentasi'],
                    'satker_level': item['satker_id__level'],
                    'peserta': self.get_data_peserta(item['id']),
                }
                serialized_data.append(serialized_item)

            current_row = 7
            numering = 0
            group_count = 0
            current_group = None

            jumlah_kegiatan_per_satker = {}

            for row in serialized_data:
                nama_satker = row['nama_satker']
                if nama_satker not in jumlah_kegiatan_per_satker:
                    jumlah_kegiatan_per_satker[nama_satker] = 0
                jumlah_kegiatan_per_satker[nama_satker] += 1

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    jumlah_kegiatan_per_satker[row['nama_satker']],
                    row['nama_observasi'],
                    numering,
                    formated_date,
                    ',\n'.join([f"{item['nama_peserta']}" for item in row['peserta']]),
                    row['deskripsi_hasil'],
                    row['simpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:13], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

class PSM_PENGUMPULAN_DATA_IKOTAN_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all()
    serializer_class = serializers.PSM_PENGUMPULAN_DATA_IKOTAN_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:

                parent_instance = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.get(pk=inserted_id)

                peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                )

    def perform_update(self, serializer):
        current_record_id = serializer.instance.pk

        serializer.save()

        models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent_id=current_record_id).delete()

        peserta_array_lama_json = self.request.data.get('peserta_lamaArray', None)
        if peserta_array_lama_json:
            data = json.loads(peserta_array_lama_json)
            if len(data) > 0:

                for entry in data:

                    parent_instance = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.get(pk=current_record_id)

                    peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                    )

        peserta_array_json = self.request.data.get('pesertaArray', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            if len(data) > 0:

                for entry in data:

                    parent_instance = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.get(pk=current_record_id)

                    peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                    )

        serializer.save(updated_by=self.request.user)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'satker_id__level','observasi', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'observasi': item['observasi'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    def get_data_peserta(self, parent):
        data = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def delete(self, request, *args, **kwargs):
        try:
            pk = self.kwargs.get('pk')
            deleted_peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(pk=pk).delete()
            return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}'}, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM DUKUNGAN STAKEHOLDER API =======
class PSM_DUKUNGAN_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_DUKUNGAN_STAKEHOLDER.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-created_at').distinct('satker__id')
    serializer_class = serializers.PSM_DUKUNGAN_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_DUKUNGAN_STAKEHOLDER_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)

        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'pemda', 'kegiatan', 'alamat', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'pemda': item['pemda'],
                'kegiatan': item['kegiatan'],
                'alamat': item['alamat'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'pemda', 'kegiatan', 'alamat', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status', 'satker_id__level'

        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'pemda': item['pemda'],
                'kegiatan': item['kegiatan'],
                'alamat': item['alamat'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerepan_anggaran': item['penyerepan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)

        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}

            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_DUKUNGAN_STAKEHOLDER.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def get_view_name(self):
        return "PSM DUKUNGAN STAKEHOLDER"

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_DUKUNGAN_STAKEHOLDER_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'Dukungan Stakeholder {tahun}'
        base_path = 'media/kegiatan/psm/dukungan_stakeholder/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN DUKUNGAN DATA STAKEHOLDER")
            sheet.merge_cells('A1:K1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:K2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'NAMA STAKEHOLDER', 'KEGIATAN YANG DILAKUKAN', 'PELAKSANA', 'JUMLAH SASARAN (ORANG) JIKA ADA', 'HASIL/DAMPAK', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('D4:D5')
            sheet.merge_cells('E4:E5')
            sheet.merge_cells('F4:F5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('K4:K5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['J4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'ANGGARAN'
            sheet.merge_cells('I4:J4')

            # Menambahkan subheader untuk 'ANGGARAN'
            subheaders = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']
            for col_num, subheader in enumerate(subheaders, start=9):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Mengatur tinggi baris untuk header dan subheader
            sheet.row_dimensions[1].height = 30
            sheet.row_dimensions[2].height = 30

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_RAKERNIS.objects.values(
                'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status'
            ).order_by(
                '-satker_id'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'satker_target': item['satker_target'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker_target__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            current_row = 6
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']
                data_row = [
                    group_count,
                    row['nama_satker'],
                    formated_date,
                    row['nama_satker_target'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                # Tambahkan data_row ke sheet
                for item, value in enumerate(data_row[:11], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    cell.alignment = Alignment(vertical='center')
                current_row += 1
                numering += 1

            # Pastikan untuk menggabungkan sel terakhir jika diperlukan
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

class PSM_DUKUNGAN_STAKEHOLDER_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_DUKUNGAN_STAKEHOLDER.objects.all()
    serializer_class = serializers.PSM_DUKUNGAN_STAKEHOLDER_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'pemda', 'kegiatan', 'alamat', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'status', 'satker_id__level'

        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'pemda': item['pemda'],
                'kegiatan': item['kegiatan'],
                'alamat': item['alamat'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

# ======= PSM KEGIATAN LAINNYA API =======
class PSM_KEGIATAN_LAINNYA_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_KEGIATAN_LAINNYA.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-created_at').distinct('satker__id')
    serializer_class = serializers.PSM_KEGIATAN_LAINNYA_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_KEGIATAN_LAINNYA_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)

        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_KEGIATAN_LAINNYA.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'kegiatan', 'tempat', 'waktu_awal', 'waktu_akhir', 'lingkungan', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'kegiatan_akun', 'uraian_kegiatan', 'drive_url', 'anggaran', 'penyerapan_anggaran', 'status', 'satker_id__level'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'kegiatan': item['kegiatan'],
                'tempat': item['tempat'],
                'waktu_awal': item['waktu_awal'],
                'waktu_akhir': item['waktu_akhir'],
                'lingkungan': item['lingkungan'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'kegiatan_akun': item['kegiatan_akun'],
                'uraian_kegiatan': item['uraian_kegiatan'],
                'drive_url': item['drive_url'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_KEGIATAN_LAINNYA.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'kegiatan', 'tempat', 'waktu_awal', 'waktu_akhir', 'lingkungan', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'kegiatan_akun', 'uraian_kegiatan', 'drive_url', 'anggaran', 'penyerapan_anggaran', 'status', 'satker_id__level'

        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'kegiatan': item['kegiatan'],
                'tempat': item['tempat'],
                'waktu_awal': item['waktu_awal'],
                'waktu_akhir': item['waktu_akhir'],
                'lingkungan': item['lingkungan'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'kegiatan_akun': item['kegiatan_akun'],
                'uraian_kegiatan': item['uraian_kegiatan'],
                'drive_url': item['drive_url'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)

        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}

            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_KEGIATAN_LAINNYA.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def get_view_name(self):
        return "PSM KEGIATAN LAINNYA"

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_KEGIATAN_LAINNYA.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_KEGIATAN_LAINNYA_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN LAINNYA {tahun}'
        base_path = 'media/kegiatan/psm/kegiatan_lainnya/exported'
        file_path = f'{base_path}/{file_name}.xlsx'
        
        # load template
        template_base_path = 'media/kegiatan/psm/kegiatan_lainnya/template'
        template_file_path = f'{template_base_path}/template.xlsx'
        
        workbook = openpyxl.load_workbook(template_file_path)
        sheet = workbook.get_sheet_by_name('Data Kegiatan')
        workbook.active = sheet
        
        try:
            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:O2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)
            
            data = models.PSM_KEGIATAN_LAINNYA.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'kegiatan', 'tempat', 'waktu_awal', 'waktu_akhir', 'lingkungan', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'kegiatan_akun__no_unit', 'kegiatan_akun__akun_kegiatan', 'uraian_kegiatan__no', 'uraian_kegiatan__uraian_kegiatan', 'uraian_kegiatan__pj', 'drive_url', 'anggaran', 'penyerapan_anggaran', 'status', 'satker_id__level'
            )
            
            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)
                
            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'kegiatan': item['kegiatan'],
                    'tempat': item['tempat'],
                    'waktu_awal': item['waktu_awal'],
                    'waktu_akhir': item['waktu_akhir'],
                    'lingkungan': item['lingkungan'],
                    'jumlah_sasaran': item['jumlah_sasaran'],
                    'hasil_dampak': item['hasil_dampak'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'kegiatan_akun_no_unit': item['kegiatan_akun__no_unit'],
                    'kegiatan_akun_akun_kegiatan': item['kegiatan_akun__akun_kegiatan'],
                    'uraian_kegiatan_no': item['uraian_kegiatan__no'],
                    'uraian_kegiatan_pj': item['uraian_kegiatan__pj'],
                    'uraian_kegiatan': item['uraian_kegiatan__uraian_kegiatan'],
                    'drive_url': item['drive_url'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'status': item['status'],
                    'dokumentasi': item['dokumentasi'],
                    'satker_level': item['satker_id__level'],
                }
                serialized_data.append(serialized_item)
                
            current_row = 7
            numering = 0
            group_count = 0
            current_group = None

            jumlah_kegiatan_per_satker = {}
            
            for row in serialized_data:
                nama_satker = row['nama_satker']
                if nama_satker not in jumlah_kegiatan_per_satker:
                    jumlah_kegiatan_per_satker[nama_satker] = 0
                jumlah_kegiatan_per_satker[nama_satker] += 1
                
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['waktu_akhir']:
                    formated_date = f'{row["waktu_awal"]} s/d {row["waktu_akhir"]}'
                else:
                    formated_date = row['waktu_awal']
                
                kode_kegiatan = f'Kegiatan: {row["kegiatan_akun_no_unit"]} - {row["kegiatan_akun_akun_kegiatan"]}'
                uraian_kegiatan = f'Uraian: {row["uraian_kegiatan_no"]} - {row["uraian_kegiatan"]} {row["uraian_kegiatan_pj"] if row["uraian_kegiatan_pj"] else ""}'  

                data_row = [
                    group_count,
                    row['nama_satker'],
                    jumlah_kegiatan_per_satker[row['nama_satker']],
                    f'{kode_kegiatan} \n\n {uraian_kegiatan}',
                    row['kegiatan'],
                    row['tempat'],
                    formated_date,
                    row['lingkungan'],
                    row['jumlah_sasaran'],
                    row['hasil_dampak'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['dokumentasi']
                ]

                for item, value in enumerate(data_row[:15], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1
                
            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

class PSM_KEGIATAN_LAINNYA_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_KEGIATAN_LAINNYA.objects.all()
    serializer_class = serializers.PSM_KEGIATAN_LAINNYA_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        data = models.PSM_KEGIATAN_LAINNYA.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'kegiatan', 'tempat', 'waktu_awal', 'waktu_akhir', 'lingkungan', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'anggaran', 'penyerapan_anggaran', 'drive_url', 'dokumentasi', 'kegiatan_akun', 'uraian_kegiatan', 'drive_url', 'anggaran', 'penyerapan_anggaran', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        # 'kegiatan_akun', 'uraian_kegiatan', 'drive_url', 'anggaran', 'penyerapan_anggaran'

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'kegiatan': item['kegiatan'],
                'tempat': item['tempat'],
                'waktu_awal': item['waktu_awal'],
                'waktu_akhir': item['waktu_akhir'],
                'lingkungan': item['lingkungan'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'kegiatan_akun': item['kegiatan_akun'],
                'uraian_kegiatan': item['uraian_kegiatan'],
                'drive_url': item['drive_url'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

# ======= PSM RAKOR PEMETAAN API =======
class PSM_RAKOR_PEMETAAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_RAKOR_PEMETAAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_RAKOR_PEMETAAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_RAKOR_PEMETAAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_RAKOR_PEMETAAN_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_RAKOR_PEMETAAN.objects.get(id=id)
        except models.PSM_RAKOR_PEMETAAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_RAKOR_PEMETAAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_RAKOR_PEMETAAN.objects.values(
            'id', 'satker_id', 'peserta', 'nama_lingkungan', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut','anggaran','penyerapan_anggaran','drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_RAKOR_PEMETAAN.objects.get(id=request.data['id'])
        except models.PSM_RAKOR_PEMETAAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_RAKOR_PEMETAAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_RAKOR_PEMETAAN.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_RAKOR_PEMETAAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_RAKOR_PEMETAAN_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'RAKOR PEMETAAN KELOMPOK SASARAN {tahun}'
        base_path = 'media/kegiatan/psm/rakor_pemetaan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="RAKOR PEMETAAN KELOMPOK SASARAN")
            sheet.merge_cells('A1:M1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:M2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

             # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','' 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('B4:B5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('M4:M5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['K4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('K4:L4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'SASARAN/LINGKUNGAN']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=11):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_RAKOR_PEMETAAN.objects.values(
                'id', 'satker_id', 'peserta', 'nama_lingkungan', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut','anggaran','penyerapan_anggaran','drive_url', 'dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'peserta': item['peserta'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_lingkungan': item['nama_lingkungan'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    '\n'.join([f"{item['nama']} - {item['jabatan']}" for item in row['peserta']]),  # Convert JSON to string,
                    row['nama_lingkungan'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:13], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM AUDIENSI API =======
class PSM_AUDIENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_AUDIENSI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_AUDIENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_AUDIENSI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_AUDIENSI_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_AUDIENSI.objects.get(id=id)
        except models.PSM_AUDIENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_AUDIENSI_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_AUDIENSI.objects.values(
            'id', 'satker_id', 'peserta', 'nama_lingkungan', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url','dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_AUDIENSI.objects.get(id=request.data['id'])
        except models.PSM_AUDIENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_AUDIENSI_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_AUDIENSI.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_AUDIENSI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_AUDIENSI_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'AUDIENSI KELOMPOK SASARAN ATAU STAKEHOLDER DALAM USAHA PENINGKATAN PERAN SERTA MASYARAKAT {tahun}'
        base_path = 'media/kegiatan/psm/audiensi/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="AUDIENSI KELOMPOK SASARAN/STAKEHOLDER DALAM USAHA PENINGKATAN PERAN SERTA MASYARAKAT")
            sheet.merge_cells('A1:M1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:M2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

             # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','' 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('B4:B5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('M4:M5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['K4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('K4:L4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'SASARAN/LINGKUNGAN']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=11):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_AUDIENSI.objects.values(
                'id', 'satker_id', 'peserta', 'nama_lingkungan', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url','dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'peserta': item['peserta'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_lingkungan': item['nama_lingkungan'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    '\n'.join([f"{item['nama']} - {item['jabatan']}" for item in row['peserta']]),  # Convert JSON to string,
                    row['nama_lingkungan'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:13], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM KONSOLIDASI_KEBIJAKAN API =======
class PSM_KONSOLIDASI_KEBIJAKAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_KONSOLIDASI_KEBIJAKAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_KONSOLIDASI_KEBIJAKAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_KONSOLIDASI_KEBIJAKAN_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.get(id=id)
        except models.PSM_KONSOLIDASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_KONSOLIDASI_KEBIJAKAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.values(
            'id', 'satker_id', 'peserta', 'stakeholder', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut','anggaran','penyerapan_anggaran','drive_url',  'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.get(id=request.data['id'])
        except models.PSM_KONSOLIDASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_KONSOLIDASI_KEBIJAKAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_KONSOLIDASI_KEBIJAKAN_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN RAPAT KONSOLIDASI KEBIJAKAN STAKEHOLDERS {tahun}'
        base_path = 'media/kegiatan/psm/konsolidasi_kebijakan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN RAPAT KONSOLIDASI KEBIJAKAN STAKEHOLDERS")
            sheet.merge_cells('A1:M1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:M2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','' 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('B4:B5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('M4:M5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['K4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('K4:L4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'STAKEHOLDER/ PENDAMPING YANG DIUNDANG']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=11):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.values(
                'id', 'satker_id', 'peserta', 'stakeholder', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut','anggaran','penyerapan_anggaran','drive_url',  'dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'peserta': item['peserta'],
                    'nama_satker': item['satker__nama_satker'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    '\n'.join([f"{item['nama']} - {item['jabatan']}" for item in row['peserta']]),  # Convert JSON to string,
                    row['stakeholder'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:13], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM WORKSHOP_PENGGIAT API =======
class PSM_WORKSHOP_PENGGIAT_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_WORKSHOP_PENGGIAT.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_WORKSHOP_PENGGIAT_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_WORKSHOP_PENGGIAT_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_WORKSHOP_PENGGIAT_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_WORKSHOP_PENGGIAT.objects.get(id=id)
        except models.PSM_WORKSHOP_PENGGIAT.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_WORKSHOP_PENGGIAT_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_WORKSHOP_PENGGIAT.objects.values(
            'id', 'satker_id', 'peserta', 'stakeholder', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_WORKSHOP_PENGGIAT.objects.get(id=request.data['id'])
        except models.PSM_WORKSHOP_PENGGIAT.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_WORKSHOP_PENGGIAT_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_WORKSHOP_PENGGIAT.objects.filter(id=id)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_WORKSHOP_PENGGIAT.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_WORKSHOP_PENGGIAT_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN WORKSHOP PENGGIAT P4GN {tahun}'
        base_path = 'media/kegiatan/psm/workshop_penggiat/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="KEGIATAN WORKSHOP PENGGIAT P4GN")
            sheet.merge_cells('A1:M1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:M2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '','' 'DESKRIPSI HASIL', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('B4:B5')

            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('M4:M5')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['K4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:G4')
            sheet.merge_cells('K4:L4')

            # Menambahkan subheader untuk 'PELAKSANAAN' & 'ANGGARAN'
            subheaders1 = ['NO', 'TANGGAL', 'PESERTA', 'STAKEHOLDER/ PENDAMPING YANG DIUNDANG']
            subheaders2 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']

            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            for col_num, subheader in enumerate(subheaders2, start=11):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_WORKSHOP_PENGGIAT.objects.values(
                'id', 'satker_id', 'peserta', 'stakeholder', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url', 'dokumentasi', 'status'
            ).order_by('-satker_id')

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'peserta': item['peserta'],
                    'nama_satker': item['satker__nama_satker'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    '\n'.join([f"{item['nama']} - {item['jabatan']}" for item in row['peserta']]),  # Convert JSON to string,
                    row['stakeholder'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:13], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM BIMBINGAN TEKNIS PENGGIAT P4GN API =======
class PSM_BIMTEK_P4GN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_BIMTEK_P4GN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_BIMTEK_P4GN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_BIMTEK_P4GN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_BIMTEK_P4GN_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            instance = serializer.save(created_by=self.request.user, status=status)
            inserted_id = instance.id
            peserta_bimtek = self.request.data.get('peserta_bimtek', None)
            if peserta_bimtek:
                data = json.loads(peserta_bimtek)
                for entry in data:
                    gender_mapping = {'L': 'L', 'P': 'P'}
                    jenis_kelamin = gender_mapping.get(entry['jenis_kelamin'], 'L')
                    parent_instance = models.PSM_BIMTEK_P4GN.objects.get(pk=inserted_id)
                    peserta = models.PSM_BIMTEK_P4GN_PESERTA.objects.create(
                        parent=parent_instance,
                        nama=entry['nama'],
                        jabatan=entry['jabatan'],
                        jenis_kelamin=jenis_kelamin,
                        alamat=entry['alamat'],
                        no_telepon=entry['no_telepon']
                    )
            return Response(serializer.data, status=201)

        return Response(serializer.errors, status=400)

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_BIMTEK_P4GN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_BIMTEK_P4GN_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)


            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_BIMTEK_P4GN.objects.filter(id=id)
            print(kegiatan  )

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_BIMTEK_P4GN.objects.get(id=id)
        except models.PSM_BIMTEK_P4GN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_BIMTEK_P4GN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_BIMTEK_P4GN.objects.values(
            'id', 'satker_id','nama_lingkungan', 'seri_pin_penggiat', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'hasil_capaian', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url',  'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'nama_lingkungan': item['nama_lingkungan'],
                'seri_pin_penggiat': item['seri_pin_penggiat'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'hasil_capaian': item['hasil_capaian'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'anggaran': item['anggaran'],
                'penyerapan_anggaran': item['penyerapan_anggaran'],
                'drive_url': item['drive_url'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level,
                'peserta_bimtek': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    def get_data_peserta(self, parent):
        data = models.PSM_BIMTEK_P4GN_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama', 'jabatan', 'jenis_kelamin', 'alamat','no_telepon'
        ).order_by('-id')

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama': item['nama'],
                'jabatan': item['jabatan'],
                'alamat': item['alamat'],
                'jenis_kelamin': item['jenis_kelamin'],
                'no_telepon': item['no_telepon']
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_BIMTEK_P4GN.objects.get(id=request.data['id'])
        except models.PSM_BIMTEK_P4GN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = serializers.PSM_BIMTEK_P4GN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            models.PSM_BIMTEK_P4GN_PESERTA.objects.filter(parent_id=request.data['id']).delete()
            peserta_bimtek = self.request.data.get('peserta_bimtek', None)
            if peserta_bimtek:
                data = json.loads(peserta_bimtek)
                for entry in data:
                    gender_mapping = {'L': 'L', 'P': 'P'}
                    jenis_kelamin = gender_mapping.get(entry['jenis_kelamin'], 'L')
                    parent_instance = models.PSM_BIMTEK_P4GN.objects.get(pk=request.data['id'])
                    peserta = models.PSM_BIMTEK_P4GN_PESERTA.objects.create(
                        parent=parent_instance,
                        nama=entry['nama'],
                        jabatan=entry['jabatan'],
                        jenis_kelamin=jenis_kelamin,
                        alamat=entry['alamat'],
                        no_telepon=entry['no_telepon']
                    )
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'BIMBINGAN TEKNIS PENGGIAT P4GN {tahun}'
        base_path = 'media/kegiatan/psm/bimtek_p4gn/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="BIMBINGAN TEKNIS PENGGIAT P4GN")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATKER PELAKSANA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '','' ,'LINGKUNGAN', '', '','','JUMLAH PESERTA', 'HASIL CAPAIAN', 'HAMBATAN/ KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', '', 'DOKUMENTASI']
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('A4:A6')
            sheet.merge_cells('B4:B6')
            sheet.merge_cells('C4:C6')
            sheet.merge_cells('K4:K6')
            sheet.merge_cells('L4:L6')
            sheet.merge_cells('M4:M6')
            sheet.merge_cells('N4:N6')
            sheet.merge_cells('O4:O6')
            sheet.merge_cells('R4:R6')

            # Menetapkan nilai ke sel pertama dari range yang akan digabungkan untuk 'ANGGARAN'
            sheet['D4'].value = 'PELAKSANAAN'
            sheet['F4'].value = 'LINGKUNGAN'
            sheet['P4'].value = 'ANGGARAN'
            # Menggabungkan sel untuk header 'PELAKSANAAN' & 'ANGGARAN'
            sheet.merge_cells('D4:E4')
            sheet.merge_cells('F4:J4')
            sheet.merge_cells('P4:Q4')

            subheaders1 = ['NO', 'TANGGAL']
            for col_num, subheader in enumerate(subheaders1, start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet.merge_cells('D5:D6')
            sheet.merge_cells('E5:E6')

            subheaders2 = ['NAMA LINGKUNGAN', 'NO SERI PIN PENGGIAT', 'PESERTA']
            for col_num, subheader in enumerate(subheaders2, start=6):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
            sheet.merge_cells('F5:F6')
            sheet.merge_cells('G5:G6')

            sheet['H5'].value = 'PESERTA'
            sheet.merge_cells('H5:J5')

            subheaders2_2 = ['NAMA DAN JABATAN', 'PRIA', 'WANITA']
            for col_num, subheader in enumerate(subheaders2_2, start=8):
                cell = sheet.cell(row=6, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders3 = ['PERENCANAAN ANGGARAN', 'PENYERAPAN ANGGARAN']
            for col_num, subheader in enumerate(subheaders3, start=16):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            sheet.merge_cells('P5:P6')
            sheet.merge_cells('Q5:Q6')

            # Menambahkan data
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_BIMTEK_P4GN.objects.values(
                'id', 'satker_id','nama_lingkungan', 'seri_pin_penggiat', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'hasil_capaian', 'kendala', 'kesimpulan', 'tindak_lanjut', 'anggaran','penyerapan_anggaran','drive_url',  'dokumentasi', 'status'
            ).order_by('-satker_id')
            
            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_lingkungan': item['nama_lingkungan'],
                    'seri_pin_penggiat': item['seri_pin_penggiat'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'hasil_capaian': item['hasil_capaian'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'anggaran': item['anggaran'],
                    'penyerapan_anggaran': item['penyerapan_anggaran'],
                    'drive_url': item['drive_url'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level,
                    'peserta_bimtek': self.get_data_peserta(item['id']),
                }
                serialized_data.append(serialized_item)

            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 7
            numering = 0
            group_count = 0
            rep_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center')

                formated_date = None
                if row['tanggal_akhir']:
                    formated_date = f'{row["tanggal_awal"]} s/d {row["tanggal_akhir"]}'
                else:
                    formated_date = row['tanggal_awal']

                # Menghitung jumlah data dengan jenis kelamin "L"
                jumlah_laki_laki = sum(1 for item in row['peserta_bimtek'] if item["jenis_kelamin"] == "L") or 0
                # Menghitung jumlah data dengan jenis kelamin "P"
                jumlah_perempuan = sum(1 for item in row['peserta_bimtek'] if item["jenis_kelamin"] == "P") or 0
                # Menghitung jumlah peserta_bimtek
                jumlah_peserta = (jumlah_laki_laki + jumlah_perempuan)

                data_row = [
                    group_count,
                    row['nama_satker'],
                    rep_count,
                    numering,
                    formated_date,
                    row['nama_lingkungan'],
                    row['seri_pin_penggiat'],
                    '\n'.join([f"{item['nama']} ({item['jenis_kelamin']}) - {item['jabatan']}" for item in row['peserta_bimtek']]),  # Convert JSON to string,
                    jumlah_laki_laki,
                    jumlah_perempuan,
                    jumlah_peserta,
                    row['hasil_capaian'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    row['penyerapan_anggaran'],
                    row['drive_url']
                ]

                for item, value in enumerate(data_row[:18], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # Autofit columns
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

class PSM_BINTEK_P4GN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_TES_URINE_COUNT.objects.all()
    serializer_class = serializers.PSM_TEST_URINE_DETEKSI_DINI_COUNT

# ======= PSM JADWAL KEGIATAN TAHUNAN API =======
class PSM_JADWAL_KEGIATAN_TAHUNAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-created_at').distinct('satker__id')
    serializer_class = serializers.PSM_JADWAL_KEGIATAN_TAHUNAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_JADWAL_KEGIATAN_TAHUNAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)

        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'nama_kegiatan', 'metode_kegiatan', 'tempat_pelaksana', 'waktu_kegiatan', 'jumlah_peserta', 'anggaran', 'satker_id__level'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'nama_kegiatan': item['nama_kegiatan'],
                'metode_kegiatan': item['metode_kegiatan'],
                'tempat_pelaksana': item['tempat_pelaksana'],
                'waktu_kegiatan': item['waktu_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'anggaran': item['anggaran'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'nama_kegiatan', 'metode_kegiatan', 'tempat_pelaksana', 'waktu_kegiatan', 'jumlah_peserta', 'anggaran', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'nama_kegiatan': item['nama_kegiatan'],
                'metode_kegiatan': item['metode_kegiatan'],
                'tempat_pelaksana': item['tempat_pelaksana'],
                'waktu_kegiatan': item['waktu_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'anggaran': item['anggaran'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)

        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}

            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def get_view_name(self):
        return "PSM KEGIATAN LAINNYA"

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_JADWAL_KEGIATAN_TAHUNAN_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

class PSM_JADWAL_KEGIATAN_TAHUNAN_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.all()
    serializer_class = serializers.PSM_JADWAL_KEGIATAN_TAHUNAN_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        early_instance = serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        # kode_id = early_instance.kode_id

        # latest_instance = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.filter(satker=satker, kode_id=kode_id).order_by('-created_at').first()

        # if latest_instance:
        #     try:
        #         second_latest_instance = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.filter(satker=satker, kode_id=kode_id, created_at__lt=latest_instance.created_at).order_by('-created_at')[1]
        #     except IndexError:
        #         second_latest_instance = latest_instance

        # def increment_string(s):
        #     if not s or not s[-1].isalpha():
        #         return s + 'A'
        #     else:
        #         last_char = s[-1]
        #         if last_char == 'Z':
        #             return increment_string(s[:-1]) + 'A'
        #         else:
        #             return s[:-1] + chr(ord(last_char) + 1)

        # if  second_latest_instance.ordering_asc:
        #     second_latest_instance = second_latest_instance.ordering_asc
        #     if 'Z' * len(second_latest_instance) == second_latest_instance:  # Check if 'Z' is repeated
        #         ordering_asc = increment_string('A' * (len(second_latest_instance) + 1))
        #     else:
        #         ordering_asc = increment_string(second_latest_instance)
        # else:
        #     ordering_asc = 'A'

        # serializer.save(created_by=self.request.user, status=status, ordering_asc=ordering_asc)
        serializer.save(created_by=self.request.user, status=status)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        data = models.PSM_JADWAL_KEGIATAN_TAHUNAN.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'kode', 'uraian', 'waktu_kegiatan', 'metode_kegiatan', 'tempat_pelaksana', 'jumlah_peserta', 'keterangan', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'kode': item['kode'],
                'uraian': item['uraian'],
                'metode_kegiatan': item['metode_kegiatan'],
                'tempat_pelaksana': item['tempat_pelaksana'],
                'waktu_kegiatan': item['waktu_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'keterangan': item['keterangan'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

class PSM_JADWAL_ViewSet(viewsets.ModelViewSet):
    queryset = models.Kegiatan_akun.objects.all()
    serializer_class = serializers.PSM_JADWAL_SERIALIZERS