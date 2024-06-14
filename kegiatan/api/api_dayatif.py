import textwrap
from django.db.models.functions import ExtractMonth, ExtractYear
from django.template.base import filter_re
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from survei.models import DataSurvei, TipeSurvei
from survey.models import survey as FormulirElektronik
from users.models import Profile, Satker

from django.db.models import Count, Q

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import re
import os
import codecs
import shutil
import datetime
import calendar
import pandas as pd
from django.conf import settings

from sidamas import pagination

from kegiatan import models
from kegiatan.api import serializers
from kegiatan.api import filters
from users.serializers import SatkerSerializer
from users.models import Satker

from kegiatan.api.helpers.api_helpers import (
    get_data_list,
    get_data_list_dukungan,
    get_data_list_queryset,
    get_data_list_bnnk,
    get_export_file_name,
    get_filtered_data,
    set_created_kegiatan_status,
    get_tanggal_kegiatan,

    kirim_kegiatan_helper,
    delete_all_kegiatan_helper,
    aksi_semua_kegiatan,
)

# ======= SATKER KEGIATAN =======
class DAYATIF_SurveiCountView(APIView):

    def get(self, request):
        current_year = timezone.now().year
        satker_id = request.query_params.get('satker', None)

        def get_counts_formulir(queryset):
            counts = queryset.annotate(
                year=ExtractYear('created'),
                month=ExtractMonth('created')
            ).values('year', 'month').annotate(
                total=Count('id')
            ).order_by('year', 'month')

            result = {}
            for item in counts:
                year = item['year']
                month = item['month']
                if year not in result:
                    result[year] = {m: 0 for m in range(1, 13)}
                    result[year]['total'] = 0
                result[year][month] = item['total']
                result[year]['total'] += item['total']

            # Ensure all months are present with a value of 0 if no data is found
            for year in result:
                for month in range(1, 13):
                    if month not in result[year]:
                        result[year][month] = 0

            # Ensure the current year is present in the result
            if current_year not in result:
                result[current_year] = {m: 0 for m in range(1, 13)}
                result[current_year]['total'] = 0

            return result

        def get_counts(queryset, satker_id):
            counts = queryset.annotate(
                year=ExtractYear('created_at'),
                month=ExtractMonth('created_at')
            ).values('year', 'month').annotate(
                total=Count('id')
            ).order_by('year', 'month')

            result = {}
            for item in counts:
                year = item['year']
                month = item['month']
                if year not in result:
                    result[year] = {m: 0 for m in range(1, 13)}
                    result[year]['total'] = 0
                result[year][month] = item['total']
                result[year]['total'] += item['total']

            # Ensure all months are present with a value of 0 if no data is found
            for year in result:
                for month in range(1, 13):
                    if month not in result[year]:
                        result[year][month] = 0

            # Ensure the current year is present in the result
            if current_year not in result:
                result[current_year] = {m: 0 for m in range(1, 13)}
                result[current_year]['total'] = 0

            return result

        tipe_survei_skm_lifeskill = TipeSurvei.objects.get(nama="SKM Life Skill")
        data_survei_skm_lifeskill = DataSurvei.objects.filter(tipe=tipe_survei_skm_lifeskill.id)

        tipe_survei_keberhasilan = TipeSurvei.objects.get(nama="Keberhasilan dan Kewirausahaan")
        data_survei_keberhasilan = DataSurvei.objects.filter(tipe=tipe_survei_keberhasilan.id)

        data_formulir_elektronik = FormulirElektronik.objects.all()


        data = {
            'formulir_elektronik': get_counts_formulir(data_formulir_elektronik),
            'skm_lifeskill': get_counts(data_survei_skm_lifeskill, satker_id),
            'keberhasilan_dan_kewirausahaan': get_counts(data_survei_keberhasilan, satker_id),
            'survei_ikrn': get_counts(data_survei_keberhasilan, satker_id),
        }

        return Response({
            'status': True,
            'data': data
        }, status=status.HTTP_200_OK)

class DAYATIF_KegiatanCountView(APIView):

    def get(self, request):
        current_year = timezone.now().year
        satker_id = request.query_params.get('satker', None)

        def get_counts(queryset, satker_id):
            if satker_id:
                queryset = queryset.filter(satker_id=satker_id)
            counts = queryset.annotate(
                year=ExtractYear('created_at'),
                month=ExtractMonth('created_at')
            ).values('year', 'month').annotate(
                total=Count('id')
            ).order_by('year', 'month')

            result = {}
            for item in counts:
                year = item['year']
                month = item['month']
                if year not in result:
                    result[year] = {m: 0 for m in range(1, 13)}
                    result[year]['total'] = 0
                result[year][month] = item['total']
                result[year]['total'] += item['total']

            # Ensure all months are present with a value of 0 if no data is found
            for year in result:
                for month in range(1, 13):
                    if month not in result[year]:
                        result[year][month] = 0

            # Ensure the current year is present in the result
            if current_year not in result:
                result[current_year] = {m: 0 for m in range(1, 13)}
                result[current_year]['total'] = 0

            return result

        data = {
            'binaan_teknis': get_counts(models.DAYATIF_BINAAN_TEKNIS.objects.all(), satker_id),
            'pemetaan_potensi': get_counts(models.DAYATIF_PEMETAAN_POTENSI.objects.all(), satker_id),
            'pemetaan_stakeholder': get_counts(models.DAYATIF_PEMETAAN_STAKEHOLDER.objects.all(), satker_id),
            'rapat_sinergi_stakeholder': get_counts(models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER.objects.all(), satker_id),
            'bimbingan_teknis_stakeholder': get_counts(models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER.objects.all(), satker_id),
            'bimbingan_teknis_lifeskill': get_counts(models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.all(), satker_id),
            'monitoring_dan_evaluasi': get_counts(models.DAYATIF_MONEV_DAYATIF.objects.all(), satker_id),
            'dukungan_stakeholder': get_counts(models.DAYATIF_DUKUNGAN_STAKEHOLDER.objects.all(), satker_id)
        }

        return Response({
            'status': True,
            'data': data
        }, status=status.HTTP_200_OK)

class DAYATIF_SatkerCountView(APIView):
    def get(self, request):
        satker_id = request.query_params.get('satker', None)

        # If satker_id is provided, filter by satker_id, else get all satkers
        if satker_id:
            satkers = Satker.objects.filter(id=satker_id).annotate(
                binaan_teknis_count=Count('dayatif_binaan_teknis'),
                pemetaan_potensi_count=Count('dayatif_pemetaan_potensi'),
                pemetaan_stakeholder_count=Count('dayatif_pemetaan_stakeholder'),
                rapat_sinergi_count=Count('dayatif_rapat_sinergi_stakeholder'),
                bimbingan_teknis_stakeholder_count=Count('dayatif_bimbingan_teknis_stakeholder'),
                bimbingan_teknis_lifeskill_count=Count('dayatif_bimbingan_teknis_lifeskill'),
                monev_dayatif_count=Count('dayatif_monev_dayatif'),
                dukungan_stakeholder_count=Count('dayatif_dukungan_stakeholder')
            )
        else:
            satkers = Satker.objects.annotate(
                binaan_teknis_count=Count('dayatif_binaan_teknis'),
                pemetaan_potensi_count=Count('dayatif_pemetaan_potensi'),
                pemetaan_stakeholder_count=Count('dayatif_pemetaan_stakeholder'),
                rapat_sinergi_count=Count('dayatif_rapat_sinergi_stakeholder'),
                bimbingan_teknis_stakeholder_count=Count('dayatif_bimbingan_teknis_stakeholder'),
                bimbingan_teknis_lifeskill_count=Count('dayatif_bimbingan_teknis_lifeskill'),
                monev_dayatif_count=Count('dayatif_monev_dayatif'),
                dukungan_stakeholder_count=Count('dayatif_dukungan_stakeholder')
            )

        response_data = []
        for satker in satkers:
            response_data.append({
                'satker': SatkerSerializer(satker).data,
                'binaan_teknis': satker.binaan_teknis_count,
                'pemetaan_potensi': satker.pemetaan_potensi_count,
                'pemetaan_stakeholder': satker.pemetaan_stakeholder_count,
                'rapat_sinergi': satker.rapat_sinergi_count,
                'bimbingan_teknis_stakeholder': satker.bimbingan_teknis_stakeholder_count,
                'bimbingan_teknis_lifeskill': satker.bimbingan_teknis_lifeskill_count,
                'monev_dayatif': satker.monev_dayatif_count,
                'dukungan_stakeholder': satker.dukungan_stakeholder_count
            })

        return Response({
            'status': True,
            'data': response_data
        }, status=status.HTTP_200_OK)

# ======= KEGIATAN SATKER =======
class DAYATIF_KEGIATAN_SATKER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_KEGIATAN_SATKER.objects.all()
    serializer_class = serializers.DAYATIF_KEGIATAN_SATKER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page100NumberPagination
    # filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['id']

# ======= BINAAN TEKNIS =======
class DAYATIF_BINAAN_TEKNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_BINAAN_TEKNIS.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_BINAAN_TEKNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.DAYATIF_BINAAN_TEKNIS_Filters

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_BINAAN_TEKNIS_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_BINAAN_TEKNIS, serializers.DAYATIF_BINAAN_TEKNIS_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_BINAAN_TEKNIS, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_BINAAN_TEKNIS)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_BINAAN_TEKNIS, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_BINAAN_TEKNIS.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'jumlah_hari_pelaksanaan', 'satker_target', 'satker_target__nama_satker', 'jumlah_peserta', 'tujuan', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)

        print('Panjang data sebelum filter:', len(data))

        data = get_filtered_data(satker, data, status, waktu)

        print('Panjang data sesudah filter:', len(  data))

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'jumlah_hari_pelaksanaan': item['jumlah_hari_pelaksanaan'],
                'satker_target': item['satker_target'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'jumlah_peserta': item['jumlah_peserta'],
                'tujuan': item['tujuan'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'REKAPITULASI PEMBINAAN TEKNIS {tahun}'
        base_path = 'media/kegiatan/dayatif/binaan_teknis/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="REKAPITULASI PEMBINAAN TEKNIS")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '', '', '', 'RINCIAN HASIL PELAKSANAAN PROGRAM PEMBERDAYAAN ALTERNATIF PADA SATKER TARGET', '', '', 'TINDAK LANJUT', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('L4:L5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:H4') # -> Horizontally

            sheet['I4'].value = 'RINCIAN HASIL PELAKSANAAN PROGRAM PEMBERDAYAAN ALTERNATIF PADA SATKER TARGET'
            sheet.merge_cells('I4:K4') # -> Horizontally

            sheet['M4'].value = 'DOKUMENTASI'
            sheet.merge_cells('M4:N4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'JUMLAH HARI PELAKSANAAN KERJA', 'SATKER TARGET', 'JUMLAH PESERTA']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['RINCIAN'] = ['TUJUAN', 'HAMBATAN/KENDALA', 'KESIMPULAN']

            for col_num, subheader in enumerate(subheaders['RINCIAN'], start=column_index_from_string('I')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('M')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= SOURCE DATA =======
            data = models.DAYATIF_BINAAN_TEKNIS.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'satker_target__nama_satker', 'jumlah_hari_pelaksanaan', 'jumlah_peserta',
                'tujuan', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'status': item['status'],
                    'satker_level': satker.level,
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker__nama_satker'],
                    'jumlah_hari_pelaksanaan': item['jumlah_hari_pelaksanaan'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'tujuan': item['tujuan'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('M'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('M'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    numering,
                    row['tanggal'],
                    row['jumlah_hari_pelaksanaan'],
                    row['nama_satker_target'],
                    row['jumlah_peserta'],
                    row['tujuan'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")',
                    dokumentasi_gambar
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['I'].width = 45 # Tujuan
            sheet.column_dimensions['M'].width = 18 # Dokumentasi
            sheet.column_dimensions['N'].width = 28 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======= PEMETAAN POTENSI =======
class DAYATIF_PEMETAAN_POTENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_PEMETAAN_POTENSI.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_PEMETAAN_POTENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.DAYATIF_PEMETAAN_POTENSI_Filters

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_PEMETAAN_POTENSI_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_PEMETAAN_POTENSI, serializers.DAYATIF_PEMETAAN_POTENSI_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_PEMETAAN_POTENSI, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_PEMETAAN_POTENSI)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_PEMETAAN_POTENSI, request)

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'PEMETAAN POTENSI SDM DAN SDA KAWASAN RAWAN NARKOBA {tahun}'
        base_path = 'media/kegiatan/dayatif/pemetaan_potensi/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="PEMETAAN POTENSI SDM DAN SDA KAWASAN RAWAN NARKOBA")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '', 'DESKRIPSI HASIL', 'HAMBATAN/KENDALA', 'KESIMPULAN', 'TINDAK LANJUT', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:F4') # -> Horizontally

            sheet['M4'].value = 'DOKUMENTASI'
            sheet.merge_cells('M4:N4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'LOKASI']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('K')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= SOURCE DATA =======
            data = models.DAYATIF_PEMETAAN_POTENSI.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi',
                'potensi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                lokasi = f"{item['nama_desa']}, {item['nama_kecamatan']}, {item['nama_kabupaten']}, {item['nama_provinsi']}"

                # response_potensi = {
                #         "potensi": {
                #             "sda": {
                #                 "swot": {
                #                     "text": "analisis",
                #                     "peluang": "peluang",
                #                     "kekuatan": "kekuatan",
                #                     "kelemahan": "kelemahan",
                #                     "tantangan": "tantangan"
                #                 },
                #                 "text": "potensi",
                #                 "kesimpulan": "kesimpulan"
                #             },
                #             "sdm": {
                #                 "swot": {
                #                     "text": "analsis",
                #                     "peluang": "peluang",
                #                     "kekuatan": "kekuatan",
                #                     "kelemahan": "kelemahan",
                #                     "tantangan": "tantangan"
                #                 },
                #                 "text": "potensi",
                #                 "kesimpulan": "kesimpulan"
                #             }
                #         },
                # }

                # formatted_potensi = f'
                #     1. Potensi SDA : potensi\n
                #     \n\tAnalisis SWOT : analisis
                #             \n\t\tKekuatan : kekuatan
                #             \n\t\tKelemahan : kelamahan
                #             \n\t\tPeluang : peluang
                #             \n\t\tTantangan : tantangan
                #             \n\t\tKesimpulan : kesimpulan
                #     2. Potensi SDM : potensi\n
                #     \n\tAnalisis SWOT : analsis
                #             \n\t\tKekuatan : kekuatan
                #             \n\t\tKelemahan : kelemahan
                #             \n\t\tPeluang : peluang
                #             \n\t\tTantangan : tantangan
                #             \n\t\tKesimpulan : kesimpulan
                # '

                sda = item['potensi']['sda']
                sdm = item['potensi']['sdm']

                formatted_potensi = textwrap.dedent(f"""
                    1. Potensi SDA : {sda['text']}
                        Analisis SWOT : {sda['swot']['text']}
                            Kekuatan : {sda['swot']['kekuatan']}
                            Kelemahan : {sda['swot']['kelemahan']}
                            Peluang : {sda['swot']['peluang']}
                            Tantangan : {sda['swot']['tantangan']}
                        Kesimpulan : {sda['kesimpulan']}

                    2. Potensi SDM : {sdm['text']}

                        Analisis SWOT : {sdm['swot']['text']}
                            Kekuatan : {sdm['swot']['kekuatan']}
                            Kelemahan : {sdm['swot']['kelemahan']}
                            Peluang : {sdm['swot']['peluang']}
                            Tantangan : {sdm['swot']['tantangan']}
                        Kesimpulan : {sdm['kesimpulan']}
                """)


                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'status': item['status'],
                    'satker_level': satker.level,
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'nama_satker': item['satker__nama_satker'],
                    'lokasi': lokasi,
                    'deskripsi': formatted_potensi,
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('K'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('L'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    numering,
                    row['tanggal'],
                    row['lokasi'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")',
                    dokumentasi_gambar
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['G'].width = 32 # Deskripsi (POTENSI)
            sheet.column_dimensions['K'].width = 18 # Dokumentasi
            sheet.column_dimensions['L'].width = 28 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======= PEMETAAN STAKEHOLDER =======
class DAYATIF_PEMETAAN_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_PEMETAAN_STAKEHOLDER.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_PEMETAAN_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.DAYATIF_PEMETAAN_STAKEHOLDER_Filters

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_PEMETAAN_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_PEMETAAN_STAKEHOLDER, serializers.DAYATIF_PEMETAAN_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_PEMETAAN_STAKEHOLDER, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_PEMETAAN_STAKEHOLDER)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_PEMETAAN_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_PEMETAAN_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)

        print('Panjang data sebelum filter:', len(data))

        data = get_filtered_data(satker, data, status, waktu)

        print('Panjang data sesudah filter:', len(data))

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],

                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],

                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'PEMETAAN STAKEHOLDER {tahun}'
        base_path = 'media/kegiatan/dayatif/pemetaan_stakeholder/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="PEMETAAN STAKEHOLDER")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '', '',
            'DESKRIPSI HASIL', 'KENDALA/HAMBATAN', 'KESIMPULAN', 'TINDAK LANJUT', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')
            # sheet.merge_cells('L4:L5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:G4') # -> Horizontally

            sheet['L4'].value = 'DOKUMENTASI'
            sheet.merge_cells('L4:M4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'LOKASI', 'STAKEHOLDER']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('L')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= SOURCE DATA =======
            data = models.DAYATIF_PEMETAAN_STAKEHOLDER.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'nama_desa', 'nama_kecamatan', 'nama_kabupaten',
                'nama_provinsi',
                'stakeholders', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                lokasi = f"{item['nama_desa']}, {item['nama_kecamatan']}, {item['nama_kabupaten']}, {item['nama_provinsi']}"

                stakeholders = item['stakeholders']

                formatted_stakeholders = "\n".join([f"{index + 1}. {stakeholder['nama']}" for index, stakeholder in enumerate(stakeholders)])

                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'status': item['status'],
                    'satker_level': satker.level,
                    'lokasi': lokasi,
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'nama_satker': item['satker__nama_satker'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'stakeholders': formatted_stakeholders,
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('L'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('M'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    numering,
                    row['tanggal'],
                    row['lokasi'],
                    row['stakeholders'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")',
                    dokumentasi_gambar
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                        cell = sheet.cell(row=current_row, column=item, value=value)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['L'].width = 18 # Dokumentasi
            sheet.column_dimensions['M'].width = 28 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'], url_path='export2', name='Export Data')
    def export_data2(self, request):
        satker = self.request.user.profile.satker
        filter_status_pengiriman = request.data.get("status", None)
        filter_rentang_waktu = request.data.get("waktu", None)

        serialized_data = self.get_flat_values(request, filter_status_pengiriman, filter_rentang_waktu)

        # return Response({
        #     'status': True,
        #     'status': filter_status_pengiriman,
        #     'rentang_waktu': filter_rentang_waktu,
        #     'data': len(serialized_data),
        # })

        # ======= GET FILE NAME =======
        file_name = ''

        tahun = datetime.datetime.now().year
        base_path = 'media/kegiatan/pemetaan_STAKEHOLDER/exported'
        main_name = 'PEMETAAN STAKEHOLDER SDM DAN SDA KAWASAN RAWAN NARKOBA'

        if satker.level != 2:
            if filter_rentang_waktu == 'semua':
                file_name = f'{main_name} {satker.nama_satker.upper()} TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan1':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 1 (Januari - Maret) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan2':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 2 (April - Juni) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan3':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 3 (July - September) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan4':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 4 (Oktober - Desember) TAHUN {tahun}'
            elif filter_rentang_waktu == 'hari_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%d %B %Y")}'
            elif filter_rentang_waktu == 'minggu_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%A, %d %B %Y")}'
            elif filter_rentang_waktu == 'bulan_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%B %Y")}'
        else:
            file_name = f'{main_name} BNNK & BNNP TAHUN {tahun}'

        file_path = f'{base_path}/{file_name}.xlsx'

        if os.path.exists(base_path): shutil.rmtree(base_path)
        os.makedirs(base_path, exist_ok=True)

        base_url = self.request.build_absolute_uri('/') + 'media/'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'{satker.nama_satker}'

        # return Response({
        #     'status': True,
        #     'file_name': file_name,
        # })

        try:
            # ======= HEADERS =======
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'JUMLAH KEGIATAN',
                'STATUS',
                'NO.',
                # 'SATUAN KERJA TARGET',
                'TANGGAL',
                'LOKASI',
                # 'JUMLAH PESERTA',
                'DESKRIPSI',
                'HAMBATAN/KENDALA',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI'
            ]

            # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

            current_row = 4

            # ======= GENERATE HEADERS =======
            for item, header in enumerate(headers[:len(headers)], start=1):
                cell = sheet.cell(row=current_row, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            current_row += 1

            no = 0
            no_child = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    no += 1

                no_child += 1

                formatted_data = {}

                # ======= STATUS =======
                status_mapping = {
                    0: 'Belum dikirim',
                    1: 'Dikirim ke BNNP',
                    2: 'Dikirim ke BNN Pusat'
                }

                formatted_data['status'] = status_mapping.get(row['status'], '-')

                formatted_data['jumlah_kegiatan'] = f"{len(serialized_data)} kali"
                formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])

                # ======= LOKASI =======
                merged_lokasi = f"{row['nama_desa']}, Kecamatan {row['nama_kecamatan']}, {row['nama_kabupaten']}, Provinsi {row['nama_provinsi']}"

                formatted_data['lokasi'] = merged_lokasi

                for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
                    celcol = sheet.cell(row=current_row, column=col, value=current_group)
                    celcol.font = openpyxl.styles.Font(bold=True)
                    celcol.alignment = Alignment(horizontal='center', vertical='center')

                # ======= COLUMN ADJUSMENT =======
                # Index start dari 1

                # No Kegiatan
                col_no = column_index_from_string('E')
                cell_no = sheet.cell(row=current_row, column=col_no, value=current_group)
                cell_no.font = openpyxl.styles.Font(bold=True)
                cell_no.alignment = Alignment(horizontal='center', vertical='center')

                # Tanggal
                col_tanggal = column_index_from_string('F')
                cell_tanggal = sheet.cell(row=current_row, column=col_tanggal, value=current_group)
                cell_tanggal.alignment = Alignment(horizontal='center', vertical='center')

                # Dokumentasi
                col_dokumentasi = column_index_from_string('L')
                cell_dokumentasi = sheet.cell(row=current_row, column=col_dokumentasi, value=current_group)
                cell_dokumentasi.alignment = Alignment(horizontal='center', vertical='center')

                data_row = [
                    no,
                    row['nama_satker'],
                    formatted_data['jumlah_kegiatan'],
                    formatted_data['status'],
                    no_child,
                    formatted_data['tanggal'],
                    formatted_data['lokasi'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    # row['dokumentasi'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
                ]

                for item, value in enumerate(data_row[:len(headers)], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1

            # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
            sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
            sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

            # ======= AUTOFIT KOLOM =======
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.3
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['L'].width = 35

            # ======= SAVE FILE =======
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diekspor!',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

# ======= RAPAT SINERGI STAKEHOLDER =======
class DAYATIF_RAPAT_SINERGI_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_RAPAT_SINERGI_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_RAPAT_SINERGI_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER, serializers.DAYATIF_RAPAT_SINERGI_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)

        print('Panjang data sebelum filter:', len(data))

        data = get_filtered_data(satker, data, status, waktu)

        print('Panjang data sesudah filter:', len(data))

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],

                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],

                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'RAPAT SINERGI STAKEHOLDER {tahun}'
        base_path = 'media/kegiatan/dayatif/rapat_sinergi_stakeholder/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="RAPAT SINERGI STAKEHOLDER")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '', '',
            'DESKRIPSI HASIL', 'KENDALA/HAMBATAN', 'KESIMPULAN', 'TINDAK LANJUT', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:G4') # -> Horizontally

            sheet['L4'].value = 'DOKUMENTASI'
            sheet.merge_cells('L4:M4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'JUMLAH PESERTA', 'STAKEHOLDER/PENDAMPING YANG HADIR']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=4):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('L')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )


            # ======= SOURCE DATA =======
            data = models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'jumlah_peserta',
                'stakeholders', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                stakeholders = item['stakeholders']

                formatted_stakeholders = "\n".join([f"{index + 1}. {stakeholder['nama']}" for index, stakeholder in enumerate(stakeholders)])

                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'nama_satker': item['satker__nama_satker'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'stakeholders': formatted_stakeholders,
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                    'status': item['status'],
                    'satker_level': satker.level
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('L'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi = f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")'

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('M'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    numering,
                    row['tanggal'],
                    row['jumlah_peserta'],
                    row['stakeholders'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    dokumentasi,
                    dokumentasi_gambar,
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                    if item in (column_index_from_string('K'), column_index_from_string('I')):
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['J'].width = 40 # Kesimpulan
            sheet.column_dimensions['L'].width = 18 # Dokumentasi
            sheet.column_dimensions['M'].width = 25 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

# ======= BIMBINGAN_TEKNIS_STAKEHOLDER =======
class DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER, serializers.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)

        print('Panjang data sebelum filter:', len(data))

        data = get_filtered_data(satker, data, status, waktu)

        print('Panjang data sesudah filter:', len(data))

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],

                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],

                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'BIMBINGAN TEKNIS STAKEHOLDER {tahun}'
        base_path = 'media/kegiatan/dayatif/bimbingan_teknis_stakeholder/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="BIMBINGAN TEKNIS STAKEHOLDER")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'JENIS_BIMBINGAN', 'PELAKSANAAN', '', '', '', '',
            'DESKRIPSI HASIL', 'KENDALA/HAMBATAN', 'KESIMPULAN', 'TINDAK LANJUT', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num, value=header)
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('D4:D5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')
            sheet.merge_cells('L4:L5')
            sheet.merge_cells('M4:M5')
            # sheet.merge_cells('N4:N5')
            # sheet.merge_cells('O4:O5')
            # sheet.merge_cells('P4:P5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['E4'].value = 'PELAKSANAAN'
            sheet.merge_cells('E4:I4') # -> Horizontally

            sheet['N4'].value = 'DOKUMENTASI'
            sheet.merge_cells('N4:O4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'TEMPAT', 'JUMLAH PESERTA', 'STAKEHOLDER/PENDAMPING']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('E')):
                cell = sheet.cell(row=5, column=col_num, value=subheader)
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('N')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= SOURCE DATA =======
            data = models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'jenis_bimbingan', 'tempat', 'jumlah_peserta',
                'stakeholders', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                stakeholders = item['stakeholders']

                formatted_stakeholders = "\n".join([f"{index + 1}. {stakeholder['nama']}" for index, stakeholder in enumerate(stakeholders)])

                def get_bimbingan(data):
                    dataMap = {
                        'bimbingan_teknis_stakeholder': 'Bimbingan Teknis Stakeholder',
                        'bimbingan_teknis_pendamping': 'Bimbingan Teknis Pendamping'
                    }

                    return dataMap[data] or '-'

                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'jumlah_peserta': f"{item['jumlah_peserta']} orang",
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'jenis_bimbingan': get_bimbingan(item['jenis_bimbingan']),
                    'tempat': item['tempat'],
                    'nama_satker': item['satker__nama_satker'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'stakeholders': formatted_stakeholders,
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                    'status': item['status'],
                    'satker_level': satker.level
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('N'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi = f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")'

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('O'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    row['jenis_bimbingan'],
                    numering,
                    row['tanggal'],
                    row['tempat'],
                    row['jumlah_peserta'],
                    row['stakeholders'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    dokumentasi,
                    dokumentasi_gambar,
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                    if item in (column_index_from_string('K'), column_index_from_string('I'), column_index_from_string('J')):
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['J'].width = 40 # Deskripsi Hasil
            sheet.column_dimensions['L'].width = 40 # Kesimpulan
            sheet.column_dimensions['N'].width = 18 # Dokumentasi
            sheet.column_dimensions['O'].width = 25 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======= BIMBINGAN_TEKNIS_LIFESKILL =======
class DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL, serializers.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)

        print('Panjang data sebelum filter:', len(data))

        data = get_filtered_data(satker, data, status, waktu)

        print('Panjang data sesudah filter:', len(data))

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],

                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],

                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def export_data_new(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'BIMBINGAN TEKNIS LIFESKILL {tahun}'
        base_path = 'media/kegiatan/dayatif/bimbingan_teknis_lifeskill/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="BIMBINGAN TEKNIS LIFESKILL")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '',
                    'KAWASAN YANG DIINTERVENSI', 'KETERAMPILAN YANG DIBERIKAN DALAM PELATIHAN', 'PESERTA', '',
                    'SINERGI', '', 'HASIL SKM', '',
                    'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('O4:O5')
            sheet.merge_cells('P4:P5')
            sheet.merge_cells('Q4:Q5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:F4')  # -> Horizontally

            sheet['I4'].value = 'PESERTA'
            sheet.merge_cells('I4:J4')  # -> Horizontally

            sheet['K4'].value = 'SINERGI'
            sheet.merge_cells('K4:L4')  # -> Horizontally

            sheet['M4'].value = 'HASIL SKM'
            sheet.merge_cells('M4:N4')  # -> Horizontally

            sheet['R4'].value = 'DOKUMENTASI'
            sheet.merge_cells('R4:S4')  # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'TEMPAT']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['PESERTA'] = ['DAFTAR', 'JUMLAH']

            for col_num, subheader in enumerate(subheaders['PESERTA'], start=column_index_from_string('I')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['SINERGI'] = ['DESA BERSINAR', 'IBM']

            for col_num, subheader in enumerate(subheaders['SINERGI'], start=column_index_from_string('K')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['SKM'] = ['NILAI (1,00-4,00)', 'KATEGORI']

            for col_num, subheader in enumerate(subheaders['SKM'], start=column_index_from_string('M')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('R')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= SOURCE DATA =======
            data = models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'tempat',
                'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi',
                'keterampilan', 'peserta', 'jumlah_peserta',
                'sinergi_desa', 'sinergi_ibm', 'hasil_skm_nilai', 'hasil_skm_kategori',
                'anggaran', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            # ======= ROW CONFIGURATION =======
            row_num = 6  # Starting row for data entries
            no = 1  # Row number

            for row in data:
                peserta_list = row['peserta'].split(',') if row['peserta'] else ['-']
                peserta_count = row['jumlah_peserta'] or 0

                for peserta in peserta_list:
                    # Satker information
                    sheet.cell(row=row_num, column=1, value=no)
                    sheet.cell(row=row_num, column=2, value=row['satker__nama_satker'])
                    sheet.cell(row=row_num, column=3, value=data.filter(satker_id=row['satker_id']).count())

                    # Pelaksanaan information
                    sheet.cell(row=row_num, column=4, value=row_num - 5)
                    sheet.cell(row=row_num, column=5, value=row['tanggal_awal'])
                    sheet.cell(row=row_num, column=6, value=row['tempat'])

                    # Kawasan yang diintervensi information
                    sheet.cell(row=row_num, column=7, value=f"{row['nama_desa']}, {row['nama_kecamatan']}, {row['nama_kabupaten']}, {row['nama_provinsi']}")

                    # Keterampilan yang diberikan information
                    sheet.cell(row=row_num, column=8, value=row['keterampilan'])

                    # Peserta information
                    sheet.cell(row=row_num, column=9, value=peserta.strip())
                    sheet.cell(row=row_num, column=10, value=peserta_count)

                    # Sinergi information
                    sheet.cell(row=row_num, column=11, value=row['sinergi_desa'])
                    sheet.cell(row=row_num, column=12, value=row['sinergi_ibm'])

                    # Hasil SKM information
                    sheet.cell(row=row_num, column=13, value=row['hasil_skm_nilai'])
                    sheet.cell(row=row_num, column=14, value=row['hasil_skm_kategori'])

                    # Kesimpulan dan tindak lanjut information
                    sheet.cell(row=row_num, column=15, value=row['kesimpulan'])
                    sheet.cell(row=row_num, column=16, value=row['tindak_lanjut'])

                    # Anggaran information
                    sheet.cell(row=row_num, column=17, value=row['anggaran'])

                    # Dokumentasi information
                    sheet.cell(row=row_num, column=18, value=row['dokumentasi'])
                    sheet.cell(row=row_num, column=19, value=row['gambar'])

                    row_num += 1
                no += 1

            # Simpan file Excel
            workbook.save(file_path)

            return Response({
                'success': True,
                'url': f'{base_url}kegiatan/dayatif/bimbingan_teknis_lifeskill/exported/{file_name}.xlsx',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'BIMBINGAN TEKNIS LIFESKILL {tahun}'
        base_path = 'media/kegiatan/dayatif/bimbingan_teknis_lifeskill/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="BIMBINGAN TEKNIS LIFESKILL")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '',
            'KAWASAN YANG DIINTERVENSI', 'KETERAMPILAN YANG DIBERIKAN DALAM PELATIHAN', 'PESERTA', '',
            'SINERGI', '', 'HASIL SKM', '',
            'KESIMPULAN', 'TINDAK LANJUT', 'ANGGARAN', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('O4:O5')
            sheet.merge_cells('P4:P5')
            sheet.merge_cells('Q4:Q5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:F4') # -> Horizontally


            sheet['I4'].value = 'PESERTA'
            sheet.merge_cells('I4:J4') # -> Horizontally

            sheet['K4'].value = 'SINERGI'
            sheet.merge_cells('K4:L4') # -> Horizontally

            sheet['M4'].value = 'HASIL SKM'
            sheet.merge_cells('M4:N4') # -> Horizontally

            sheet['R4'].value = 'DOKUMENTASI'
            sheet.merge_cells('R4:S4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'TEMPAT']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['PESERTA'] = ['DAFTAR', 'JUMLAH']

            for col_num, subheader in enumerate(subheaders['PESERTA'], start=column_index_from_string('I')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['SINERGI'] = ['DESA BERSINAR', 'IBM']

            for col_num, subheader in enumerate(subheaders['SINERGI'], start=column_index_from_string('K')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['SKM'] = ['NILAI (1,00-4,00)', 'KATEGORI']

            for col_num, subheader in enumerate(subheaders['SKM'], start=column_index_from_string('M')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('R')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )


            # ======= SOURCE DATA =======
            data = models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'tempat',
                'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi',
                'keterampilan', 'peserta', 'jumlah_peserta',
                'sinergi_desa', 'sinergi_ibm', 'hasil_skm_nilai', 'hasil_skm_kategori',
                'anggaran', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                peserta = item['peserta']

                formatted_peserta = "\n".join([f"{index + 1}. {item['nama']} - ({item['jenis_kelamin']})" for index, item in enumerate(peserta)])

                lokasi = f"{item['nama_desa']}, {item['nama_kecamatan']}, {item['nama_kabupaten']}, {item['nama_provinsi']}"

                def get_status(data: bool) -> str:
                    return 'Ya' if data else 'Tidak'

                def get_kategori(data: str) -> str:
                    data_map = {
                        'buruk': 'Buruk',
                        'cukup': 'Cukup',
                        'baik': 'Baik',
                        'sangat_baik': 'Sangat Baik',
                    }

                    return data_map[data] or '-'

                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'nama_satker': item['satker__nama_satker'],
                    'tempat': item['tempat'],
                    'lokasi': lokasi,
                    'keterampilan': item['keterampilan'],
                    'peserta': formatted_peserta,
                    'jumlah_peserta': f"{item['jumlah_peserta']} orang",
                    'sinergi_desa': get_status(item['sinergi_desa']),
                    'sinergi_ibm': get_status(item['sinergi_ibm']),
                    'hasil_skm_nilai': f"{item['hasil_skm_nilai']:.2f}",
                    'hasil_skm_kategori': get_kategori(item['hasil_skm_kategori']),
                    'anggaran': item['anggaran'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                    'status': item['status'],
                    'satker_level': satker.level
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('R'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi = f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")'

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('S'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    numering,
                    row['tanggal'],
                    row['tempat'],
                    row['lokasi'],
                    row['keterampilan'],
                    row['peserta'],
                    row['jumlah_peserta'],
                    row['sinergi_desa'],
                    row['sinergi_ibm'],
                    row['hasil_skm_nilai'],
                    row['hasil_skm_kategori'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['anggaran'],
                    dokumentasi,
                    dokumentasi_gambar,
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                    if item in (column_index_from_string('K'), column_index_from_string('I'), column_index_from_string('G')):
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['K'].width = 12 # Sinergi : Desa
            sheet.column_dimensions['I'].width = 30 # Daftar Peserta
            sheet.column_dimensions['G'].width = 75 # Kawasan Yang Diintervensi
            sheet.column_dimensions['M'].width = 12 # Hasil SKM : Nilai
            sheet.column_dimensions['H'].width = 22 # Keterampilan
            sheet.column_dimensions['R'].width = 18 # Dokumentasi
            sheet.column_dimensions['S'].width = 25 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======= MONEV DAYATIF =======
class DAYATIF_MONEV_DAYATIF_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_MONEV_DAYATIF.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_MONEV_DAYATIF_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_MONEV_DAYATIF_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_MONEV_DAYATIF, serializers.DAYATIF_MONEV_DAYATIF_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_MONEV_DAYATIF, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_MONEV_DAYATIF)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_MONEV_DAYATIF, request)

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'REKAPITULASI KEGIATAN MONITORING DAN EVALUASI PROGRAM PEMBERDAYAAN ALTERNATIF {tahun}'
        base_path = 'media/kegiatan/dayatif/monev_dayatif/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="REKAPITULASI KEGIATAN MONITORING DAN EVALUASI PROGRAM PEMBERDAYAAN ALTERNATIF")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'JUMLAH KEGIATAN', 'PELAKSANAAN', '', '',
            'JENIS', 'PERIODE', 'KAWASAN YANG DIINTERVENSI', 'KETERAMPILAN YANG DIBERIKAN DALAM PELATIHAN', 'PESERTA', '',
            'SINERGI', '', 'HASIL SKM', '', 'HASIL INDEKS KEWIRAUSAHAAN', '',
            'STATUS KERAWANAN KAWASAN', '', '',
            'KESIMPULAN', 'TINDAK LANJUT', 'DOKUMENTASI', '']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('V4:V5')
            sheet.merge_cells('W4:W5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'PELAKSANAAN'
            sheet.merge_cells('D4:F4') # -> Horizontally

            sheet['K4'].value = 'PESERTA'
            sheet.merge_cells('K4:L4') # -> Horizontally

            sheet['M4'].value = 'SINERGI'
            sheet.merge_cells('M4:N4') # -> Horizontally

            sheet['O4'].value = 'HASIL SKM'
            sheet.merge_cells('O4:P4') # -> Horizontally

            sheet['Q4'].value = 'HASIL INDEKS KEWIRAUSAHAAN'
            sheet.merge_cells('Q4:R4') # -> Horizontally

            sheet['S4'].value = 'STATUS KERAWANAN KAWASAN'
            sheet.merge_cells('S4:U4') # -> Horizontally

            sheet['X4'].value = 'DOKUMENTASI'
            sheet.merge_cells('X4:Y4') # -> Horizontally

            subheaders = {}

            subheaders['PELAKSANAAN'] = ['NO', 'TANGGAL', 'TEMPAT']

            for col_num, subheader in enumerate(subheaders['PELAKSANAAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['PESERTA'] = ['DAFTAR', 'JUMLAH']

            for col_num, subheader in enumerate(subheaders['PESERTA'], start=column_index_from_string('K')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['SINERGI'] = ['DESA BERSINAR', 'IBM']

            for col_num, subheader in enumerate(subheaders['SINERGI'], start=column_index_from_string('M')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['SKM'] = ['NILAI (1,00-4,00)', 'KATEGORI']

            for col_num, subheader in enumerate(subheaders['SKM'], start=column_index_from_string('O')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['KEWIRAUSAHAAN'] = ['NILAI (1,00-4,00)', 'KATEGORI']

            for col_num, subheader in enumerate(subheaders['KEWIRAUSAHAAN'], start=column_index_from_string('Q')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['KAWASAN'] = ['AWAL', 'AKHIR', 'PULIH']

            for col_num, subheader in enumerate(subheaders['KAWASAN'], start=column_index_from_string('S')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            subheaders['DOKUMENTASI'] = ['FILE', 'GAMBAR']

            for col_num, subheader in enumerate(subheaders['DOKUMENTASI'], start=column_index_from_string('X')):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = subheader
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )


            # ======= SOURCE DATA =======
            data = models.DAYATIF_MONEV_DAYATIF.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'jenis', 'periode', 'tempat',
                'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi',
                'keterampilan', 'peserta', 'jumlah_peserta',
                'sinergi_desa', 'sinergi_ibm', 'hasil_skm_nilai', 'hasil_skm_kategori',
                'hasil_indeks_kewirausahaan_nilai', 'hasil_indeks_kewirausahaan_kategori',
                'status_kerawanan_kawasan_awal_kategori', 'status_kerawanan_kawasan_akhir_nilai', 'status_kerawanan_kawasan_akhir_kategori', 'status_kerawanan_kawasan_kepulihan',
                'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'gambar', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:
                peserta = item['peserta']

                formatted_peserta = "\n".join([f"{index + 1}. {item['nama']} - ({item['jenis_kelamin']})" for index, item in enumerate(peserta)])

                lokasi = f"{item['nama_desa']}, {item['nama_kecamatan']}, {item['nama_kabupaten']}, {item['nama_provinsi']}"

                def get_status(data: bool) -> str:
                    return 'Ya' if data else 'Tidak'

                def get_kategori(data: str) -> str:
                    data_map = {
                        'buruk': 'Buruk',
                        'cukup': 'Cukup',
                        'baik': 'Baik',
                        'sangat_baik': 'Sangat Baik',
                    }

                    return data_map[data] or '-'

                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'tanggal': get_tanggal_kegiatan(item['tanggal_awal'], item['tanggal_akhir']),
                    'nama_satker': item['satker__nama_satker'],
                    'jenis': item['jenis'],
                    'periode': item['periode'],
                    'tempat': item['tempat'],
                    'lokasi': lokasi,
                    'keterampilan': item['keterampilan'],
                    'peserta': formatted_peserta,
                    'jumlah_peserta': f"{item['jumlah_peserta']} orang",
                    'sinergi_desa': get_status(item['sinergi_desa']),
                    'sinergi_ibm': get_status(item['sinergi_ibm']),
                    'hasil_skm_nilai': f"{item['hasil_skm_nilai']:.2f}",
                    'hasil_skm_kategori': get_kategori(item['hasil_skm_kategori']),
                    'hasil_indeks_kewirausahaan_nilai': item['hasil_indeks_kewirausahaan_nilai'],
                    'hasil_indeks_kewirausahaan_kategori': item['hasil_indeks_kewirausahaan_kategori'],

                    'status_kerawanan_kawasan_awal_kategori': item['status_kerawanan_kawasan_awal_kategori'],
                    'status_kerawanan_kawasan_akhir_nilai': item['status_kerawanan_kawasan_akhir_nilai'],
                    'status_kerawanan_kawasan_akhir_kategori': item['status_kerawanan_kawasan_akhir_kategori'],
                    'status_kerawanan_kawasan_kepulihan': item['status_kerawanan_kawasan_kepulihan'],

                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'gambar': item['gambar'],
                    'status': item['status'],
                    'satker_level': satker.level
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('X'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi = f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")'

                dokumentasi_gambar = 'Tidak ada'

                if row['gambar']:
                    dokumentasi_gambar = f'=HYPERLINK("{base_url + row["gambar"]}","Link dokumentasi gambar")'
                    cell_dokumentasi_gambar = sheet.cell(row=current_row, column=column_index_from_string('Y'), value=current_group)
                    cell_dokumentasi_gambar.font = Font(color="0000FF")

                status_kerawanan_kawasan_akhir = f"{row['status_kerawanan_kawasan_akhir_nilai']} ({row['status_kerawanan_kawasan_akhir_kategori']})"

                data_row = [
                    group_count,
                    row['nama_satker'],
                    f'{rep_count} kali',
                    numering,
                    row['tanggal'],
                    row['jenis'],
                    row['periode'],
                    row['tempat'],
                    row['lokasi'],
                    row['keterampilan'],
                    row['peserta'],
                    row['jumlah_peserta'],
                    row['sinergi_desa'],
                    row['sinergi_ibm'],
                    row['hasil_skm_nilai'],
                    row['hasil_skm_kategori'],
                    row['hasil_indeks_kewirausahaan_nilai'],
                    row['hasil_indeks_kewirausahaan_kategori'],
                    row['status_kerawanan_kawasan_awal_kategori'],
                    status_kerawanan_kawasan_akhir,
                    row['status_kerawanan_kawasan_kepulihan'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    dokumentasi,
                    dokumentasi_gambar,
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                    if item in (column_index_from_string('K'), column_index_from_string('I'), column_index_from_string('G')):
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            # sheet.column_dimensions['K'].width = 12 # Sinergi : Desa
            # sheet.column_dimensions['I'].width = 30 # Daftar Peserta
            # sheet.column_dimensions['G'].width = 75 # Kawasan Yang Diintervensi
            # sheet.column_dimensions['M'].width = 12 # Hasil SKM : Nilai
            # sheet.column_dimensions['H'].width = 22 # Keterampilan
            sheet.column_dimensions['X'].width = 18 # Dokumentasi
            sheet.column_dimensions['Y'].width = 25 # Dokumentasi Gambar

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


# ======= DUKUNGAN_STAKEHOLDER =======
class DAYATIF_DUKUNGAN_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_DUKUNGAN_STAKEHOLDER.objects.all()
    serializer_class = serializers.DAYATIF_DUKUNGAN_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list_dukungan(request, self.queryset, serializers.DAYATIF_DUKUNGAN_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_DUKUNGAN_STAKEHOLDER, serializers.DAYATIF_DUKUNGAN_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_DUKUNGAN_STAKEHOLDER, pk)

    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_DUKUNGAN_STAKEHOLDER)

    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_DUKUNGAN_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_DUKUNGAN_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'stakeholder', 'jumlah_peserta', 'jenis', 'bentuk', 'jumlah',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'jumlah_sasaran', 'pengaruh', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)

        print('Panjang data sebelum filter:', len(data))

        data = get_filtered_data(satker, data, status, waktu)

        print('Panjang data sesudah filter:', len(data))

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'stakeholder': item['stakeholder'],
                'jumlah_peserta': item['jumlah_peserta'],
                'jenis': item['jenis'],
                'bentuk': item['bentuk'],
                'jumlah': item['jumlah'],

                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],

                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'jumlah_sasaran': item['jumlah_sasaran'],

                'pengaruh': item['pengaruh'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)

        return serialized_data

    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        base_url = self.request.build_absolute_uri('/') + 'media/'
        tahun = datetime.datetime.now().year

        file_name = f'DUKUNGAN STAKEHOLDER {tahun}'
        base_path = 'media/kegiatan/dayatif/dukungan_stakeholder/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            # Menentukan warna latar belakang dan warna teks
            bg_color = PatternFill(start_color="2ca8dc", end_color="2ca8dc", fill_type="solid")
            text_color = Font(color="FFFFFF")

            # Mengatur judul di baris pertama
            cell = sheet.cell(row=1, column=1, value="DUKUNGAN STAKEHOLDER")
            sheet.merge_cells('A1:R1')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Mengatur tahun di baris kedua
            cell = sheet.cell(row=2, column=1, value=f"TAHUN {tahun}")
            sheet.merge_cells('A2:R2')
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Menambahkan header utama
            headers = ['NO', 'SATUAN KERJA', 'NAMA STAKEHOLDER', 'DUKUNGAN YANG DIBERIKAN', '', '',
            'KAWASAN YANG DIINTERVENSI', 'JUMLAH SASARAN', 'PENGARUH/MANFAAT', 'KESIMPULAN', 'TINDAK LANJUT', 'DOKUMENTASI']

            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_num)
                cell.value = header
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )

            # ======= MERGE HEADERS CELLS =======
            sheet.merge_cells('A4:A5')
            sheet.merge_cells('B4:B5')
            sheet.merge_cells('C4:C5')
            sheet.merge_cells('G4:G5')
            sheet.merge_cells('H4:H5')
            sheet.merge_cells('I4:I5')
            sheet.merge_cells('J4:J5')
            sheet.merge_cells('K4:K5')
            sheet.merge_cells('L4:L5')
            sheet.merge_cells('M4:M5')

            # ======= MERGED CELLS CONFIGURATION =======
            sheet['D4'].value = 'DUKUNGAN YANG DIBERIKAN'
            sheet.merge_cells('D4:F4') # -> Horizontally

            subheaders = {}

            subheaders['DUKUNGAN'] = ['JENIS', 'BENTUK', 'JUMLAH']

            for col_num, subheader in enumerate(subheaders['DUKUNGAN'], start=column_index_from_string('D')):
                cell = sheet.cell(row=5, column=col_num, value=subheader)
                cell.fill = bg_color
                cell.font = text_color
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # ======= SOURCE DATA =======
            data = models.DAYATIF_DUKUNGAN_STAKEHOLDER.objects.values(
                'id', 'satker_id', 'satker__nama_satker',
                'stakeholder', 'jenis', 'bentuk', 'jumlah',
                'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi',
                'jumlah_sasaran', 'pengaruh', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
            ).order_by(
                '-satker_id'
            )

            # ======= DATA FILTERING =======
            satker = self.request.user.profile.satker

            if satker.level == 1:
                data = data.filter(satker_id=satker)
            elif satker.level == 0:
                data = data.filter(satker__provinsi_id=satker.provinsi, status__gt=0)
            elif satker.level == 2:
                data = data.filter(status=2)

            serialized_data = []

            for item in data:

                lokasi = f"{item['nama_desa']}, {item['nama_kecamatan']}, {item['nama_kabupaten']}, {item['nama_provinsi']}"

                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'satker_level': satker.level,
                    'status': item['status'],
                    'stakeholder': item['stakeholder'],
                    'jenis': item['jenis'],
                    'bentuk': item['bentuk'],
                    'jumlah': item['jumlah'],
                    'lokasi': lokasi,
                    'jumlah_sasaran': item['jumlah_sasaran'],
                    'pengaruh': item['pengaruh'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                }
                serialized_data.append(serialized_item)

            # Jumlah kegiatan
            rep_counts = {}
            for row in serialized_data:
                satker_name = row['nama_satker']
                if satker_name not in rep_counts:
                    rep_counts[satker_name] = 0
                rep_counts[satker_name] += 1

            current_row = 6
            numering = 0 # Inside numbering
            group_count = 0
            rep_count = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                        sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    group_count += 1
                    numering = 1

                rep_count = rep_counts[current_group]

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # ======= CELL STYLING =======
                cell_dokumentasi = sheet.cell(row=current_row, column=column_index_from_string('L'), value=current_group)
                cell_dokumentasi.font = Font(color="0000FF")

                dokumentasi = f'=HYPERLINK("{base_url + row["dokumentasi"]}","Link dokumentasi")'

                data_row = [
                    group_count,
                    row['nama_satker'],
                    row['stakeholder'],
                    row['jenis'],
                    row['bentuk'],
                    row['jumlah'],
                    row['lokasi'],
                    row['jumlah_sasaran'],
                    row['pengaruh'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    dokumentasi,
                ]

                for item, value in enumerate(data_row[:column_index_from_string('Z')], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                    if item in (column_index_from_string('K'), column_index_from_string('I'), column_index_from_string('J')):
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                current_row += 1
                numering += 1

            # Gabungkan sel-sel untuk grup terakhir jika diperlukan
            # Pastikan untuk melakukan merge terakhir setelah loop
            if current_group is not None:
                end_merge_row = current_row - 1
                sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
                sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')

            # ======= AUTOFIT COLUMNS =======
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column  # Get the column number (1-indexed)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['A'].width = 8 # No
            sheet.column_dimensions['J'].width = 40 # Deskripsi Hasil
            sheet.column_dimensions['L'].width = 18 # Dokumentais

            # ======= SAVE FILE =======
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file': f'/{file_path}'
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
