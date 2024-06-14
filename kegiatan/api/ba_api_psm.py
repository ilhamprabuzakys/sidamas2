from django.db.models import F, Count, Q, Subquery, OuterRef
from django.shortcuts import get_object_or_404
from collections import defaultdict
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from users.models import Profile, Satker

import re
import os
import shutil
import codecs
import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import json

from sidamas import pagination

from kegiatan import models
from kegiatan.api import serializers
from kegiatan.api import filters
from users.models import Satker

def get_kegiatan_satker_status(request):
    satker = request.user.profile.satker
    
    """
        Mapping :
        Satker Level       Status
        0 : BNNP        -> 2 : Pusat
        1 : BNNK        -> 1 : BNNP
        2 : Pusat       -> 2 : Pusat
    """

    status_map = {
        0: 1,
        1: 0,
        2: 2
    }
    
    status = status_map.get(satker.level, None)
    
    return status

def get_tanggal_kegiatan(self, tanggal_awal, tanggal_akhir=None):
    start = datetime.date.fromisoformat(f'{tanggal_awal}')
    start_date = start.strftime('%-d')
    start_month = start.strftime('%B')
    start_year = start.strftime('%Y')
    
    nama_bulan = {
        "January": "Januari",
        "February": "Februari",
        "March": "Maret",
        "April": "April",
        "May": "Mei",
        "June": "Juni",
        "July": "Juli",
        "August": "Agustus",
        "September": "September",
        "October": "Oktober",
        "November": "November",
        "December": "Desember"
    }

    if tanggal_akhir:
        end = datetime.date.fromisoformat(f'{tanggal_akhir}')
        end_date = end.strftime('%-d')
        end_month = end.strftime('%B')
        end_year = end.strftime('%Y')

        if start_month == end_month and start_year == end_year:
            return f"{start_date} - {end_date} {nama_bulan[start_month]} {start_year}"
        elif start_year != end_year:
            return f"{start.strftime('%-d')} {nama_bulan[start_month]} {start_year} - {end.strftime('%-d')} {nama_bulan[end_month]} {end_year}"
        else:
            return f"{start.strftime('%-d')} {nama_bulan[start_month]} - {end.strftime('%-d')} {nama_bulan[end_month]} {end_year}"
    else:
        return f"{start.strftime('%-d')} {nama_bulan[start_month]} {start_year}"

# ======= PSM RAKERNIS API =======
    
class PSM_RAKERNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_RAKERNIS.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_RAKERNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_RAKERNIS_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_RAKERNIS.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_RAKERNIS_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_RAKERNIS.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_RAKERNIS.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI RAPAT KERJA TEKNIS {tahun}'
        base_path = 'media/kegiatan/psm/rakernis/exported'
        file_path = f'{base_path}/{file_name}.xlsx'
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'
        
        try:
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_RAKERNIS.objects.values(
                'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
            ).order_by(
                'satker__nama_satker'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)
        
            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'satker_target': item['satker_target'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker_target__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            # tb_header
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'NO. KEGIATAN',
                'TANGGAL',
                'SATUAN KERJA YANG DIUNDANG',
                'DESKRIPSI',
                'KENDALA',
                'KESIMPULAN',
                'REKOMENDASI TINDAK LANJUT'
            ]

            for item, header in enumerate(headers[:9], start=1):
                cell = sheet.cell(row=1, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            current_row = 2
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center')

                data_row = [
                    group_count,
                    row['nama_satker'],
                    numering,
                    row['tanggal_awal'],
                    row['nama_satker_target'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut']
                ]

                for item, value in enumerate(data_row[:9], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1
                numering += 1

            # Merge cell untuk grup terakhir
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_RAKERNIS_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_RAKERNIS.objects.get(id=request.data['id'])
        except models.PSM_RAKERNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_RAKERNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_RAKERNIS.objects.get(id=id)
        except models.PSM_RAKERNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_RAKERNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
    
    def get_view_name(self):
        return "PSM RAKERNIS"

# ======= PSM BINAAN TEKNIS API =======
class PSM_BINAAN_TEKNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_BINAAN_TEKNIS.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_BINAAN_TEKNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_BINAAN_TEKNIS_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset
    
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_BINAAN_TEKNIS.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_BINAAN_TEKNIS_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_BINAAN_TEKNIS.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_BINAAN_TEKNIS.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_BINAAN_TEKNIS_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_BINAAN_TEKNIS.objects.get(id=request.data['id'])
        except models.PSM_BINAAN_TEKNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_BINAAN_TEKNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_BINAAN_TEKNIS.objects.get(id=id)
        except models.PSM_BINAAN_TEKNIS.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_BINAAN_TEKNIS_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI PEMBINAAN TEKNIS {tahun}'
        base_path = 'media/kegiatan/psm/bintek/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'

        try:
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_BINAAN_TEKNIS.objects.values(
                'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
            ).order_by(
                'satker__nama_satker'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)

            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'satker_target': item['satker_target'],
                    'nama_satker': item['satker__nama_satker'],
                    'nama_satker_target': item['satker_target__nama_satker'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)
            
            # tb_header
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'NO. KEGIATAN',
                'TANGGAL',
                'SATUAN KERJA YANG DIBINTEK',
                'DESKRIPSI',
                'KENDALA',
                'KESIMPULAN',
                'REKOMENDASI TINDAK LANJUT'
            ]
            
            for item, header in enumerate(headers[:9], start=1):
                    cell = sheet.cell(row=1, column=item, value=header)
                    cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            current_row = 2
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center')

                data_row = [
                    group_count,
                    row['nama_satker'],
                    numering,
                    row['tanggal_awal'],
                    row['nama_satker_target'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut']
                ]

                for item, value in enumerate(data_row[:9], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1
                numering += 1

            # Merge cell untuk grup terakhir
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)

            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    def get_view_name(self):
        return "PSM BINTEK"

# ======= PSM ASISTENSI API =======
class PSM_ASISTENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_ASISTENSI.objects.all().filter(satker__parent__isnull=True).order_by('satker__id','-tanggal_awal', 'satker__order').distinct('satker__id')
    serializer_class = serializers.PSM_ASISTENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_ASISTENSI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        
        if satker_level == 1:
            # bnnk
            queryset = queryset.filter(satker__id=satker)
        elif satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        elif satker_level == 2:
            # pusat
            queryset = queryset.filter(status=2)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_ASISTENSI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_ASISTENSI_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_ASISTENSI.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
            'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_ASISTENSI.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN ASISTENSI KOTAN{tahun}'
        base_path = 'media/kegiatan/psm/asistensi/exported'
        file_path = f'{base_path}/{file_name}.xlsx'
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'
        
        try:
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_ASISTENSI.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
                'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
            ).order_by(
                'satker__nama_satker'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)
        
            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            # tb_header
            headers = [
                'NO.',
                'SATUAN KERJA',
                'JUMLAH KEGIATAN (KALI)',  # Menyesuaikan headers dengan penambahan kolom 'SATKER ID'
                'PELAKSANAAN',
                'NO',
                'TANGGAL',
                'JML PESERTA (ORG)',
                'STAKEHOLDER YANG DIASISTENSI DALAM RANGKA KOTAN',
                'DESKRIPSI HASIL',
                'KENDALA/HAMBATAN',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI',
            ]

            for item, header in enumerate(headers[:9], start=1):
                cell = sheet.cell(row=1, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            current_row = 2
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center')

                data_row = [
                    group_count,
                    row['nama_satker'],
                    numering,
                    row['jumlah_kegiatan'],
                    row['tanggal'],
                    row['jumlah_peserta'],
                    row['stakeholder'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['dokumentasi'],
                ]

                for item, value in enumerate(data_row[:9], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1
                numering += 1

            # Merge cell untuk grup terakhir
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_ASISTENSI_CREATE_UPDATE_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_ASISTENSI.objects.get(id=request.data['id'])
        except models.PSM_ASISTENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_ASISTENSI_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_ASISTENSI.objects.get(id=id)
        except models.PSM_ASISTENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_ASISTENSI_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM ASISTENSI"

# ======= PSM SINKRONISASI KEBIJAKAN API =======
class PSM_SINKRONISASI_KEBIJAKAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_SINKRONISASI_KEBIJAKAN.objects.all().filter(satker__parent__isnull=True).order_by('satker__id','-tanggal_awal', 'satker__order').distinct('satker__id')
    serializer_class = serializers.PSM_SINKRONISASI_KEBIJAKAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_SINKRONISASI_KEBIJAKAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        
        if satker_level == 1:
            # bnnk
            queryset = queryset.filter(satker__id=satker)
        elif satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        elif satker_level == 2:
            # pusat
            queryset = queryset.filter(status=2)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_SINKRONISASI_KEBIJAKAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_SINKRONISASI_KEBIJAKAN_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_SINKRONISASI_KEBIJAKAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
            'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_SINKRONISASI_KEBIJAKAN.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN SINKRONISASI KEBIJAKAN{tahun}'
        base_path = 'media/kegiatan/psm/sinkronisasi_kebijakan/exported'
        file_path = f'{base_path}/{file_name}.xlsx'
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'
        
        try:
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_SINKRONISASI_KEBIJAKAN.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
                'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
            ).order_by(
                'satker__nama_satker'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)
        
            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            # tb_header
            headers = [
                'NO.',
                'SATUAN KERJA',
                'JUMLAH KEGIATAN (KALI)',  # Menyesuaikan headers dengan penambahan kolom 'SATKER ID'
                'PELAKSANAAN',
                'NO',
                'TANGGAL',
                'JML PESERTA (ORG)',
                'STAKEHOLDER YANG DIASISTENSI DALAM RANGKA KOTAN',
                'DESKRIPSI HASIL',
                'KENDALA/HAMBATAN',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI',
            ]

            for item, header in enumerate(headers[:9], start=1):
                cell = sheet.cell(row=1, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            current_row = 2
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center')

                data_row = [
                    group_count,
                    row['nama_satker'],
                    numering,
                    row['jumlah_kegiatan'],
                    row['tanggal'],
                    row['jumlah_peserta'],
                    row['stakeholder'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['dokumentasi'],
                ]

                for item, value in enumerate(data_row[:9], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1
                numering += 1

            # Merge cell untuk grup terakhir
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_SINKRONISASI_KEBIJAKAN_CREATE_UPDATE_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_SINKRONISASI_KEBIJAKAN.objects.get(id=request.data['id'])
        except models.PSM_SINKRONISASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_SINKRONISASI_KEBIJAKAN_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_SINKRONISASI_KEBIJAKAN.objects.get(id=id)
        except models.PSM_SINKRONISASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_SINKRONISASI_KEBIJAKAN_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM SINKRONISASI KEBIJAKAN"

# ======= PSM WORKSHOP TEMATIK API =======
class PSM_WORKSHOP_TEMATIK_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_WORKSHOP_TEMATIK.objects.all().filter(satker__parent__isnull=True).order_by('satker__id','-tanggal_awal', 'satker__order').distinct('satker__id')
    serializer_class = serializers.PSM_WORKSHOP_TEMATIK_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_WORKSHOP_TEMATIK_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        
        if satker_level == 1:
            # bnnk
            queryset = queryset.filter(satker__id=satker)
        elif satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        elif satker_level == 2:
            # pusat
            queryset = queryset.filter(status=2)
        return queryset

    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_WORKSHOP_TEMATIK.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_WORKSHOP_TEMATIK_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # untuk edit
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_WORKSHOP_TEMATIK.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
            'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_WORKSHOP_TEMATIK.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'KEGIATAN WORKSHOP TEMATIK{tahun}'
        base_path = 'media/kegiatan/psm/workshop_tematik/exported'
        file_path = f'{base_path}/{file_name}.xlsx'
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'
        
        try:
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            data = models.PSM_WORKSHOP_TEMATIK.objects.values(
                'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'jumlah_kegiatan',
                'jumlah_peserta', 'stakeholder', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
            ).order_by(
                'satker__nama_satker'
            )

            if satker_level == 1:
                data = data.filter(satker_id=satker)
            elif satker_level == 0:
                data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
            elif satker_level == 2:
                data = data.filter(status=2)
        
            serialized_data = []
            for item in data:
                serialized_item = {
                    'id': item['id'],
                    'satker_id': item['satker_id'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_peserta': item['jumlah_peserta'],
                    'stakeholder': item['stakeholder'],
                    'tanggal_awal': item['tanggal_awal'],
                    'tanggal_akhir': item['tanggal_akhir'],
                    'deskripsi': item['deskripsi'],
                    'kendala': item['kendala'],
                    'kesimpulan': item['kesimpulan'],
                    'tindak_lanjut': item['tindak_lanjut'],
                    'dokumentasi': item['dokumentasi'],
                    'status': item['status'],
                    'satker_level': satker_level
                }
                serialized_data.append(serialized_item)

            # tb_header
            headers = [
                'NO.',
                'SATUAN KERJA',
                'JUMLAH KEGIATAN (KALI)',  # Menyesuaikan headers dengan penambahan kolom 'SATKER ID'
                'PELAKSANAAN',
                'NO',
                'TANGGAL',
                'JML PESERTA (ORG)',
                'STAKEHOLDER YANG DIASISTENSI DALAM RANGKA KOTAN',
                'DESKRIPSI HASIL',
                'KENDALA/HAMBATAN',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI',
            ]

            for item, header in enumerate(headers[:9], start=1):
                cell = sheet.cell(row=1, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            current_row = 2
            numering = 0
            group_count = 0
            current_group = None

            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    numering = 1  # Atur ulang nomor ketika grup berubah
                    group_count += 1

                group_cell = sheet.cell(row=current_row, column=2, value=current_group)
                group_cell.font = openpyxl.styles.Font(bold=True)
                group_cell.alignment = Alignment(horizontal='center')

                data_row = [
                    group_count,
                    row['nama_satker'],
                    numering,
                    row['jumlah_kegiatan'],
                    row['tanggal'],
                    row['jumlah_peserta'],
                    row['stakeholder'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    row['dokumentasi'],
                ]

                for item, value in enumerate(data_row[:9], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1
                numering += 1

            # Merge cell untuk grup terakhir
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')

            # Autofit kolom
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # save file
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_WORKSHOP_TEMATIK_CREATE_UPDATE_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_WORKSHOP_TEMATIK.objects.get(id=request.data['id'])
        except models.PSM_WORKSHOP_TEMATIK.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_WORKSHOP_TEMATIK_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_WORKSHOP_TEMATIK.objects.get(id=id)
        except models.PSM_WORKSHOP_TEMATIK.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_WORKSHOP_TEMATIK_CREATE_UPDATE_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)

    def get_view_name(self):
        return "PSM WORKSHOP TEMATIK"
    
# ======= PSM TES URIN DETEKSI DINI API =======
class PSM_TES_URINE_DETEKSI_DINI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_TES_URINE_DETEKSI_DINI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_TES_URINE_DETEKSI_DINI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_TES_URINE_DETEKSI_DINI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
        
    @action(detail=True, methods=['delete'])
    def delete_kegiatan_peserta(self, request, pk=None):
        # hapus data kegiatan sama seluruh peserta yang terkoneksi
        try:
            deleted_peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(pk=pk).delete()
            return Response({
                'message': 'Berhasil menghapus data',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}',
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)
        
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)
            
            satker_parent = {}
            
            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        satker = self.request.user.profile.satker
       
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI {satker.nama_satker.upper()} TAHUN {tahun}' if satker.level < 2 else f'REKAPITULASI PEMBINAAN TEKNIS BNNK & BNNP TAHUN {tahun}'
        base_path = 'media/kegiatan/binaan_teknis/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        base_url = self.request.build_absolute_uri('/')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'Data Kegiatan   '

        try:
            serialized_data = self.get_flat_values(request)

            # ======= HEADERS =======
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'JUMLAH KEGIATAN',
                'STATUS',
                'NO.',
                'TANGGAL',
                'NAMA LINGKUNGAN',
                'PRIA',
                'WANITA',
                'JUMLAH PESERTA',
                'DAFTAR NAMA PESERTA',
                'HASIL TEST URINE',
                'TINDAK LANJUT',
                'DOKUMENTASI'
            ]
            
            # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

            current_row = 4
            
            # ======= GENERATE HEADERS =======
            for item, header in enumerate(headers[:len(headers)], start=1):
                cell = sheet.cell(row=current_row, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            current_row += 1
            
            no = 0
            no_child = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    no += 1
                
                no_child += 1
                
                formatted_data = {}
                
                # ======= STATUS =======
                status_mapping = {
                    0: 'Belum dikirim',
                    1: 'Dikirim ke BNNP',
                    2: 'Dikirim ke BNN Pusat'
                }
                
                formatted_data['status'] = status_mapping.get(row['status'], '-')
                
                formatted_data['jumlah_kegiatan'] = len(row)
                # formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])

                formatted_data['pria'] = self.count_participants(row['peserta'], 'Laki-Laki')
                formatted_data['wanita'] = self.count_participants(row['peserta'], 'Perempuan')
                # print(row['peserta'])

                formatted_data['jumlah_peserta'] = len(row['peserta'])
                formatted_data['daftar_peserta'] = self.format_addresses(row['peserta'])
                formatted_data['hasil_test'] = self.count_hasil_test(row['peserta'])
                
                for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
                    celcol = sheet.cell(row=current_row, column=col, value=current_group)
                    celcol.font = openpyxl.styles.Font(bold=True)
                    celcol.alignment = Alignment(horizontal='center', vertical='center')
                    
                cell_e = sheet.cell(row=current_row, column=5, value=current_group)
                cell_e.font = openpyxl.styles.Font(bold=True)
                cell_e.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_g = sheet.cell(row=current_row, column=7, value=current_group)
                cell_g.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_h = sheet.cell(row=current_row, column=8, value=current_group)
                cell_h.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_i = sheet.cell(row=current_row, column=9, value=current_group)
                cell_i.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_n = sheet.cell(row=current_row, column=14, value=current_group)
                cell_n.alignment = Alignment(horizontal='center', vertical='center')

                data_row = [
                    no,
                    row['nama_satker'],
                    formatted_data['jumlah_kegiatan'],
                    formatted_data['status'],
                    no_child,
                    row['tanggal_awal'],
                    row['nama_lingkungan'],
                    formatted_data['pria'],
                    formatted_data['wanita'],
                    formatted_data['jumlah_peserta'],
                    formatted_data['daftar_peserta'],
                    # row[0]['nama_peserta'],
                    formatted_data['hasil_test'],
                    row['tindak_lanjut'],
                    # row['dokumentasi'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
                ]

                for item, value in enumerate(data_row[:len(headers)], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1

            # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
            sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
            sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

            # ======= AUTOFIT KOLOM =======
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.3
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
            
            sheet.column_dimensions['N'].width = 40

            # ======= SAVE FILE =======
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': serialized_data,
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:
                gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                status_mapping = {'P': 'Positif', 'L': 'Negatif'}
                
                jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                hasil_test = status_mapping.get(entry['status'], 'Negatif')

                parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=inserted_id)
                
                peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                    jenis_kelamin=jenis_kelamin,
                    hasil_test=hasil_test,
                    alamat=entry['alamat']
                )

    def get_view_name(self):
        return "PSM TEST URINE"
    
    def get_data_peserta(self, parent):
        data = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'hasil_test', 'alamat'
        ).order_by('-id')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'hasil_test': item['hasil_test'],
                'alamat': item['alamat']
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    def process_peserta_array(array_json, current_record_id):
        if array_json is not None:
            data = json.loads(array_json)
            gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
            status_mapping = {'P': 'Positif', 'L': 'Negatif'}
            parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=current_record_id)

            for entry in data:
                try:
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat']
                    )
                except KeyError:
                    # Handle missing keys in entry
                    pass  # Or log the error

    def get_flat_values(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_TES_URINE_DETEKSI_DINI_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # helper
    def count_participants(self, data, gender):
        count = 0
        for row in data:
            if row['jenis_kelamin'] == gender:
                count += 1
        return count

    def format_addresses(self, objects):
        print(objects)
        result = ""
        for obj in objects:
            result += f"{obj['nama_peserta']} - {obj['alamat']}, "
        # Remove the trailing comma and space
        result = result.rstrip(", ")
        return result
    
    def count_hasil_test(self, data):
        count_positif = 0
        count_negatif = 0

        for row in data:
            if row['hasil_test'] == "Positif":
                count_positif += 1
            elif row['hasil_test'] == "Negatif":
                count_negatif += 1

        return "Positif - {} Negatif - {}".format(count_positif, count_negatif)
    
    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return serializers.PSM_TES_URINE_DETEKSI_DINI_CREATE_UPDATE_Serializer
        return super().get_serializer_class()

class PSM_TES_URNIE_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_TES_URINE_DETEKSI_DINI.objects.all()
    serializer_class = serializers.PSM_TES_URINE_DETEKSI_DINI_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_update(self, serializer):
        current_record_id = serializer.instance.pk

        serializer.save()

        models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent_id=current_record_id).delete()

        peserta_array_lama_json = self.request.data.get('peserta_lamaArray', None)
        if peserta_array_lama_json:
            data = json.loads(peserta_array_lama_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                    status_mapping = {'P': 'Positif', 'L': 'Negatif'}
                    
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=current_record_id)
                    
                    peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat']
                    )

        peserta_array_json = self.request.data.get('pesertaArray', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                    status_mapping = {'P': 'Positif', 'L': 'Negatif'}
                    
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    parent_instance = models.PSM_TES_URINE_DETEKSI_DINI.objects.get(pk=current_record_id)
                    
                    peserta = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat']
                    )

        serializer.save(updated_by=self.request.user)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_TES_URINE_DETEKSI_DINI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)


    def get_data_peserta(self, parent):
        data = models.PSM_TES_URINE_DETEKSI_DINI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'hasil_test', 'alamat'
        ).order_by('-id')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'hasil_test': item['hasil_test'],
                'alamat': item['alamat']
            }
            serialized_data.append(serialized_item)

        return serialized_data

# ======= PSM MONITORING DAN EVALUASI SUPERVISI KEGIATAN KOTAN API =======
class PSM_MONITORING_DAN_EVALUASI_SUPERVISI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
        
    @action(detail=True, methods=['delete'])
    def delete_kegiatan_peserta(self, request, pk=None):
        # hapus data kegiatan sama seluruh peserta yang terkoneksi
        try:
            deleted_peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).delete()
            return Response({
                'message': 'Berhasil menghapus data',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}',
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)
        
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)
            
            satker_parent = {}
            
            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def export_data(self, request):
        satker = self.request.user.profile.satker
       
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI {satker.nama_satker.upper()} TAHUN {tahun}' if satker.level < 2 else f'REKAPITULASI PEMBINAAN TEKNIS BNNK & BNNP TAHUN {tahun}'
        base_path = 'media/kegiatan/binaan_teknis/exported'
        file_path = f'{base_path}/{file_name}.xlsx'

        base_url = self.request.build_absolute_uri('/')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'Data Kegiatan   '

        try:
            serialized_data = self.get_flat_values(request)

            # ======= HEADERS =======
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'JUMLAH KEGIATAN',
                'STATUS',
                'NO.',
                'TANGGAL',
                'NAMA LINGKUNGAN',
                'PRIA',
                'WANITA',
                'JUMLAH PESERTA',
                'DAFTAR NAMA PESERTA',
                'HASIL TEST URINE',
                'TINDAK LANJUT',
                'DOKUMENTASI'
            ]
            
            # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

            current_row = 4
            
            # ======= GENERATE HEADERS =======
            for item, header in enumerate(headers[:len(headers)], start=1):
                cell = sheet.cell(row=current_row, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            current_row += 1
            
            no = 0
            no_child = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    no += 1
                
                no_child += 1
                
                formatted_data = {}
                
                # ======= STATUS =======
                status_mapping = {
                    0: 'Belum dikirim',
                    1: 'Dikirim ke BNNP',
                    2: 'Dikirim ke BNN Pusat'
                }
                
                formatted_data['status'] = status_mapping.get(row['status'], '-')
                
                formatted_data['jumlah_kegiatan'] = len(row)
                # formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])

                formatted_data['pria'] = self.count_participants(row['peserta'], 'Laki-Laki')
                formatted_data['wanita'] = self.count_participants(row['peserta'], 'Perempuan')
                # print(row['peserta'])

                formatted_data['jumlah_peserta'] = len(row['peserta'])
                formatted_data['daftar_peserta'] = self.format_addresses(row['peserta'])
                formatted_data['hasil_test'] = self.count_hasil_test(row['peserta'])
                
                for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
                    celcol = sheet.cell(row=current_row, column=col, value=current_group)
                    celcol.font = openpyxl.styles.Font(bold=True)
                    celcol.alignment = Alignment(horizontal='center', vertical='center')
                    
                cell_e = sheet.cell(row=current_row, column=5, value=current_group)
                cell_e.font = openpyxl.styles.Font(bold=True)
                cell_e.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_g = sheet.cell(row=current_row, column=7, value=current_group)
                cell_g.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_h = sheet.cell(row=current_row, column=8, value=current_group)
                cell_h.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_i = sheet.cell(row=current_row, column=9, value=current_group)
                cell_i.alignment = Alignment(horizontal='center', vertical='center')
                
                cell_n = sheet.cell(row=current_row, column=14, value=current_group)
                cell_n.alignment = Alignment(horizontal='center', vertical='center')

                data_row = [
                    no,
                    row['nama_satker'],
                    formatted_data['jumlah_kegiatan'],
                    formatted_data['status'],
                    no_child,
                    row['tanggal_awal'],
                    row['nama_lingkungan'],
                    formatted_data['pria'],
                    formatted_data['wanita'],
                    formatted_data['jumlah_peserta'],
                    formatted_data['daftar_peserta'],
                    # row[0]['nama_peserta'],
                    formatted_data['hasil_test'],
                    row['tindak_lanjut'],
                    # row['dokumentasi'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
                ]

                for item, value in enumerate(data_row[:len(headers)], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1

            # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
            sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
            sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

            # ======= AUTOFIT KOLOM =======
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.3
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
            
            sheet.column_dimensions['N'].width = 40

            # ======= SAVE FILE =======
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': serialized_data,
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:
                gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                status_mapping = {'P': 'Positif', 'L': 'Negatif'}
                
                jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                hasil_test = status_mapping.get(entry['status'], 'Negatif')

                parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=inserted_id)
                
                peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                    jenis_kelamin=jenis_kelamin,
                    hasil_test=hasil_test,
                    alamat=entry['alamat']
                )

    def get_view_name(self):
        return "PSM MONITORING DAN EVALUASI SUPERVISI KEGIATAN KOTAN"
    
    def get_data_peserta(self, parent):
        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'hasil_test', 'alamat'
        ).order_by('-id')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'hasil_test': item['hasil_test'],
                'alamat': item['alamat']
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    def process_peserta_array(array_json, current_record_id):
        if array_json is not None:
            data = json.loads(array_json)
            gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
            status_mapping = {'P': 'Positif', 'L': 'Negatif'}
            parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=current_record_id)

            for entry in data:
                try:
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')
                    hasil_test = status_mapping.get(entry['status'], 'Negatif')

                    peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        hasil_test=hasil_test,
                        alamat=entry['alamat']
                    )
                except KeyError:
                    # Handle missing keys in entry
                    pass  # Or log the error

    def get_flat_values(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # helper
    def count_participants(self, data, gender):
        count = 0
        for row in data:
            if row['jenis_kelamin'] == gender:
                count += 1
        return count

    def format_addresses(self, objects):
        print(objects)
        result = ""
        for obj in objects:
            result += f"{obj['nama_peserta']} - {obj['alamat']}, "
        # Remove the trailing comma and space
        result = result.rstrip(", ")
        return result
    
    def count_hasil_test(self, data):
        count_positif = 0
        count_negatif = 0

        for row in data:
            if row['hasil_test'] == "Positif":
                count_positif += 1
            elif row['hasil_test'] == "Negatif":
                count_negatif += 1

        return "Positif - {} Negatif - {}".format(count_positif, count_negatif)
    
    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CREATE_UPDATE_Serializer
        return super().get_serializer_class()

class PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.all()
    serializer_class = serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:
                gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                
                jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')

                parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=inserted_id)
                
                peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                    jenis_kelamin=jenis_kelamin,
                    jabatan=entry['jabatan']
                )

    def perform_update(self, serializer):
        current_record_id = serializer.instance.pk

        serializer.save()

        models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent_id=current_record_id).delete()

        peserta_array_lama_json = self.request.data.get('peserta_lamaArray', None)
        if peserta_array_lama_json:
            data = json.loads(peserta_array_lama_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                    
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')

                    parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=current_record_id)
                    
                    peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        jabatan=entry['jabatan']
                    )

        peserta_array_json = self.request.data.get('pesertaArray', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            if len(data) > 0:

                for entry in data:
                    gender_mapping = {'N': 'Laki-Laki', 'P': 'Perempuan'}
                    
                    jenis_kelamin = gender_mapping.get(entry['gender'], 'Laki-Laki')

                    parent_instance = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.get(pk=current_record_id)
                    
                    peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                        jenis_kelamin=jenis_kelamin,
                        jabatan=entry['jabatan']
                    )

        serializer.save(updated_by=self.request.user)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'nama_lingkungan', 'status_indeks', 'nilai_ikp', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'nama_lingkungan': item['nama_lingkungan'],
                'status_indeks': item['status_indeks'],
                'tindak_lanjut': item['tindak_lanjut'],
                'nilai_ikp': item['nilai_ikp'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': satker_level,
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    def get_data_peserta(self, parent):
        data = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta', 'jenis_kelamin', 'jabatan'
        ).order_by('-id')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
                'jenis_kelamin': item['jenis_kelamin'],
                'jabatan': item['jabatan']
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    def delete(self, request, *args, **kwargs):
        try:
            pk = self.kwargs.get('pk')
            deleted_peserta = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_MONITORING_DAN_EVALUASI_SUPERVISI.objects.filter(pk=pk).delete()
            return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}'}, status=status.HTTP_400_BAD_REQUEST)

# ======= PSM PENGUMPULAN DATA IKOTAN API =======
class PSM_PENGUMPULAN_DATA_IKOTAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_PENGUMPULAN_DATA_IKOTAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_PENGUMPULAN_DATA_IKOTAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)
        
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)
            
            satker_parent = {}
            
            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    def get_view_name(self):
        return "PSM PENGUMPULAN DATA IKOTAN"
    
    def get_data_peserta(self, parent):
        data = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta'
        ).order_by('-id')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta']
            }
            serialized_data.append(serialized_item)

        return serialized_data

    def get_flat_values(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_PENGUMPULAN_DATA_IKOTAN_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    # helper
    def count_participants(self, data, gender):
        count = 0
        for row in data:
            if row['jenis_kelamin'] == gender:
                count += 1
        return count

    def format_addresses(self, objects):
        print(objects)
        result = ""
        for obj in objects:
            result += f"{obj['nama_peserta']} - {obj['alamat']}, "
        # Remove the trailing comma and space
        result = result.rstrip(", ")
        return result
    
    def count_hasil_test(self, data):
        count_positif = 0
        count_negatif = 0

        for row in data:
            if row['hasil_test'] == "Positif":
                count_positif += 1
            elif row['hasil_test'] == "Negatif":
                count_negatif += 1

        return "Positif - {} Negatif - {}".format(count_positif, count_negatif)
    
    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return serializers.PSM_MONITORING_DAN_EVALUASI_SUPERVISI_CREATE_UPDATE_Serializer
        return super().get_serializer_class()

class PSM_PENGUMPULAN_DATA_IKOTAN_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.all()
    serializer_class = serializers.PSM_PENGUMPULAN_DATA_IKOTAN_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)
        inserted_id = instance.id

        peserta_array_json = self.request.data.get('peserta_array', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            for entry in data:

                parent_instance = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.get(pk=inserted_id)
                
                peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.create(
                    parent=parent_instance,
                    nama_peserta=entry['nama'],
                )

    def perform_update(self, serializer):
        current_record_id = serializer.instance.pk

        serializer.save()

        models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent_id=current_record_id).delete()

        peserta_array_lama_json = self.request.data.get('peserta_lamaArray', None)
        if peserta_array_lama_json:
            data = json.loads(peserta_array_lama_json)
            if len(data) > 0:

                for entry in data:

                    parent_instance = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.get(pk=current_record_id)
                    
                    peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                    )

        peserta_array_json = self.request.data.get('pesertaArray', None)
        if peserta_array_json:
            data = json.loads(peserta_array_json)
            if len(data) > 0:

                for entry in data:

                    parent_instance = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.get(pk=current_record_id)
                    
                    peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.create(
                        parent=parent_instance,
                        nama_peserta=entry['nama'],
                    )

        serializer.save(updated_by=self.request.user)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        data = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'satker_id__level','observasi', 'tanggal_awal', 'tanggal_akhir', 'deskripsi_hasil', 'simpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'observasi': item['observasi'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi_hasil': item['deskripsi_hasil'],
                'simpulan': item['simpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
                'peserta': self.get_data_peserta(item['id']),
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    def get_data_peserta(self, parent):
        data = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent=parent).values(
            'id', 'nama_peserta'
        ).order_by('-id')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'nama_peserta': item['nama_peserta'],
            }
            serialized_data.append(serialized_item)

        return serialized_data
    
    def delete(self, request, *args, **kwargs):
        try:
            pk = self.kwargs.get('pk')
            deleted_peserta = models.PSM_PENGUMPULAN_DATA_IKOTAN_PESERTA.objects.filter(parent_id=pk).delete()
            deleted_kegiatan = models.PSM_PENGUMPULAN_DATA_IKOTAN.objects.filter(pk=pk).delete()
            return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}'}, status=status.HTTP_400_BAD_REQUEST)


# ======= PSM DUKUNGAN STAKEHOLDER API =======
class PSM_DUKUNGAN_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_DUKUNGAN_STAKEHOLDER.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-created_at').distinct('satker__id')
    serializer_class = serializers.PSM_PENGUMPULAN_DATA_IKOTAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_DUKUNGAN_STAKEHOLDER_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False)
    def get_detail_data(self, request):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'pemda', 'kegiatan', 'alamat', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status', 'satker_id__level'
        ).order_by(
            'satker__nama_satker'
        )

        if satker_level == 1:
            data = data.filter(satker_id=satker)
        elif satker_level == 0:
            data = data.filter(satker__provinsi_id=satker_provinsi, status__gt=0)
        elif satker_level == 2:
            data = data.filter(status=2)
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'pemda': item['pemda'],
                'kegiatan': item['kegiatan'],
                'alamat': item['alamat'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'pemda', 'kegiatan', 'alamat', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status', 'satker_id__level'

        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'pemda': item['pemda'],
                'kegiatan': item['kegiatan'],
                'alamat': item['alamat'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id_ = request.data.get("id_", None)
        
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)
            
            satker_parent = {}
            
            if user_satker_level == 1:
                satker_parent['id'] = user_satker
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''

            kegiatan = models.PSM_DUKUNGAN_STAKEHOLDER.objects.filter(id=id_)

            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari ID {id_}',
                'id': id_,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    def get_view_name(self):
        return "PSM DUKUNGAN STAKEHOLDER"
    
    # untuk bnnk
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')

        detail = []

        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_DUKUNGAN_STAKEHOLDER_DATA_Serializer(data, many=True).data
            }
        ]

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

class PSM_DUKUNGAN_STAKEHOLDER_CURD_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_DUKUNGAN_STAKEHOLDER.objects.all()
    serializer_class = serializers.PSM_DUKUNGAN_STAKEHOLDER_CREATE_UPDATE_Serializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]

    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        serializer.save()

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        instance = serializer.save(created_by=self.request.user, status=status)

    @action(detail=True)
    def get_detail_data_detail(self, request, pk=None):
        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        data = models.PSM_DUKUNGAN_STAKEHOLDER.objects.filter(pk=pk).values(
            'id', 'satker_id', 'satker__nama_satker', 'pemda', 'kegiatan', 'alamat', 'jumlah_sasaran', 'hasil_dampak', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status', 'satker_id__level'

        ).order_by(
            'satker__nama_satker'
        )

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'pemda': item['pemda'],
                'kegiatan': item['kegiatan'],
                'alamat': item['alamat'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                'hasil_dampak': item['hasil_dampak'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'status': item['status'],
                'dokumentasi': item['dokumentasi'],
                'satker_level': item['satker_id__level'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

# ======= PSM RAKOR PEMETAAN API =======
class PSM_RAKOR_PEMETAAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_RAKOR_PEMETAAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_RAKOR_PEMETAAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_RAKOR_PEMETAAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_RAKOR_PEMETAAN_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)

    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_RAKOR_PEMETAAN.objects.get(id=id)
        except models.PSM_RAKOR_PEMETAAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_RAKOR_PEMETAAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
    
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_RAKOR_PEMETAAN.objects.values(
            'id', 'satker_id', 'peserta', 'nama_lingkungan', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_RAKOR_PEMETAAN.objects.get(id=request.data['id'])
        except models.PSM_RAKOR_PEMETAAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_RAKOR_PEMETAAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                user_nama_satker_parent = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker_parent)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_RAKOR_PEMETAAN.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_RAKOR_PEMETAAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_RAKOR_PEMETAAN_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)
    
# ======= PSM AUDIENSI API =======
class PSM_AUDIENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_AUDIENSI.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_AUDIENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_AUDIENSI_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_AUDIENSI_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.AUDIENSI.objects.get(id=id)
        except models.AUDIENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.AUDIENSI_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
    
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_AUDIENSI.objects.values(
            'id', 'satker_id', 'peserta', 'nama_lingkungan', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'nama_lingkungan': item['nama_lingkungan'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_AUDIENSI.objects.get(id=request.data['id'])
        except models.PSM_AUDIENSI.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_AUDIENSI_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_AUDIENSI.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_AUDIENSI.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_AUDIENSI_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)
    
# ======= PSM KONSOLIDASI_KEBIJAKAN API =======
class PSM_KONSOLIDASI_KEBIJAKAN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_KONSOLIDASI_KEBIJAKAN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_KONSOLIDASI_KEBIJAKAN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_KONSOLIDASI_KEBIJAKAN_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.get(id=id)
        except models.PSM_KONSOLIDASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_KONSOLIDASI_KEBIJAKAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
    
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.values(
            'id', 'satker_id', 'peserta', 'stakeholder', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.get(id=request.data['id'])
        except models.PSM_KONSOLIDASI_KEBIJAKAN.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_KONSOLIDASI_KEBIJAKAN_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_KONSOLIDASI_KEBIJAKAN.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_KONSOLIDASI_KEBIJAKAN_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)
    
# ======= PSM WORKSHOP_PENGGIAT API =======
class PSM_WORKSHOP_PENGGIAT_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_WORKSHOP_PENGGIAT.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal', 'satker__satker_order').distinct('satker__id')
    serializer_class = serializers.PSM_WORKSHOP_PENGGIAT_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_WORKSHOP_PENGGIAT_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_WORKSHOP_PENGGIAT_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['delete'])
    def perform_delete(self, request):
        id = request.GET.get('id')
        try:
            instance = models.PSM_WORKSHOP_PENGGIAT.objects.get(id=id)
        except models.PSM_WORKSHOP_PENGGIAT.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = serializers.PSM_WORKSHOP_PENGGIAT_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            instance.delete()
        return Response({'message': 'Berhasil menghapus data'}, status=status.HTTP_200_OK)
    
    @action(detail=False)
    def get_detail_data(self, request):
        id = request.GET.get('id')

        user_id = request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        data = models.PSM_WORKSHOP_PENGGIAT.objects.values(
            'id', 'satker_id', 'peserta', 'stakeholder', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker').filter(id=id)

        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'peserta': item['peserta'],
                'nama_satker': item['satker__nama_satker'],
                'stakeholder': item['stakeholder'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker_level
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['patch'])
    def perform_update(self, request):
        try:
            instance = models.PSM_WORKSHOP_PENGGIAT.objects.get(id=request.data['id'])
        except models.PSM_WORKSHOP_PENGGIAT.DoesNotExist:
            return Response({"message": "Data not found"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.PSM_WORKSHOP_PENGGIAT_CRUD_Serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(updated_by=self.request.user)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        id = request.data.get("id")
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)

            print(id)

            satker_parent = {}
            if user_satker_level == 1:
                user_satker_parent = Profile.objects.values_list('satker__parent', flat=True).get(user_id=user_id)
                satker_parent['id'] = user_satker_parent
                satker_parent['keterangan'] = 'BNN Pusat'
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_WORKSHOP_PENGGIAT.objects.filter(id=id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=2)
                
            return Response({
                'id': id,
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {id}',
                'id': id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False)
    def get_data_bnnk(self, request):
        # get user data
        user_id = request.user.id
        user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)

        # get satker data
        satker = Satker.objects.values().get(id=user_satker)
        data = models.PSM_WORKSHOP_PENGGIAT.objects.all().filter(satker_id=user_satker).order_by('satker__nama_satker')
        serialized_data = [
            {
                'satker': satker,
                'data': serializers.PSM_WORKSHOP_PENGGIAT_DATA_Serializer(data, many=True).data
            }
        ]
        return Response(serialized_data, status=status.HTTP_200_OK)
    
# ======= PSM BIMBINGAN TEKNIS PENGGIAT P4GN API =======
class PSM_BIMTEK_P4GN_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_BIMTEK_P4GN.objects.all().filter(satker__parent__isnull=True).order_by('-satker__id','-tanggal_awal').distinct('satker__id')
    serializer_class = serializers.PSM_BIMTEK_P4GN_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_BIMTEK_P4GN_Filters

    def get_queryset(self):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        queryset = super().get_queryset()
        if satker_level == 0:
            # bnnp
            queryset = queryset.filter(satker__provinsi_id=satker_provinsi, status__gt = 0)
        return queryset

    @action(detail=False, methods=['post'])
    def perform_create(self, request):
        print(request.data)
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=self.request.user.id)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)

        serializer = serializers.PSM_BIMTEK_P4GN_CRUD_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=self.request.user, status=status)
            return Response(serializer.data, status=201)
        
        return Response(serializer.errors, status=400)