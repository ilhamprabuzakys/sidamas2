from django.db.models import F, Count, Q, Subquery, OuterRef
from django.shortcuts import get_object_or_404
from collections import defaultdict
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from users.models import Profile, Satker

import re
import os
import shutil
import codecs
import datetime
import openpyxl

from sidamas import pagination

from kegiatan import models
from kegiatan.api import serializers
from kegiatan.api import filters
from users.models import Satker

class PSM_RAKERNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_RAKERNIS.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.PSM_RAKERNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_RAKERNIS_Filters

    @action(detail=False)
    # def get_all_data(self, request):
    #     user_id = request.user.id
    #     satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
    #     satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
    #     satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

    #     queryset = None
    #     search_value = request.GET.get('search[value]', None)

    #     if satker_level == 1:
    #         queryset = models.PSM_RAKERNIS.objects.filter(
    #             satker_id=satker,
    #         ).values(
    #             'satker_id', 'satker__nama_satker', 'status'
    #         ).annotate(
    #             jumlah_kegiatan=Count('id')
    #         ).order_by(
    #             '-created_at'
    #         ).distinct()
    #     elif satker_level == 0:
    #         queryset = models.PSM_RAKERNIS.objects.filter(
    #             satker__provinsi_id=satker_provinsi, status__gt=0
    #         ).values(
    #             'satker_id', 'satker__nama_satker', 'status'
    #         ).annotate(
    #             jumlah_kegiatan=Count('id')
    #         ).order_by(
    #             '-created_at'
    #         ).distinct()
    #     elif satker_level == 2:
    #         queryset = models.PSM_RAKERNIS.objects.filter(
    #             status=2
    #         ).values(
    #             'satker_id', 'satker__nama_satker', 'status'
    #         ).annotate(
    #             jumlah_kegiatan=Count('id')
    #         ).order_by(
    #             '-created_at'
    #         ).distinct()
    #     else:
    #         raise ValueError(f"Invalid satker level: {satker_level}")

    #     if search_value:
    #         queryset = queryset.filter(
    #             Q(satker__nama_satker__icontains=search_value)
    #         )

    #     serialized_data = []
    #     for row in queryset:
    #         serialized_data.append({
    #             'nama_satker': row['satker__nama_satker'],
    #             'satker_id': row['satker_id'],
    #             'jumlah_kegiatan': row['jumlah_kegiatan'],
    #             'satker_level': satker_level,
    #             'status': row['status'],
    #             'pelaksanaan': self.get_detail_data(row['satker_id'])
    #         })

    #     return Response(serialized_data, status=status.HTTP_200_OK)

    # def get_all_data(self, request):
    #     # persiapan
    #     user_id = request.user.id
    #     satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
    #     satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
    #     satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

    #     queryset = None
    #     search_value = request.GET.get('search[value]', None)

    #     # logika leveling
    #     if satker_level == 1:
    #         # BNNK
    #         queryset = models.PSM_RAKERNIS.objects.filter(
    #             satker_id=satker,
    #         ).values(
    #             'satker_id', 'satker__nama_satker', 'status'
    #         ).annotate(
    #             jumlah_kegiatan=Count('id')
    #         ).order_by(
    #             '-created_at'
    #         )
    #     elif satker_level == 0:
    #         # BNNP
    #         queryset = models.PSM_RAKERNIS.objects.filter(
    #             satker__provinsi_id=satker_provinsi, status__gt=0
    #         ).values(
    #             'satker_id', 'satker__nama_satker', 'status'
    #         ).annotate(
    #             jumlah_kegiatan=Count('id')
    #         ).order_by(
    #             '-created_at'
    #         )
    #     elif satker_level == 2:
    #         # Pusat
    #         queryset = models.PSM_RAKERNIS.objects.filter(
    #             status=2
    #         ).values(
    #             'satker_id', 'satker__nama_satker', 'status'
    #         ).annotate(
    #             jumlah_kegiatan=Count('id')
    #         ).order_by(
    #             '-created_at'
    #         )
    #     else:
    #         raise ValueError(f"Invalid satker level: {satker_level}")

    #     if search_value:
    #         queryset = queryset.filter(
    #             Q(satker__nama_satker__icontains=search_value)
    #         )

    #     serialized_data = [
    #         {
    #             'nama_satker': row['satker__nama_satker'],
    #             'satker_id': row['satker_id'],
    #             'jumlah_kegiatan': row['jumlah_kegiatan'],
    #             'satker_level': satker_level,
    #             'status': row['status'],
    #             'pelaksanaan': self.get_detail_data(row['satker_id'])
    #         }
    #         for row in queryset
    #     ]

    #     return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=False)
    def get_detail_data(self, satker_id=None):
        data = models.PSM_RAKERNIS.objects.all().values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
            }
            serialized_data.append(serialized_item)

        # return serialized_data
        return Response(serialized_data, status=status.HTTP_200_OK)


    @action(detail=True, methods=['delete'])
    def delete_all_kegiatan(self, request, pk=None):
        try:
            deleted_count, _ = models.PSM_BINAAN_TEKNIS.objects.filter(satker_id=pk).delete()
            return Response({
                'status': deleted_count > 0,
                'deleted_count' : deleted_count,
                'message': 'Data kegiatan berhasil dihapus' if deleted_count > 0 else f'Data kegiatan dari Satker ID {pk} tidak ditemukan',
                'satker_id': pk,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}',
                'satker_id': pk,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=['post'])
    def kirim_kegiatan(self, request):
        satker_id = request.data.get("satker_id", None)
        
        try:
            # persiapan
            user_id = request.user.id
            user_satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            user_nama_satker = Satker.objects.values_list('nama_satker', flat=True).get(id=user_satker)
            user_satker_level = Satker.objects.values_list('level', flat=True).get(id=user_satker)
            
            # satker_instance = Satker.objects.filter(pk=satker).first()
            satker_parent = {}
            
            if user_satker_level == 1:
                satker_parent_instance = user_satker.parent
                satker_parent['id'] = user_satker.parent
                satker_parent['keterangan'] = user_nama_satker
            elif user_satker_level == 0:
                satker_parent['id'] = 213
                satker_parent['keterangan'] = 'BNN Pusat'
            else:
                satker_parent['id'] = 0
                satker_parent['keterangan'] = ''
                
            kegiatan = models.PSM_RAKERNIS.objects.filter(satker_id=satker_id)
            
            if user_satker_level == 0:
                kegiatan.update(status=2)
            elif user_satker_level == 1:
                kegiatan.update(status=1)
                
            print(user_satker_level)
            return Response({
                'status': True,
                'message': f'Data kegiatan dari Satuan Kerja {user_nama_satker} berhasil dikirim ke {satker_parent.get("keterangan")}',
                'parent': satker_parent,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengirim kegiatan dari Satuan Kerja ID {satker_id}',
                'satker_id': satker_id,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def export_data(self, request):
        tahun = datetime.datetime.now().year
        file_name = f'REKAPITULASI PEMBINAAN TEKNIS BNNP & BNNK TAHUN {tahun}'
        base_path = 'media/kegiatan/binaan_teknis/exported'
        file_path = f'{base_path}/{file_name}.xlsx'
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Data Kegiatan'
        
        try:
            user_id = request.user.id
            satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
            satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
            satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

            queryset = None
            
            if satker_level == 1:
                # BNNK
                queryset = models.PSM_RAKERNIS.objects.filter(
                    satker_id=satker,
                ).values(
                    'satker_id', 'satker__nama_satker', 'status'
                ).annotate(
                    jumlah_kegiatan=Count('id')
                ).order_by(
                    '-created_at'
                )
            elif satker_level == 0:
                # BNNP
                queryset = models.PSM_RAKERNIS.objects.filter(
                    satker__provinsi_id=satker_provinsi, status__gt=0
                ).values(
                    'satker_id', 'satker__nama_satker', 'status'
                ).annotate(
                    jumlah_kegiatan=Count('id')
                ).order_by(
                    '-created_at'
                )
            elif satker_level == 2:
                # Pusat
                queryset = models.PSM_RAKERNIS.objects.filter(
                    status = 2
                ).values(
                    'satker_id', 'satker__nama_satker', 'status'
                ).annotate(
                    jumlah_kegiatan=Count('id')
                ).order_by(
                    '-created_at'
                )
            else:
                raise ValueError(f"Invalid satker level: {satker_level}")
            
            serialized_data = []
            for index, item in enumerate(queryset, start=1):
                data = models.PSM_RAKERNIS.objects.filter(satker_id=item['satker_id']).values(
                    'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
                    'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
                ).order_by('satker__nama_satker')

                detail_data = data
                serialized_item = {
                    'nomor': index,
                    'status': item['status'],
                    'satker_id': item['satker_id'],
                    'nama_satker': item['satker__nama_satker'],
                    'jumlah_kegiatan': item['jumlah_kegiatan'],
                    'detail_kegiatan': detail_data
                }
                serialized_data.append(serialized_item)

            # tb_header
            headers = ['No', 'Satuan Kerja Pelaksana', 'Jumlah Kegiatan', 'Status']
            for item, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)  # Membuat teks tebal
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            # save file
            shutil.rmtree(base_path)
            os.makedirs(base_path, exist_ok=True)
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diexport',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengexport kegiatan dari Satuan Kerja',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
            
    
    def get(self, request):
        return self.get_all_data(request)
    
    def perform_create(self, serializer):
        user_id = self.request.user.id
        satker = Profile.objects.values_list('satker', flat=True).get(user_id=user_id)
        satker_provinsi = Satker.objects.values_list('provinsi_id', flat=True).get(id=satker)
        satker_level = Satker.objects.values_list('level', flat=True).get(id=satker)

        status_map = {1: 0, 0: 1, 2: 2}
        status = status_map.get(satker_level, None)
        serializer.save(created_by=self.request.user, status=status)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def get_view_name(self):
        return "PSM RAKERNIS"

class PSM_BINAAN_TEKNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_BINAAN_TEKNIS.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.PSM_BINAAN_TEKNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_BINAAN_TEKNIS_Filters

    @action(detail=False)
    def get_all_data(self, request):
        search_value = request.GET.get('search[value]', None)
        
        queryset = models.PSM_BINAAN_TEKNIS.objects.values(
            'satker_id', 'satker__nama_satker'
        ).annotate(
            jumlah_kegiatan=Count('id')
        ).order_by(
            'satker_id'
        )

        if search_value:
            queryset = queryset.filter(
                Q(satker__nama_satker__icontains=search_value)
            )

        # Mengambil 50 data teratas setelah filter diterapkan
        queryset = queryset[:50]

        serialized_data = [
            {
                'id': idx + 1,
                'nama_satker': row['satker__nama_satker'],
                'satker_id': row['satker_id'],
                'jumlah_kegiatan': row['jumlah_kegiatan'],
            }
            for idx, row in enumerate(queryset)
        ]

        return Response(serialized_data, status=status.HTTP_200_OK)


    @action(detail=False)
    def get_detail_data(self, request):
        satker_id = request.GET["satker_id"]
        data = models.PSM_BINAAN_TEKNIS.objects.filter(satker_id=satker_id).values(
            'id', 'satker_id', 'satker_target', 'satker_target__nama_satker', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi'
        ).order_by('satker__nama_satker')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'satker_target': item['satker_target'],
                'nama_satker': item['satker__nama_satker'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
            }
            serialized_data.append(serialized_item)

        return Response(serialized_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'])
    def delete_all_kegiatan(self, request, pk=None):
        try:
            deleted_count, _ = models.PSM_BINAAN_TEKNIS.objects.filter(satker_id=pk).delete()
            return Response({
                'status': deleted_count > 0,
                'deleted_count' : deleted_count,
                'message': 'Data kegiatan berhasil dihapus' if deleted_count > 0 else f'Data kegiatan dari Satker ID {pk} tidak ditemukan',
                'satker_id': pk,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal saat menghapus semua kegiatan pada satker_id {pk}',
                'satker_id': pk,
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        return self.get_all_data(request)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def get_view_name(self):
        return "PSM BINTEK"
    
class PSM_ASISTENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.PSM_ASISTENSI.objects.all().order_by('-tanggal')
    serializer_class = serializers.PSM_ASISTENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.PSM_ASISTENSI_Filters

    @action(detail=False)
    def get_all_data(self, request):
        search_value = request.GET.get('search[value]', None)
        
        queryset = models.PSM_ASISTENSI.objects.values(
            'satker_id', 'satker__nama_satker'
        ).annotate(
            jumlah_kegiatan=Count('id')
        ).order_by(
            'satker_id'
        )

        if search_value:
            queryset = queryset.filter(
                Q(satker__nama_satker__icontains=search_value)
            )

        # Mengambil 50 data teratas setelah filter diterapkan
        queryset = queryset[:50]

        serialized_data = [
            {
                'id': idx + 1,
                'nama_satker': row['satker__nama_satker'],
                'satker_id': row['satker_id'],
                'jumlah_kegiatan': row['jumlah_kegiatan'],
            }
            for idx, row in enumerate(queryset)
        ]

        return Response(serialized_data, status=status.HTTP_200_OK)


    @action(detail=False)
    def get_detail_data(self, request):
        satker_id = request.GET["satker_id"]
        data = models.PSM_ASISTENSI.objects.filter(satker_id=satker_id).values(
            'id', 'satker_id', 'satker__nama_satker', 'jumlah_kegiatan', 'tanggal', 'jumlah_peserta', 'stakeholder',
            'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi'
        ).order_by('satker__nama_satker')
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'jumlah_kegiatan': item['jumlah_kegiatan'],
                'tanggal': item['tanggal'],
                'jumlah_peserta': item['jumlah_peserta'],
                'stakeholder': item['stakeholder'],
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
            }
            serialized_data.append(serialized_item)

        return Response(serialized_data, status=status.HTTP_200_OK)

    def get(self, request):
        return self.get_all_data(request)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def get_view_name(self):
        return "PSM ASISTENSI"