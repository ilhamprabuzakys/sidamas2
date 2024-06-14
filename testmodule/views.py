from django.views import generic
from django.urls import reverse_lazy
from django.shortcuts import render
from import_export.admin import ExportMixin
from .resources import UserResource, ImageResource
from django.views import View
from django.http import HttpResponse
from openpyxl import Workbook
from tablib import Dataset
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
from django.conf import settings
from . import models
from . import forms


class userListView(generic.ListView):
    model = models.user
    form_class = forms.userForm


class userCreateView(generic.CreateView):
    model = models.user
    form_class = forms.userForm


class userDetailView(generic.DetailView):
    model = models.user
    form_class = forms.userForm


class userUpdateView(generic.UpdateView):
    model = models.user
    form_class = forms.userForm
    pk_url_kwarg = "pk"


class userDeleteView(generic.DeleteView):
    model = models.user
    success_url = reverse_lazy("testmodule_user_list")


def testdatatables(request):
    return render(request, 'module/coba_tabel.html')

def testchart(request):
    return render(request, 'module/coba_chartjs.html')

def testeform(request):
    return render(request, 'module/coba_efom.html')

def testsurvey(request):
    return render(request, 'module/coba_survey.html')

def testinteraktif(request):
    return render(request, 'module/coba_interaktif.html')

class ExportDataView(View):
    def get(self, request, *args, **kwargs):
        # Create a new Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Get the column titles from the UserResource
        user_resource = UserResource()
        column_titles = [field.column_name for field in user_resource.get_user_fields()]

        # Write the column titles to the first row of the worksheet
        for col, title in enumerate(column_titles):
            ws.cell(row=1, column=col + 1, value=title)

        # Create a dataset using the UserResource
        dataset = user_resource.export()

        # Iterate through the dataset and write data to the worksheet
        for row, record in enumerate(dataset, start=2):  # Start from the second row
            for col, value in enumerate(record):
                ws.cell(row=row, column=col + 1, value=value)

        # Create an HTTP response with the exported Excel data
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="coba_download.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response
    
class ExportDataImg(View):
    def get(self, request, *args, **kwargs):
        # Create a new Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Add headers for the title and image columns
        ws['A1'] = 'Title'
        ws['B1'] = 'Image'

        # Get the queryset of images
        images = models.image.objects.all()

        # Create an HTTP response with the exported Excel data
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="image_download.xlsx'

        # Set the row height for all rows to 200
        for i in range(2, len(images) + 2):
            ws.row_dimensions[i].height = 200

        for i, img in enumerate(images):
            # Get the image file path and create an Image object
            image_path = os.path.join(settings.MEDIA_ROOT, str(img.image))
            img_obj = Image(image_path)

            # Calculate the desired image width and height (e.g., 200x200)
            img_obj.width = 200
            img_obj.height = 200

            # Add the Image object to the worksheet in the appropriate cell
            ws.add_image(img_obj, f'B{i + 2}')

            # Add the title in the 'A' column
            ws[f'A{i + 2}'] = img.title

            ws.column_dimensions[get_column_letter(1)].width = 200

        # Save the workbook to the response
        wb.save(response)

        return response