from django.views import View, generic
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from . import models
from . import forms
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse
from django.urls import reverse

# excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from django.http import HttpResponse
from django.http import JsonResponse
import datetime
import json
import os.path

# Define global variables outside of any function
data_array_responden = None
data_sigma_nilai = None
data_nrr = None
data_nrrt = None
data_responden = None
data_nrrtu = None
data_ikm = None
data_conclusion = None
data_keterangan = None
data_kode = None
data_judul_layanan = None
data_alamat = None
data_telpfax = None
data_bulan = None
data_tahun = None
data_satker = None

class GlobalPermissionMixin:
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_superuser == False and request.user.is_staff == False and request.user.profile.role is None or request.user.profile.satker is None or request.user.profile.is_verified == False:
            user = self.request.user
            message = "Maaf " + user.username + ", anda tidak memiliki hak akses untuk mengunjungi halaman ini."
            print(message)
            return HttpResponseRedirect(reverse("dashboard:profile"))
        return super().dispatch(request, *args, **kwargs)

class SurveyBaseView(GlobalPermissionMixin, LoginRequiredMixin):

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class hasil_surveiView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/psm/test_urine/hasil_survei.html"

    def get(self, request, pk):
        return render(request, self.template_name)

# gk kepake tpi di simpen dulu
class hasil_survei_teView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/hasil_survei_te.html"

    def get(self, request, pk):
        return render(request, self.template_name)
# gk kepake

class hasil_survei_lsView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/dayatif/life_skill/hasil_survei_ls.html"

    def get(self, request, pk):
        return render(request, self.template_name)

class hasil_survei_kkView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/dayatif/kewirausahaan_keberhasilan/hasil_survei_kk.html"

    def get(self, request, pk):
        return render(request, self.template_name)

class hasil_survei_triwulanView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/psm/test_urine/hasil_triwulan.html"

    def get(self, request, bulan, tahun, satker):
        return render(request, self.template_name)

class hasil_survei_triwulan_lsView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/dayatif/life_skill/hasil_triwulan_ls.html"

    def get(self, request, bulan, tahun, satker):
        return render(request, self.template_name)

class hasil_survei_triwulan_kkView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/dayatif/kewirausahaan_keberhasilan/hasil_survei_triwulan.html"

    def get(self, request, bulan, tahun, satker):
        return render(request, self.template_name)

class chart_surveiView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/chart_survei.html"

    def get(self, request, pk):
        return render(request, self.template_name)

class chart_survei_kkView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/dayatif/kewirausahaan_keberhasilan/chart_survei.html"

    def get(self, request, pk):
        return render(request, self.template_name)

class triwulan_surveiView(SurveyBaseView, View):
    template_name = "survei/hasil_survei/chart_triwulan.html"

    def get(self, request, bulan, tahun):
        return render(request, self.template_name)

# download xlsx
def generate_spreadsheet_tu(request):
    if request.method == 'POST':
        global data_array_responden
        global data_sigma_nilai
        global data_nrr
        global data_nrrt
        global data_responden
        global data_nrrtu
        global data_ikm
        global data_kode
        global data_judul_layanan
        global data_alamat
        global data_telpfax
        global data_bulan
        global data_tahun
        global data_satker

        # Retrieve the raw JSON data from the request body
        data = json.loads(request.body)

        # Access the 'jawaban' key from the JSON data
        responden = data.get('jawaban')
        sigma_nilai = data.get('sigma_nilai')
        nrr = data.get('nrr')
        nrrt = data.get('nrrt')
        total_responden = data.get('responden')
        nrrtu = data.get('nrrtu')
        ikm = data.get('ikm1')
        kode = data.get('kode')
        unitpelayanan = data.get('unitpelayanan')
        alamat = data.get('alamat')
        tlpFax = data.get('tlpFax')
        bulan = data.get ('bulan')
        tahun = data.get ('tahun')
        satker = data.get ('satker')

        data_array_responden = responden
        data_sigma_nilai = sigma_nilai
        data_nrr = nrr
        data_nrrt = nrrt
        data_responden = total_responden
        data_nrrtu = nrrtu
        data_ikm = ikm
        data_kode = kode
        data_judul_layanan = unitpelayanan
        data_alamat = alamat
        data_telpfax = tlpFax
        data_bulan = bulan
        data_tahun = tahun
        data_satker = satker

        response_data = {
            'message': 'data sudah di set file segera terdownload',
        }
        return JsonResponse(response_data)

def generate_spreadsheet_ls(request):
    if request.method == 'POST':
        global data_array_responden
        global data_nrr
        global data_nrrt
        global data_conclusion
        global data_keterangan
        global data_kode
        global data_judul_layanan
        global data_alamat
        global data_bulan
        global data_tahun
        global data_satker

        data = json.loads(request.body)

        responden = data.get('jawaban')
        nrr = data.get('nrr')
        nrrt = data.get('nrrt')
        conclusion = data.get('kesimpulan')
        keterangan = data.get('keterangan')
        kode = data.get('kode')
        unitpelayanan = data.get('unitpelayanan')
        alamat = data.get('alamat')
        bulan = data.get ('bulan')
        tahun = data.get ('tahun')
        satker = data.get ('satker')

        data_array_responden = responden
        data_nrr = nrr
        data_nrrt = nrrt
        data_conclusion = conclusion
        data_keterangan = keterangan
        data_kode = kode
        data_judul_layanan = unitpelayanan
        data_alamat = alamat
        data_bulan = bulan
        data_tahun = tahun
        data_satker = satker

        response_data = {
            'message': 'data sudah di set file segera terdownload',
        }
        return JsonResponse(response_data)

def generate_spreadsheet_kk(request):
    if request.method == 'POST':
        global data_array_responden
        global data_nrr
        global data_nrrt
        global data_conclusion
        global data_keterangan
        global data_kode
        global data_judul_layanan
        global data_alamat
        global data_bulan
        global data_tahun
        global data_satker

        data = json.loads(request.body)

        responden = data.get('jawaban')
        nrr = data.get('nrr')
        nrrt = data.get('nrrt')
        conclusion = data.get('kesimpulan')
        keterangan = data.get('keterangan')
        kode = data.get('kode')
        unitpelayanan = data.get('unitpelayanan')
        alamat = data.get('alamat')
        bulan = data.get ('bulan')
        tahun = data.get ('tahun')
        satker = data.get ('satker')

        data_array_responden = responden
        data_nrr = nrr
        data_nrrt = nrrt
        data_conclusion = conclusion
        data_keterangan = keterangan
        data_kode = kode
        data_judul_layanan = unitpelayanan
        data_alamat = alamat
        data_bulan = bulan
        data_tahun = tahun
        data_satker = satker

        response_data = {
            'message': 'data sudah di set file segera terdownload',
        }
        return JsonResponse(response_data)

def spreedsheet_tu(request):
    global data_array_responden
    global data_sigma_nilai
    global data_nrr
    global data_nrrt
    global data_responden
    global data_nrrtu
    global data_ikm
    global data_kode
    global data_judul_layanan
    global data_alamat
    global data_telpfax
    global data_bulan
    global data_tahun
    global data_satker

    current_year = datetime.datetime.now().year
    classification = grade_classification_tu(data_ikm)
    termin = get_termin(data_bulan)

    #
    data_survei_obj = models.DataSurvei.objects.get(id=data_kode)
    if data_survei_obj.satker is not None:
        satker = data_survei_obj.satker.nama_satker if data_survei_obj.satker.nama_satker else "-"
    else:
        satker = "Satuan Kerja"
    tipe = data_survei_obj.tipe.nama
    tanggal_awal = data_survei_obj.tanggal_awal
    tanggal_akhir = data_survei_obj.tanggal_akhir

    wb = load_workbook("/home/intioptima/sidamas/templates/template_skm_test_urine.xlsx")
    ws = wb.active

    # current_row = 9  # Starting row
    current_row = 3

    sentence = "PER RESPONDEN DAN PER UNSUR PELAYANAN TAHUN {}"
    modified_sentence = sentence.format(current_year)
    ws.cell(row=current_row, column=2, value=modified_sentence)
    ws.cell(row=current_row, column=2).font = Font(bold=True)

    ws.cell(row=current_row+1, column=2, value="Unit Pelayanan : " + (data_judul_layanan if data_judul_layanan is not None else ""))
    ws.cell(row=current_row+2, column=2, value="Alamat : " + (data_alamat if data_alamat is not None else ""))
    ws.cell(row=current_row+3, column=2, value="Telp/Fax : " + (data_telpfax if data_telpfax is not None else ""))

    for index, rowData in enumerate(data_array_responden, start=1):
        ws.insert_rows(current_row+6)
        row = [index]
        for cellData in rowData:
            for key, value in cellData.items():
                row.append(int(value))
        for i, val in enumerate(row, start=2):  # Use enumerate to iterate over both index and value
            ws.cell(row=current_row+6, column=i, value=val)
        current_row += 1  # Increment the row number for the next insertion

    all_cells = ws['B9:K{}'.format(current_row+6)]
    all_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))
    for row in all_cells:
        for cell in row:
            cell.border = all_border
            cell.number_format = '#'

    dynamic_row1 = current_row + 6 # Start at the current row
    for i, val in enumerate(data_sigma_nilai, start=3):
        ws.cell(row=dynamic_row1, column=i, value=val)

    dynamic_row2 = current_row + 7 # Start at the current row
    for i, val in enumerate(data_nrr, start=3):
        ws.cell(row=dynamic_row2, column=i, value=val)

    dynamic_row3 = current_row + 8 # Start at the current row
    for i, val in enumerate(data_nrrt, start=3):
        ws.cell(row=dynamic_row3, column=i, value=val)

    ws.cell(row=current_row + 9, column=11, value=data_responden)
    ws.cell(row=current_row + 10, column=11, value=data_nrrtu)
    ws.cell(row=current_row + 13, column=11, value=data_nrr[0])
    ws.cell(row=current_row + 14, column=11, value=data_nrr[1])
    ws.cell(row=current_row + 15, column=11, value=data_nrr[2])
    ws.cell(row=current_row + 16, column=11, value=data_nrr[3])
    ws.cell(row=current_row + 17, column=11, value=data_nrr[4])
    ws.cell(row=current_row + 18, column=11, value=data_nrr[5])
    ws.cell(row=current_row + 19, column=11, value=data_nrr[6])
    ws.cell(row=current_row + 20, column=11, value=data_nrr[7])
    ws.cell(row=current_row + 21, column=11, value=data_nrr[8])
    ws.cell(row=current_row + 23, column=2, value="SKM UNIT PELAYANAN : " + str(data_ikm) + " " + classification)


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, satker, tanggal_awal, tanggal_akhir)

    if data_bulan is not None and data_satker is not None:
        if data_bulan == "T0" and data_satker == 0:
            response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, data_tahun, tanggal_awal, tanggal_akhir)
        else:
            response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, termin, satker, data_tahun, tanggal_awal, tanggal_akhir)
    else:
        response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, satker, data_tahun, tanggal_awal, tanggal_akhir)


    wb.save(response)

    return response

def spreedsheet_ls(request):
    global data_array_responden
    global data_nrr
    global data_nrrt
    global data_conclusion
    global data_keterangan
    global data_kode
    global data_judul_layanan
    global data_alamat
    global data_bulan
    global data_tahun
    global data_satker

    #
    termin = get_termin(data_bulan)


    data_survei_obj = models.DataSurvei.objects.get(id=data_kode)
    satker = data_survei_obj.satker.nama_satker
    tipe = data_survei_obj.tipe.nama
    tanggal_awal = data_survei_obj.tanggal_awal
    tanggal_akhir = data_survei_obj.tanggal_akhir

    wb = load_workbook("/home/intioptima/sidamas/templates/template_skm_life_skill.xlsx")
    ws = wb.active

    # current_row = 7
    current_row = 2
    row_index = 1

    ws.cell(row=current_row, column=2, value="" + (data_judul_layanan if data_judul_layanan is not None else ""))
    ws.cell(row=current_row+1, column=2, value="" + (data_alamat if data_alamat is not None else ""))
    ws.cell(row=current_row, column=2).font = Font(bold=True)
    ws.cell(row=current_row+1, column=2).font = Font(bold=True)

    # for row_data in data_array_responden:
    #     ws.insert_rows(current_row+5)
    #     row_data_with_index = [row_index] + row_data
    #     for col, value in enumerate(row_data_with_index, start=2):
    #         ws.cell(row=current_row+5, column=col, value=value)
    #     current_row += 1
    #     row_index += 1

    for row_data in data_array_responden:
        ws.insert_rows(current_row + 5)
        row_data_with_index = [row_index] + row_data
        for col, value in enumerate(row_data_with_index, start=2):
            original_value = value
            formatted_value = '-'
            if isinstance(value, (str)):
                formatted_value = original_value
            else:
                if value is not None:
                    if col == 2:
                        formatted_value = value
                    else:
                        formatted_value = f'{value:.2f}'

            ws.cell(row=current_row + 5, column=col, value=formatted_value)
        current_row += 1
        row_index += 1

    ws.insert_rows(current_row+5)
    for col, value in enumerate(data_nrr, start=3):
            ws.cell(row=current_row+5, column=col, value=value)
    current_row += 1

    ws.insert_rows(current_row+5)
    for col, value in enumerate(data_nrrt, start=3):
        original_value = value
        formatted_value = ''
        if isinstance(value, (str)):
            formatted_value = original_value
        else:
            if value is not None:
                formatted_value = f'{value:.2f}'

        ws.cell(row=current_row+5, column=col, value=formatted_value)
    current_row += 1

    ws.insert_rows(current_row+5)
    for col, value in enumerate(data_conclusion, start=3):
            ws.cell(row=current_row+5, column=col, value=value)
    current_row += 1

    dynamic = current_row + 7

    lines = data_keterangan.split('\n')

    for i, line in enumerate(lines):
        ws.insert_rows(dynamic)

    left_border_style = Border(left=Side(style='medium'))
    right_border_style = Border(right=Side(style='medium'))

    for i, line in enumerate(lines):
        current_row_for_line = current_row + 7 + i
        ws.cell(row=current_row_for_line, column=2, value=line)
        ws.cell(row=current_row_for_line, column=2).alignment = Alignment(wrapText=True)
        ws.merge_cells(start_row=current_row_for_line, start_column=2, end_row=current_row_for_line, end_column=15)

        # Apply left border style to column 2
        left_cell = ws.cell(row=current_row_for_line, column=2)
        left_cell.border = left_border_style

        # Apply right border style to column 15
        right_cell = ws.cell(row=current_row_for_line, column=15)
        right_cell.border = right_border_style
        all_cells = ws['B7:O{}'.format(current_row+5)]
        all_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))
        for row in all_cells:
            for cell in row:
                cell.border = all_border
                cell.number_format = '#'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename="export data survey.xlsx"'
    # response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, satker, tanggal_awal, tanggal_akhir)

    if data_bulan is not None and data_satker is not None:
        if data_bulan == "T0" and data_satker == 0:
            response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, data_tahun, tanggal_awal, tanggal_akhir)
        else:
            response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, termin, satker, data_tahun, tanggal_awal, tanggal_akhir)
    else:
        response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, satker, data_tahun, tanggal_awal, tanggal_akhir)


    wb.save(response)

    return response

def spreedsheet_kk(request):
    global data_array_responden
    global data_nrr
    global data_nrrt
    global data_conclusion
    global data_keterangan
    global data_kode
    global data_judul_layanan
    global data_alamat
    global data_bulan
    global data_tahun
    global data_satker

    #
    termin = get_termin(data_bulan)

    data_survei_obj = models.DataSurvei.objects.get(id=data_kode)
    satker = data_survei_obj.satker.nama_satker
    tipe = data_survei_obj.tipe.nama
    tanggal_awal = data_survei_obj.tanggal_awal
    tanggal_akhir = data_survei_obj.tanggal_akhir

    wb = load_workbook("/home/intioptima/sidamas/templates/template_skm_keberhasilan_kewirausahaan.xlsx")
    ws = wb.active

    # current_row = 7
    current_row = 2
    row_index = 1

    ws.cell(row=current_row, column=2, value="" + (data_judul_layanan if data_judul_layanan is not None else ""))
    ws.cell(row=current_row+1, column=2, value="" + (data_alamat if data_alamat is not None else ""))
    ws.cell(row=current_row, column=2).font = Font(bold=True)
    ws.cell(row=current_row+1, column=2).font = Font(bold=True)

    for row_data in data_array_responden:
        ws.insert_rows(current_row+5)
        row_data_with_index = [row_index] + row_data
        for col, value in enumerate(row_data_with_index, start=2):
            original_value = value
            formatted_value = '-'
            if isinstance(value, (str)):
                formatted_value = original_value
            else:
                if value is not None:
                    if col == 2:
                        formatted_value = value
                    else:
                        formatted_value = f'{value:.2f}'
            ws.cell(row=current_row+5, column=col, value=formatted_value)
        current_row += 1
        row_index += 1

    ws.insert_rows(current_row+5)
    for col, value in enumerate(data_nrr, start=3):
            ws.cell(row=current_row+5, column=col, value=value)
    current_row += 1

    ws.insert_rows(current_row+5)
    for col, value in enumerate(data_nrrt, start=3):
        original_value = value
        formatted_value = ''
        if isinstance(value, (str)):
            formatted_value = original_value
        else:
            if value is not None:
                formatted_value = f'{value:.2f}'
        ws.cell(row=current_row+5, column=col, value=formatted_value)
    current_row += 1

    ws.insert_rows(current_row+5)
    for col, value in enumerate(data_conclusion, start=3):
        ws.cell(row=current_row+5, column=col, value=value)
    current_row += 1

    dynamic = current_row + 7

    lines = data_keterangan.split('\n')

    for i, line in enumerate(lines):
        ws.insert_rows(dynamic)

    left_border_style = Border(left=Side(style='medium'))
    right_border_style = Border(right=Side(style='medium'))

    for i, line in enumerate(lines):
        current_row_for_line = current_row + 7 + i
        ws.cell(row=current_row_for_line, column=2, value=line)
        ws.cell(row=current_row_for_line, column=2).alignment = Alignment(wrapText=True)
        ws.merge_cells(start_row=current_row_for_line, start_column=2, end_row=current_row_for_line, end_column=16)

        # Apply left border style to column 2
        left_cell = ws.cell(row=current_row_for_line, column=2)
        left_cell.border = left_border_style

        # Apply right border style to column 15
        right_cell = ws.cell(row=current_row_for_line, column=16)
        right_cell.border = right_border_style

        all_cells = ws['B7:P{}'.format(current_row+5)]
        all_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))
        for row in all_cells:
            for cell in row:
                cell.border = all_border
                cell.number_format = '#'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename="export data survey.xlsx"'
    # response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, satker, tanggal_awal, tanggal_akhir)
    if data_bulan is not None and data_satker is not None:
        if data_bulan == "T0" and data_satker == 0:
            response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, data_tahun, tanggal_awal, tanggal_akhir)
        else:
            response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, termin, satker, data_tahun, tanggal_awal, tanggal_akhir)
    else:
        response['Content-Disposition'] = 'attachment; filename="Survei {} {} {} {}.xlsx"'.format(tipe, satker, data_tahun, tanggal_awal, tanggal_akhir)


    wb.save(response)

    return response
# download xlsx

# helper
def grade_classification_tu(value):
    if 88.31 <= value <= 100.00:
        return "A (Sangat Baik)"
    elif 76.61 <= value <= 88.30:
        return "B (Baik)"
    elif 65.00 <= value <= 76.60:
        return "C (Kurang Baik)"
    elif 25.00 <= value <= 64.99:
        return "D (Tidak Baik)"
    else:
        return "Invalid Value"

def get_termin(term):
    if term == "T1":
        return "Termin 1"
    elif term == "T2":
        return "Termin 2"
    elif term == "T3":
        return "Termin 3"
    elif term == "T4":
        return "Termin 4"
    else:
        return "Invalid term"
# helper