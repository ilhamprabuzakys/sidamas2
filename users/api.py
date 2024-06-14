import calendar
import datetime
import os
import shutil
from django.contrib.auth import login
from django.db.models import F
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
import pandas as pd
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.contrib.auth.models import User

from sidamas import pagination

from . import serializers
from . import models
from . import filters
from . import signals

class ProfileViewSet(viewsets.ModelViewSet):
    queryset = models.Profile.objects.all()
    serializer_class = serializers.ProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=True, methods=['put'])
    def update_is_staff(self, request, pk=None):
        try:
            profile = self.get_object()
            user = profile.user
            user.is_staff = True
            user.save()

            profile.is_verified = True
            profile.save()

            return Response({'detail': 'is_staff updated successfully.'}, status=status.HTTP_200_OK)
        except models.Profile.DoesNotExist:
            return Response({'detail': 'Profile not found.'}, status=status.HTTP_404_NOT_FOUND)

class UserViewSet(viewsets.ModelViewSet):
    # queryset = User.objects.annotate(profile_created_at=F('profile__created_at')).order_by('-profile_created_at')
    queryset = User.objects.order_by('-id')
    serializer_class = serializers.UserSerializer
    filter_backends = [DjangoFilterBackend, ]
    filterset_class = filters.UserFilter
    permission_classes = [permissions.IsAuthenticated,]
    pagination_class = pagination.Page10NumberPagination

    @action(detail=False, methods=['POST'], url_path='reset-password', name='Reset Password')
    def reset_password(self, request):
        id = self.request.data.get('id')

        try:
            user = User.objects.get(pk=id)

            username = user.username

            new_password = signals.generate_random_password()
            user.set_password(new_password)
            user.save()

            signals.send_user_creation_email(user, new_password, "Reset Akun")

            return Response({
                'status': True,
                'message': f'Password untuk user {username} telah berhasil direset',
            }, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({
                'status': False,
                'message': f'User dengan id {id} tidak ditemukan.',
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['POST'], url_path='update-password', name='Update Password')
    def update_password(self, request):

        user = request.user

        current_password = self.request.data.get('current_password')
        new_password = self.request.data.get('new_password')


        if not user.check_password(current_password):
            return Response({'error': 'Password anda saat ini salah.'}, status=status.HTTP_400_BAD_REQUEST)

        if new_password == current_password:
            return Response({'error': 'Password baru tidak boleh sama dengan password saat ini.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.save()

        # login(request, user)
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')

        return Response({'message': 'Password anda berhasil diperbarui.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST'], url_path='import', name='Import Data')
    def import_data(self, request):
        try:
            file = request.FILES.get('file')

            if not file:
                raise Exception("File must be provided ...")

            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            data_count = { 'created': 0, 'skipped': 0 }
            users = []

            for row in sheet.iter_rows(min_row=3, max_col=5, values_only=True):
                no, nama, username, email, notelp = row

                if not nama and not username and not email:
                    continue

                if User.objects.filter(username=username).exists():
                    continue

                user = User.objects.create_user(username=username, email=email, first_name=nama)
                user.set_password(username)

                user.profile.notelp = notelp
                user.profile.save()

                data_count['created'] += 1

            return Response({
                'status': True,
                'data': [],
                'users': users,
                'count': data_count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'data': [],
                'message': str(e),
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET'], url_path='export', name='Export Data')
    def export_data(self, request):
        try:
            base_template = '/home/intioptima/sidamas/media/templates/users/registrasi.xlsx'

            wb = openpyxl.load_workbook(base_template)
            ws = wb.active

            start_row = 3
            current_row = start_row

            users = models.UserLengkap.objects.order_by('-id')
            idx = 0

            alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

            for user in users:
                idx += 1

                ws.cell(row=current_row, column=1).value = idx
                ws.cell(row=current_row, column=1).alignment = alignment
                ws.cell(row=current_row, column=1).border = border

                ws.cell(row=current_row, column=2).value = user.first_name
                ws.cell(row=current_row, column=2).alignment = alignment
                ws.cell(row=current_row, column=2).border = border

                ws.cell(row=current_row, column=3).value = user.username
                ws.cell(row=current_row, column=3).alignment = alignment
                ws.cell(row=current_row, column=3).border = border

                ws.cell(row=current_row, column=4).value = user.email
                ws.cell(row=current_row, column=4).alignment = alignment
                ws.cell(row=current_row, column=4).border = border

                ws.cell(row=current_row, column=5).value = user.notelp
                ws.cell(row=current_row, column=5).alignment = alignment
                ws.cell(row=current_row, column=5).border = border

                current_row += 1

            for col in range(2, ws.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col)
                column_dimension = ws.column_dimensions[column_letter]
                column_dimension.auto_size = True
                column_dimension.width *= 1.2

            current_date = datetime.datetime.now()
            tahun = current_date.year
            bulan = calendar.month_name[current_date.month]

            base_url = self.request.build_absolute_uri('/') + 'media/'
            file_name = f'Data Pengguna SIDAMAS {bulan} - {tahun}.xlsx'
            file_path = f'/home/intioptima/sidamas/media/exported/users/{file_name}'
            link = f'{base_url}/exported/users/{file_name}'

            wb.save(file_path)

            return Response({
                'status': True,
                'file_path': link
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': str(e),
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET'], url_path='export_data', name='Export Data Old')
    def export(self, request):
        tahun = datetime.datetime.now().year

        try:
            data = models.UserLengkap.objects.all()
            df = pd.DataFrame.from_records(data.values())
            df.index += 1

            # return Response({
            #     'status': True,
            #     'df' : df,
            #     'message': 'Export data successfully',
            # }, status=status.HTTP_200_OK)

            base_path = 'media/exported/users'
            file_name = f'DATA PENGGUNA - TAHUN {tahun}'
            file_extension = 'xlsx'
            file_path = f'{base_path}/{file_name}.{file_extension}'

            if os.path.exists(base_path): shutil.rmtree(base_path)
            os.makedirs(base_path, exist_ok=True)

            # Rename header
            df.rename(columns={
                'first_name': 'Nama Depan',
                'last_name': 'Nama Belakang',
                'username': 'Username',
                'email': 'Email',
                'notelp': 'Nomor Telepon',
                'nama_satker': 'Satker',
                'role': 'Direktorat',
            }, inplace=True)

            columns = ['Nama Depan', 'Nama Belakang', 'Username', 'Email', 'Nomor Telepon', 'Satker', 'Direktorat']

            if type == 'csv':
                df.to_csv(file_path, mode='w', columns=columns, index=True, index_label='NO')
            else:
                # Membuat workbook dan worksheet baru
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = f'DATA PENGGUNA - TAHUN {tahun}'

                # Menulis data ke worksheet
                for row_idx, row in enumerate(df.values, start=1):
                    for col_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

                # Menambahkan judul kolom
                column_titles = ['NO', 'Nama', 'Username', 'Email', 'Nomor Telepon', 'Satker', 'Direktorat']
                for col_idx, title in enumerate(column_titles, start=1):
                    worksheet.cell(row=1, column=col_idx, value=title)

                # Mengatur lebar kolom
                column_widths = [6, 28, 25, 30, 30, 30, 30]
                for col_idx, width in enumerate(column_widths, start=1):
                    worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = width

                # Menyimpan workbook ke file
                workbook.save(file_path)

                # Add border to all cells
                # max_row = len(df.index) + 1
                # max_col = len(df.columns) + 1
                # for row in range(1, max_row):
                #     for col in range(1, max_col):
                #         cell = worksheet.cell(row=row, column=col)
                #         cell.border = openpyxl.styles.Border(
                #             left=openpyxl.styles.Side(border_style='thin', color='FF000000'),
                #             right=openpyxl.styles.Side(border_style='thin', color='FF000000'),
                #             top=openpyxl.styles.Side(border_style='thin', color='FF000000'),
                #             bottom=openpyxl.styles.Side(border_style='thin', color='FF000000')
                #         )

                # writer.save()
                # writer.close()

            return Response({
                'status': True,
                'file_path' : f'/{file_path}',
                'message': 'Export data successfully',
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': False,
                'message': 'Export data failed',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    # Membuat data profile ketika dibuatnya user
    def perform_create(self, serializer):
        satker_id = self.request.data.get('satker')
        notelp = self.request.data.get('notelp')
        direktorat = self.request.user.profile.role

        user_instance = serializer.save()

        signals.create_profile_for_user(sender=self.__class__, instance=user_instance, created=True, satker_id=satker_id, notelp=notelp, direktorat=direktorat)

    # Memperbarui data profile ketika diperbaruinya data user
    def perform_update(self, serializer):
        data = { "satker": self.request.data.get('satker'), "notelp": self.request.data.get('notelp') }

        profile_instance = serializer.instance.profile
        profile_serializer = serializers.UpdateProfileSerializer(profile_instance, data=data, partial=True)

        if not profile_serializer.is_valid():
            serializer.save()
        else:
            profile_serializer.save()
            serializer.save()

class SatkerViewSet(viewsets.ModelViewSet):
    queryset = models.Satker.objects.all()
    serializer_class = serializers.SatkerSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, ]
    filterset_class = filters.SatkerFilter

class reg_provincesViewSet(viewsets.ModelViewSet):
    queryset = models.reg_provinces.objects.all()
    serializer_class = serializers.reg_provincesSerializer
    permission_classes = [permissions.IsAuthenticated]

class reg_regenciesViewSet(viewsets.ModelViewSet):
    queryset = models.reg_regencies.objects.all()
    serializer_class = serializers.reg_regenciesSerializer
    permission_classes = [permissions.IsAuthenticated]

class reg_districtViewSet(viewsets.ModelViewSet):
    queryset = models.reg_district.objects.all()
    serializer_class = serializers.reg_districtSerializer
    permission_classes = [permissions.IsAuthenticated]

class reg_villagesViewSet(viewsets.ModelViewSet):
    queryset = models.reg_villages.objects.all()
    serializer_class = serializers.reg_villagesSerializer
    permission_classes = [permissions.IsAuthenticated]

class UraianKegiatanViewSet(viewsets.ModelViewSet):
    queryset = models.Uraian_kegiatan.objects.all()
    serializer_class = serializers.uraian_kegiatanSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, ]
    filterset_class = filters.UraianKegiatanFilter

    @action(detail=False, methods=['post'])
    def filters(self, request):
        id = request.data.get("id")
        uraian_kegiatan = self.queryset.filter(kegiatan_akun=id)
        serializer = self.get_serializer(uraian_kegiatan, many=True)
        return Response(serializer.data)

class KegiatanAkunViewSet(viewsets.ModelViewSet):
    queryset = models.Kegiatan_akun.objects.all()
    serializer_class = serializers.KegiatanAkunSerializer
    permission_classes = [permissions.IsAuthenticated]